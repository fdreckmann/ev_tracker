import os, json, time, sqlite3, logging, threading, requests, hashlib, secrets, functools, re
from typing import Optional
import smtplib, email
import xml.etree.ElementTree as ET
from datetime import datetime, timezone, timedelta
from pathlib import Path
from flask import Flask, render_template, request, jsonify, send_file, make_response, session, redirect, url_for
from providers import get_provider, get_all_capabilities, get_config_fields, PROVIDERS
from meter_providers import read_meter as _read_meter_impl, MeterResult

logging.basicConfig(level=logging.INFO, format="%(asctime)s [%(levelname)s] %(message)s")
log = logging.getLogger(__name__)

APP_VERSION   = "2.0.1"

CHANGELOG = [
    {"version":"2.0.0","changes":[
        "Export-Lokalisierung vollständig: alle Spaltenüberschriften und Labels auf Deutsch/Englisch",
        "Platzhalter-Ersetzung erweitert: {{total_sessions}}, {{total_km}}, {{avg_charge_power_kw}} u.v.m.",
        "Export-Vorschau: echte XLSX-Datei wird erzeugt und als Grid zurückgegeben",
        "Download-Token: Vorschau-XLSX per /api/export/download/<token> abrufbar (30 Min. gültig)",
        "Zählertest: Body-Parameter werden ohne Speichern für den Test verwendet",
        "Zählertest: Erweiterte Antwort mit provider, endpoint, raw_value, unit, normalized_from",
        "meter_value_unit Default auf 'auto' geändert",
        "to_row(): duration_hours aus Sekunden berechnet, charger_type aus Standort abgeleitet",
    ]},
    {"version":"1.9.9","changes":[
        "Passkey (WebAuthn/FIDO2) Authentifizierung: Anmelden per Fingerabdruck, Face ID oder Hardware-Key",
        "Passkey-Registrierung und -Verwaltung in den Sicherheitseinstellungen",
        "Mehrere Passkeys pro Benutzer möglich (mit eigenem Namen)",
        "Passkey-Login-Button auf der Anmeldeseite",
    ]},
    {"version":"1.9.3","changes":[
        "Automatische Template-Analyse mit Konfidenz-Score",
        "Zellmapping für Einzelzellen (Header/Footer-Bereich)",
        "Platzhalter-Erkennung ({{field_name}}) im Template",
        "Neuer Analyse-Button mit Vorschau und Übernahme-Funktion",
    ]},
    {"version":"1.9.0","changes":[
        "Passwort-Reset per E-Mail (Token-basiert)",
        "Benutzer-Einladungen per E-Mail",
        "Brute-force-Schutz: Account-Sperrung nach 5 Fehlversuchen",
        "2FA Backup-Codes (8 Einmalcodes pro Benutzer)",
        "Verbesserte Sicherheit: CSRF-Token in AJAX-Requests",
    ]},
    {"version":"1.8.0","changes":[
        "Vollständige Benutzerverwaltung (Admin + Benutzer-Rollen)",
        "Login mit E-Mail + Passwort (Multi-User)",
        "Setup-Assistent beim ersten Start",
        "Admin kann Benutzer anlegen, bearbeiten, deaktivieren, löschen",
        "2FA pro Benutzer (TOTP) mit Code-Bestätigung",
        "Eigenes Passwort und 2FA im Profil-Bereich änderbar",
        "Audit-Log mit Benutzer-Zuordnung",
    ]},
    {"version":"1.7.0","changes":[
        "Neuer Konfig-Bereich mit Sidebar-Navigation",
        "SMTP-Konfiguration mit Verbindungstest und Testmail",
        "Export-Vorlagen: mehrere Vorlagen speicherbar",
        "Audit-Log für sicherheitsrelevante Aktionen",
        "2FA-Status und Reset verbessert",
    ]},
    {"version": "1.6.1", "changes": [
        "SSO Login via Google-Konto (OAuth2)",
        "SSO Login via Microsoft-Konto / Azure AD (OAuth2)",
        "OAuth-Konfiguration + Redirect-URI-Anzeige im Sicherheits-Tab",
        "Login-Seite: Google- und Microsoft-Buttons werden automatisch angezeigt",
    ]},
    {"version": "1.6.0", "changes": [
        "Multi-Auto-Support: beliebig viele Fahrzeuge gleichzeitig tracken",
        "Jedes Fahrzeug läuft in eigenem Tracker-Thread mit eigenem Provider",
        "Fahrzeugverwaltung im Konfig-Tab (hinzufügen, bearbeiten, löschen)",
        "Ladevorgänge-Liste: Filter nach Fahrzeug, Fahrzeug-Spalte bei mehreren Autos",
        "Bestehende Sessions automatisch als Primärfahrzeug (v0) zugeordnet",
    ]},
    {"version": "1.5.0", "changes": [
        "Neue Provider: Hyundai / Kia (Bluelink/UVO), Renault / Dacia, Polestar, Audi (MyAudi)",
        "Login mit Passwort + optionaler 2FA (TOTP · Google Authenticator / Authy)",
        "Zählerstand-Quelle auf Konfig-Seite verschoben",
        "Update: Container-Erkennung per Image (funktioniert jetzt auf Unraid)",
        "Update: Remote-Versionsnummer und Changelog vor Installation sichtbar",
        "Update-Status: Live-Log wird im Browser angezeigt",
    ]},
    {"version": "1.4.6", "changes": [
        "Zählerstand im Dashboard als eigene Kachel (nur wenn Daten vorhanden)",
        "Zählerstand Alt→Neu in der Ladevorgänge-Liste (Spalte erscheint nur wenn Daten vorhanden)",
        "Zählerstand Alt/Neu im Session-Detail-Modal",
    ]},
    {"version": "1.4.5", "changes": [
        "Update-Mechanismus: HTTP/1.0 für Pull behoben (HTTP/1.1 Keep-Alive verhinderte Abschluss)",
        "Stop mit Force-Delete, besseres Logging für Fehlerdiagnose",
    ]},
    {"version": "1.4.4", "changes": [
        "Wallbox-Quellen: go-e Charger, openWB, WARP Charger, EVCC, Webasto, Alfen, Juice Charger",
        "EVCC als Universalquelle für KEBA, ABL, Mennekes, Heidelberg, Wallbe, NRGKick u.v.m.",
        "Erweiterte Konfiguration: EVCC-Port, Loadpoint-Index, Alfen-Passwort",
    ]},
    {"version": "1.4.3", "changes": [
        "Zählerstand-Quelle: Home Assistant, Shelly oder Tasmota",
        "Zählerstand wird beim Start und Ende jeder Session automatisch erfasst",
        "Fallback auf berechneten Zählerstand wenn keine Quelle konfiguriert",
    ]},
    {"version": "1.4.2", "changes": [
        "Zählerstand Alt/Neu wird automatisch aus Anfangszählerstand + kumulierten kWh berechnet",
        "Neues Konfigurationsfeld: Anfangszählerstand (kWh) im Export-Panel",
    ]},
    {"version": "1.4.1", "changes": [
        "Schrift in Dashboard-Kacheln skaliert automatisch — kein Überlauf mehr",
        "Schrift passt sich auch bei Fenstergrößenänderung neu an",
    ]},
    {"version": "1.4.0", "changes": [
        "Template-Kopfzeilen automatisch befüllen (Abrechnungsmonat, Kennzeichen, Fahrer, Abteilung, Kostenstelle, Gesamtkosten)",
        "Neue Template-Felder: Ladedauer, Kilometerstand, Ladekosten, Lademenge",
        "Konfigurationsfelder für Fahrer, Kennzeichen, Abteilung, Kostenstelle",
        "Kritischer Export-Fehler (NameError) behoben",
    ]},
    {"version": "1.3.0", "changes": [
        "Update-Tab als eigener Reiter neben Backup",
        "Changelog mit Versionshistorie im Update-Tab",
        "Update-Badge öffnet direkt Update-Tab und prüft automatisch",
        "Browser-Cache deaktiviert — kein manuelles Löschen mehr nötig",
    ]},
    {"version": "1.2.0", "changes": [
        "Excel-Template Editor mit visueller Vorschau",
        "Startzeile per Klick wählbar, Live-Vorschau mit echten Daten",
        "Template-Spalten-Zuordnung wird persistent gespeichert",
        "Export-Fehler werden als Toast angezeigt",
    ]},
    {"version": "1.1.0", "changes": [
        "Software-Update Kachel neben Backup",
        "Fortschrittsbalken beim Update-Install",
        "Dev-Channel und dev-Branch eingeführt",
        "Batteriekapazität als freies Zahlenfeld",
    ]},
]

DATA_DIR      = Path(os.environ.get("DATA_DIR", "/data"))
CONFIG_FILE   = DATA_DIR / "config.json"
DB_PATH       = DATA_DIR / "sessions.db"
EXPORT_DIR    = DATA_DIR / "exports"
TEMPLATE_PATH = DATA_DIR / "template.xlsx"
BACKUP_DIR    = DATA_DIR / "backups"
SIGNATURE_DIR  = DATA_DIR / "signatures"
SIGNATURE_PATH = SIGNATURE_DIR / "default_signature.png"

app = Flask(__name__)
app.config["MAX_CONTENT_LENGTH"] = 10 * 1024 * 1024

# ── Auth helpers ─────────────────────────────────────────────────────────────

def _get_secret_key():
    key_file = DATA_DIR / ".secret_key"
    if key_file.exists():
        return key_file.read_text().strip()
    key = secrets.token_hex(32)
    DATA_DIR.mkdir(parents=True, exist_ok=True)
    key_file.write_text(key)
    return key

def _hash_password(pw: str) -> str:
    return hashlib.sha256(pw.encode()).hexdigest()

def _password_ok(pw: str) -> "str | None":
    """Returns error string or None if password is acceptable."""
    if len(pw) < 8:
        return "Mindestens 8 Zeichen"
    if not any(c.isdigit() for c in pw):
        return "Mindestens eine Zahl erforderlich"
    return None

# ── User DB helpers ───────────────────────────────────────────────────────────
def _get_db():
    con = sqlite3.connect(DB_PATH)
    con.row_factory = sqlite3.Row
    return con

def _has_users() -> bool:
    try:
        con = _get_db()
        count = con.execute("SELECT COUNT(*) FROM users").fetchone()[0]
        con.close()
        return count > 0
    except Exception:
        return False

def _get_user_by_email(email: str):
    try:
        con = _get_db()
        row = con.execute("SELECT * FROM users WHERE email=?", (email.strip().lower(),)).fetchone()
        con.close()
        return dict(row) if row else None
    except Exception:
        return None

def _get_user_by_id(uid):
    try:
        con = _get_db()
        row = con.execute("SELECT * FROM users WHERE id=?", (uid,)).fetchone()
        con.close()
        return dict(row) if row else None
    except Exception:
        return None

def _current_user():
    uid = session.get("user_id")
    return _get_user_by_id(uid) if uid else None

def require_login(f):
    @functools.wraps(f)
    def wrapper(*args, **kwargs):
        if not session.get("user_id"):
            if request.path.startswith("/api/"):
                return jsonify({"error": "Nicht eingeloggt", "login_required": True}), 401
            return redirect(url_for("login_page", next=request.path))
        return f(*args, **kwargs)
    return wrapper

def require_admin(f):
    @functools.wraps(f)
    def wrapper(*args, **kwargs):
        if not session.get("user_id"):
            if request.path.startswith("/api/"):
                return jsonify({"error": "Nicht eingeloggt", "login_required": True}), 401
            return redirect(url_for("login_page", next=request.path))
        user = _get_user_by_id(session.get("user_id"))
        # Allow if legacy role is admin, OR if user has admin:all permission via roles table
        is_admin = (session.get("user_role") == "admin") or \
                   (user and ("admin:all" in _get_user_permissions(user["id"])))
        if not is_admin:
            if request.path.startswith("/api/"):
                return jsonify({"error": "Admin-Berechtigung erforderlich"}), 403
            return redirect(url_for("index"))
        return f(*args, **kwargs)
    return wrapper

def require_auth(f):
    """Legacy alias — kept for backward compat."""
    return require_login(f)

# ── Permission System ─────────────────────────────────────────────────────────

ALL_PERMISSIONS = {
    # Dashboard
    "dashboard:view":           {"label": "Dashboard ansehen",              "group": "Dashboard"},
    # Fahrzeuge
    "vehicles:view":            {"label": "Fahrzeuge ansehen",              "group": "Fahrzeuge"},
    "vehicles:create":          {"label": "Fahrzeug erstellen",             "group": "Fahrzeuge"},
    "vehicles:edit":            {"label": "Fahrzeug bearbeiten",            "group": "Fahrzeuge"},
    "vehicles:delete":          {"label": "Fahrzeug löschen",               "group": "Fahrzeuge"},
    "vehicles:switch":          {"label": "Fahrzeug wechseln",              "group": "Fahrzeuge"},
    "vehicles:image_manage":    {"label": "Fahrzeugbild verwalten",         "group": "Fahrzeuge"},
    "vehicles:provider_configure": {"label": "Provider konfigurieren",     "group": "Fahrzeuge"},
    # Ladevorgänge
    "sessions:view":            {"label": "Ladevorgänge ansehen",           "group": "Ladevorgänge"},
    "sessions:create":          {"label": "Ladevorgang erstellen",          "group": "Ladevorgänge"},
    "sessions:edit":            {"label": "Ladevorgang bearbeiten",         "group": "Ladevorgänge"},
    "sessions:delete":          {"label": "Ladevorgang löschen",            "group": "Ladevorgänge"},
    "sessions:manual_add":      {"label": "Manuell hinzufügen",             "group": "Ladevorgänge"},
    "sessions:validate":        {"label": "Ladevorgang validieren",         "group": "Ladevorgänge"},
    "sessions:ignore_warnings": {"label": "Warnungen ignorieren",           "group": "Ladevorgänge"},
    # Analyse
    "analytics:view":           {"label": "Analyse ansehen",                "group": "Analyse"},
    # Export
    "export:view":              {"label": "Export-Bereich ansehen",         "group": "Export"},
    "export:create":            {"label": "Export erstellen",               "group": "Export"},
    "export:preview":           {"label": "Export-Vorschau",                "group": "Export"},
    "export:download":          {"label": "Export herunterladen",           "group": "Export"},
    "export:templates_view":    {"label": "Templates ansehen",              "group": "Export"},
    "export:templates_manage":  {"label": "Templates verwalten",            "group": "Export"},
    "export:signature_use":     {"label": "Signatur im Export nutzen",      "group": "Export"},
    # Templates
    "templates:view":           {"label": "Templates ansehen",              "group": "Templates"},
    "templates:upload":         {"label": "Template hochladen",             "group": "Templates"},
    "templates:edit_mapping":   {"label": "Template-Mapping bearbeiten",    "group": "Templates"},
    "templates:delete":         {"label": "Template löschen",               "group": "Templates"},
    "templates:gallery_use":    {"label": "Template-Galerie nutzen",        "group": "Templates"},
    # Signatur
    "signature:view":           {"label": "Signatur ansehen",               "group": "Signatur"},
    "signature:upload":         {"label": "Signatur hochladen",             "group": "Signatur"},
    "signature:draw":           {"label": "Signatur zeichnen",              "group": "Signatur"},
    "signature:delete":         {"label": "Signatur löschen",               "group": "Signatur"},
    "signature:use_in_export":  {"label": "Signatur im Export verwenden",   "group": "Signatur"},
    # Zählerstand
    "meter:view":               {"label": "Zählerstand ansehen",            "group": "Zählerstand"},
    "meter:test":               {"label": "Zählerstand testen",             "group": "Zählerstand"},
    "meter:configure":          {"label": "Zählerstand konfigurieren",      "group": "Zählerstand"},
    # Provider
    "providers:view":           {"label": "Provider ansehen",               "group": "Provider"},
    "providers:configure":      {"label": "Provider konfigurieren",         "group": "Provider"},
    "providers:test":           {"label": "Provider testen",                "group": "Provider"},
    # Einstellungen
    "settings:view":            {"label": "Einstellungen ansehen",          "group": "Einstellungen"},
    "settings:edit":            {"label": "Einstellungen bearbeiten",       "group": "Einstellungen"},
    # Benutzer
    "users:view":               {"label": "Benutzer ansehen",               "group": "Benutzer"},
    "users:create":             {"label": "Benutzer erstellen",             "group": "Benutzer"},
    "users:edit":               {"label": "Benutzer bearbeiten",            "group": "Benutzer"},
    "users:delete":             {"label": "Benutzer löschen",               "group": "Benutzer"},
    "users:reset_password":     {"label": "Passwort zurücksetzen",          "group": "Benutzer"},
    "users:manage_2fa":         {"label": "2FA verwalten",                  "group": "Benutzer"},
    "users:manage_permissions": {"label": "Rollen & Rechte verwalten",      "group": "Benutzer"},
    # Backup
    "backup:view":              {"label": "Backup ansehen",                 "group": "Backup"},
    "backup:create":            {"label": "Backup erstellen",               "group": "Backup"},
    "backup:download":          {"label": "Backup herunterladen",           "group": "Backup"},
    "backup:restore":           {"label": "Backup wiederherstellen",        "group": "Backup"},
    "backup:delete":            {"label": "Backup löschen",                 "group": "Backup"},
    # Updates
    "updates:view":             {"label": "Updates ansehen",                "group": "Updates"},
    "updates:check":            {"label": "Updates prüfen",                 "group": "Updates"},
    "updates:start":            {"label": "Update starten",                 "group": "Updates"},
    "updates:history":          {"label": "Update-Historie ansehen",        "group": "Updates"},
    # Audit/Sicherheit
    "audit:view":               {"label": "Audit-Log ansehen",              "group": "Sicherheit"},
    "security:view":            {"label": "Sicherheit ansehen",             "group": "Sicherheit"},
    "api_tokens:manage":        {"label": "API-Tokens verwalten",           "group": "Sicherheit"},
    # System
    "system:status":            {"label": "Systemstatus ansehen",           "group": "System"},
    "system:logs":              {"label": "Logs ansehen",                   "group": "System"},
    "system:health":            {"label": "Health-Check",                   "group": "System"},
    # Reports
    "reports:view":             {"label": "Reports ansehen",                "group": "Reports"},
    "reports:configure":        {"label": "Reports konfigurieren",          "group": "Reports"},
    "reports:send":             {"label": "Report senden",                  "group": "Reports"},
    "reports:history":          {"label": "Report-Historie ansehen",        "group": "Reports"},
    # Admin-Sonderrecht
    "admin:all":                {"label": "Vollzugriff (Admin)",            "group": "Admin"},
}

DEFAULT_ROLE_PERMISSIONS = {
    "admin": ["admin:all"],
    "user": [
        "dashboard:view", "vehicles:view", "vehicles:switch",
        "sessions:view", "sessions:create", "sessions:edit", "sessions:manual_add",
        "analytics:view",
        "export:view", "export:create", "export:preview", "export:download",
        "export:templates_view", "export:signature_use",
        "templates:view", "templates:gallery_use",
        "signature:view", "signature:upload", "signature:draw",
        "signature:delete", "signature:use_in_export",
        "meter:view", "meter:test",
        "providers:view",
        "settings:view",
        "reports:view", "reports:history",
    ],
    "readonly": [
        "dashboard:view", "vehicles:view", "sessions:view",
        "analytics:view", "export:view", "export:preview", "export:download",
    ],
}

DEFAULT_CONFIG = {
    # Provider selection
    "provider":             "ha",        # ha | vw | tesla | volvo | bmw | mercedes
    "car_name":             "Mein EV",

    # HA provider fields
    "ha_url":               "http://homeassistant.local:8123",
    "ha_token":             "",
    "charging_sensor":      "",
    "soc_sensor":           "",
    "odo_sensor":           "",
    "power_sensor":         "",
    "charge_speed_sensor":  "",
    "charge_type_sensor":   "",
    "location_sensor":      "",
    "home_states":          "home,zuhause",
    "dc_threshold_kw":      22.0,

    # VW provider fields
    "vw_username":          "",
    "vw_password":          "",
    "vw_vin":               "",

    # Tesla provider fields
    "tesla_email":          "",
    "tesla_vin":            "",

    # Volvo provider fields
    "volvo_api_key":        "",
    "volvo_access_token":   "",
    "volvo_vin":            "",

    # BMW provider fields
    "bmw_username":         "",
    "bmw_password":         "",
    "bmw_vin":              "",
    "bmw_region":           "rest_of_world",

    # Mercedes provider fields
    "mercedes_token":       "",
    "mercedes_vin":         "",

    # Shared location fields (non-HA providers)
    "home_lat":             "",
    "home_lon":             "",
    "home_radius_m":        200,

    # Pricing
    "battery_capacity_kwh": 77.0,
    "price_per_kwh_home":   0.30,
    "price_per_kwh_ac":     0.45,
    "price_per_kwh_dc":     0.75,
    "entsoe_api_key":       "",
    "entsoe_ac_markup":     3.0,
    "entsoe_dc_markup":     6.0,

    # Notifications
    "notify_service":       "",

    # System
    "poll_interval":        60,
    "backup_cron":          "",
    "update_channel":       "latest",  # latest | nightly | dev
    "template_mapping":     {},        # {col_index_str: field_name}
    "template_start_row":   None,      # row number where data starts (1-based)
    "template_header_row":  None,      # row number of column headers (1-based)
    "export_language":      "de",      # export language: de|en
    "active_template":      {"source": None, "template_id": None, "name": None},
    "template_fahrer":        "",      # driver name for template header
    "template_kennzeichen":   "",      # license plate for template header
    "template_abteilung":     "",      # department for template header
    "template_kostenstelle":  "",      # cost center for template header
    "template_meter_start":   0.0,    # starting electricity meter reading (kWh) fallback
    "meter_source":      "none",  # none|ha|shelly|tasmota|go_e|openwb|warp|evcc|webasto|alfen|juice
    "meter_sensor":      "",      # HA entity_id
    "meter_device_ip":   "",      # IP for all device-based sources
    "meter_evcc_port":   7070,    # EVCC port (default 7070)
    "meter_evcc_lp":     0,       # EVCC loadpoint index (0-based)
    "meter_alfen_pass":  "admin", # Alfen web UI password (deprecated, use meter_password)
    "meter_device_scheme": "http",  # http|https
    "meter_device_port": "",        # optional port override
    "meter_username":    "",        # username for auth (Shelly, Tasmota, Alfen, Generic)
    "meter_password":    "",        # password for auth
    "meter_channel":     0,         # Shelly channel/emeter index
    "meter_phase_mode":  "total",   # Shelly: total|a|b|c
    "meter_json_path":   "",        # Tasmota/Generic: dot-separated JSON path
    "meter_value_unit":  "auto",    # Generic: auto|kwh|wh|mwh
    "meter_value_factor": 1.0,      # Generic: multiply factor
    "meter_generic_url": "",        # Generic HTTP: full URL
    "meter_openwb_lp":   1,         # openWB loadpoint (1-based)
    "meter_warp_meter_index": 0,    # WARP meter index
    "meter_timeout_seconds": 8,     # request timeout
    "meter_verify_ssl":  True,      # verify SSL certificate
    "meter_prefer_meter_delta": False,  # use meter delta instead of SoC for kwh_charged

    # Auth — password + TOTP
    "auth_password_hash": "",
    "auth_totp_secret":   "",

    # OAuth2 SSO — Google
    "oauth_google_client_id":     "",
    "oauth_google_client_secret": "",

    # OAuth2 SSO — Microsoft
    "oauth_microsoft_client_id":     "",
    "oauth_microsoft_client_secret": "",
    "oauth_microsoft_tenant":        "common",  # "common" = any MS account, or tenant-id

    # Base URL for OAuth redirect URIs (e.g. https://ev.example.com)
    # Leave empty to auto-detect from request
    "oauth_base_url": "",

    # SMTP
    "smtp_host":      "",
    "smtp_port":      587,
    "smtp_tls":       "starttls",   # starttls | ssl | none
    "smtp_user":      "",
    "smtp_password":  "",
    "smtp_from_name": "EV Tracker",
    "smtp_from_email":"",
    "smtp_reply_to":  "",

    # Export templates (stored as list in config)
    "export_templates": [],  # [{id, name, mapping, start_row, is_default}]

    # Signature
    "signature":        {"source": None, "created_at": None},
    "signature_mapping": {},          # {"cell":"B42","width":220,"height":80,"offset_x":0,"offset_y":0}
    "export_include_signature": False,
    "signature_padding_px": 24,

    # Email Reports
    "report_email_enabled":          False,
    "report_email_recipients":       [],
    "report_email_schedule_type":    "monthly",
    "report_email_time":             "08:00",
    "report_email_weekday":          1,
    "report_email_day_of_month":     1,
    "report_email_month":            1,
    "report_email_custom_days":      14,
    "report_email_cron":             "",
    "report_email_period_mode":      "previous_period",
    "report_email_custom_start_date": "",
    "report_email_custom_end_date":   "",
    "report_email_location_filter":  "all",
    "report_email_vehicle_filter":   "all",
    "report_email_include_excel":    True,
    "report_email_include_summary":  True,
    "report_email_language":         "auto",
    "report_email_include_signature": False,
    "report_email_template_id":      None,
    "report_email_last_sent_key":    "",

    # Multi-vehicle: additional vehicles beyond the primary (v0)
    "extra_vehicles":    [],
}

# Fields that belong to a vehicle config (vs. app-level config)
VEHICLE_SPECIFIC_KEYS = {
    "provider","car_name","poll_interval","battery_capacity_kwh",
    "home_lat","home_lon","home_radius_m","dc_threshold_kw",
    "ha_url","ha_token","charging_sensor","soc_sensor","odo_sensor",
    "power_sensor","charge_speed_sensor","charge_type_sensor","location_sensor","home_states",
    "vw_username","vw_password","vw_vin","vw_update_interval",
    "tesla_email","tesla_vin",
    "volvo_api_key","volvo_access_token","volvo_vin",
    "bmw_username","bmw_password","bmw_vin","bmw_region",
    "mercedes_token","mercedes_vin",
    "hk_brand","hk_username","hk_password","hk_pin","hk_region","hk_vin",
    "renault_username","renault_password","renault_locale","renault_account","renault_vin",
    "polestar_username","polestar_password","polestar_vin",
    "audi_username","audi_password","audi_vin",
    # Stellantis
    "stellantis_brand","stellantis_username","stellantis_password","stellantis_vin",
    # Ford
    "ford_username","ford_password","ford_vin",
    # MG / SAIC
    "mg_username","mg_password","mg_vin","mg_region",
    # Toyota / Lexus
    "toyota_username","toyota_password","toyota_vin","toyota_locale","toyota_region",
    # Nissan
    "nissan_username","nissan_password","nissan_vin","nissan_region",
    # Porsche
    "porsche_username","porsche_password","porsche_vin",
    # JLR
    "jlr_username","jlr_password","jlr_vin",
    # XPeng (stub)
    # BYD (stub)
    # Aggregators
    "tronity_client_id","tronity_client_secret","tronity_vehicle_id",
    "enode_client_id","enode_client_secret","enode_user_id","enode_vehicle_id",
    "smartcar_access_token","smartcar_vehicle_id",
}

def load_config():
    if CONFIG_FILE.exists():
        with open(CONFIG_FILE) as f:
            return {**DEFAULT_CONFIG, **json.load(f)}
    return DEFAULT_CONFIG.copy()

def save_config(cfg):
    DATA_DIR.mkdir(parents=True, exist_ok=True)
    with open(CONFIG_FILE, "w") as f:
        json.dump(cfg, f, indent=2)

def _audit(action: str, details: str = "", ip: str = ""):
    uid = session.get("user_id") if session else None
    try:
        con = sqlite3.connect(DB_PATH)
        con.execute(
            "INSERT INTO audit_log (ts, action, details, ip, user_id) VALUES (?,?,?,?,?)",
            (datetime.utcnow().isoformat(), action, details.strip(), ip, uid))
        con.commit(); con.close()
    except Exception:
        pass

def get_all_vehicles(cfg=None) -> list[dict]:
    """Returns primary vehicle (v0 from flat config) plus extra vehicles."""
    if cfg is None:
        cfg = load_config()
    primary = {
        "id":   "v0",
        "name": cfg.get("car_name", "Mein EV"),
        "provider": cfg.get("provider", "ha"),
        "active": True,
        **{k: cfg[k] for k in VEHICLE_SPECIFIC_KEYS if k in cfg and k not in ("provider","car_name")},
    }
    extras = cfg.get("extra_vehicles", [])
    return [primary] + extras

def build_vehicle_config(vehicle: dict, cfg=None) -> dict:
    """Merge app-level config with vehicle-specific fields for provider initialization."""
    if cfg is None:
        cfg = load_config()
    merged = dict(cfg)
    merged.update(vehicle)
    if "name" in vehicle:
        merged["car_name"] = vehicle["name"]
    return merged

# ── Database ──────────────────────────────────────────────────────────────────
def init_db():
    DATA_DIR.mkdir(parents=True, exist_ok=True)
    con = sqlite3.connect(DB_PATH)
    con.row_factory = sqlite3.Row
    con.execute("""CREATE TABLE IF NOT EXISTS sessions (
        id            INTEGER PRIMARY KEY AUTOINCREMENT,
        start_ts      TEXT, end_ts TEXT,
        odo_start     REAL, odo_end REAL,
        soc_start     REAL, soc_end REAL,
        kwh_charged   REAL, cost_eur REAL,
        cost_manual   INTEGER DEFAULT 0,
        location      TEXT DEFAULT 'unknown',
        charger_type  TEXT DEFAULT 'unknown',
        max_power_kw  REAL,
        price_per_kwh REAL,
        entsoe_spot   REAL,
        provider      TEXT DEFAULT 'ha',
        meter_old     REAL, meter_new REAL,
        vehicle_id    TEXT DEFAULT 'v0'
    )""")
    # migrate existing DB
    for col, typedef in [("meter_old","REAL"),("meter_new","REAL"),("vehicle_id","TEXT DEFAULT 'v0'")]:
        try: con.execute(f"ALTER TABLE sessions ADD COLUMN {col} {typedef}")
        except Exception: pass
    con.execute("""CREATE TABLE IF NOT EXISTS session_points (
        id INTEGER PRIMARY KEY AUTOINCREMENT,
        session_id INTEGER NOT NULL,
        ts TEXT NOT NULL,
        soc REAL, power_kw REAL
    )""")
    con.execute("""CREATE TABLE IF NOT EXISTS audit_log (
    id      INTEGER PRIMARY KEY AUTOINCREMENT,
    ts      TEXT NOT NULL,
    action  TEXT NOT NULL,
    details TEXT,
    ip      TEXT
)""")
    # live migration
    for col, typedef in [
        ("cost_manual",  "INTEGER DEFAULT 0"),
        ("charger_type", "TEXT DEFAULT 'unknown'"),
        ("max_power_kw", "REAL"),
        ("price_per_kwh","REAL"),
        ("entsoe_spot",  "REAL"),
        ("provider",     "TEXT DEFAULT 'ha'"),
        ("kwh_source",      "TEXT DEFAULT 'soc'"),
        ("meter_delta_kwh", "REAL"),
        ("meter_error",     "TEXT"),
    ]:
        try:
            con.execute(f"ALTER TABLE sessions ADD COLUMN {col} {typedef}")
        except sqlite3.OperationalError: pass
    con.execute("""CREATE TABLE IF NOT EXISTS users (
        id            INTEGER PRIMARY KEY AUTOINCREMENT,
        name          TEXT NOT NULL,
        email         TEXT UNIQUE NOT NULL,
        password_hash TEXT NOT NULL DEFAULT '',
        role          TEXT NOT NULL DEFAULT 'user',
        status        TEXT NOT NULL DEFAULT 'active',
        totp_secret   TEXT NOT NULL DEFAULT '',
        totp_enabled  INTEGER NOT NULL DEFAULT 0,
        created_at    TEXT NOT NULL,
        updated_at    TEXT NOT NULL,
        last_login_at TEXT
    )""")
    con.execute("""CREATE TABLE IF NOT EXISTS password_reset_tokens (
        id         INTEGER PRIMARY KEY AUTOINCREMENT,
        user_id    INTEGER NOT NULL,
        token_hash TEXT NOT NULL,
        expires_at TEXT NOT NULL,
        used_at    TEXT,
        created_at TEXT NOT NULL
    )""")
    con.execute("""CREATE TABLE IF NOT EXISTS invite_tokens (
        id         INTEGER PRIMARY KEY AUTOINCREMENT,
        user_id    INTEGER NOT NULL,
        token_hash TEXT NOT NULL,
        expires_at TEXT NOT NULL,
        used_at    TEXT,
        created_at TEXT NOT NULL
    )""")
    con.execute("""CREATE TABLE IF NOT EXISTS totp_backup_codes (
        id       INTEGER PRIMARY KEY AUTOINCREMENT,
        user_id  INTEGER NOT NULL,
        code_hash TEXT NOT NULL,
        used_at  TEXT
    )""")
    # live migration: add brute-force columns to users
    for col, typedef in [
        ("failed_attempts", "INTEGER NOT NULL DEFAULT 0"),
        ("locked_until",    "TEXT"),
    ]:
        try: con.execute(f"ALTER TABLE users ADD COLUMN {col} {typedef}")
        except Exception: pass
    # live migration: add user_id column to audit_log
    try: con.execute("ALTER TABLE audit_log ADD COLUMN user_id INTEGER")
    except Exception: pass
    con.execute("""CREATE TABLE IF NOT EXISTS webauthn_credentials (
        id            INTEGER PRIMARY KEY AUTOINCREMENT,
        user_id       INTEGER NOT NULL,
        credential_id TEXT NOT NULL UNIQUE,
        public_key    TEXT NOT NULL,
        sign_count    INTEGER NOT NULL DEFAULT 0,
        name          TEXT NOT NULL DEFAULT 'Passkey',
        created_at    TEXT NOT NULL,
        last_used_at  TEXT
    )""")
    # Roles
    con.execute("""CREATE TABLE IF NOT EXISTS roles (
        id          INTEGER PRIMARY KEY AUTOINCREMENT,
        name        TEXT UNIQUE NOT NULL,
        description TEXT DEFAULT '',
        is_system   INTEGER DEFAULT 0,
        created_at  TEXT,
        updated_at  TEXT
    )""")
    # Role → Permission mapping
    con.execute("""CREATE TABLE IF NOT EXISTS role_permissions (
        role_id         INTEGER NOT NULL,
        permission_key  TEXT NOT NULL,
        PRIMARY KEY (role_id, permission_key),
        FOREIGN KEY (role_id) REFERENCES roles(id) ON DELETE CASCADE
    )""")
    # User → Role mapping (mehrere Rollen pro User möglich)
    con.execute("""CREATE TABLE IF NOT EXISTS user_roles (
        user_id INTEGER NOT NULL,
        role_id INTEGER NOT NULL,
        PRIMARY KEY (user_id, role_id),
        FOREIGN KEY (user_id) REFERENCES users(id) ON DELETE CASCADE,
        FOREIGN KEY (role_id) REFERENCES roles(id) ON DELETE CASCADE
    )""")
    # Vehicle access scope (Vorbereitung, noch nicht aktiv)
    con.execute("""CREATE TABLE IF NOT EXISTS user_vehicle_access (
        user_id    INTEGER NOT NULL,
        vehicle_id INTEGER NOT NULL,
        can_view   INTEGER DEFAULT 1,
        can_edit   INTEGER DEFAULT 0,
        can_export INTEGER DEFAULT 0,
        PRIMARY KEY (user_id, vehicle_id)
    )""")
    con.execute("""CREATE TABLE IF NOT EXISTS email_report_history (
        id            INTEGER PRIMARY KEY AUTOINCREMENT,
        sent_at       TEXT NOT NULL,
        schedule_type TEXT,
        period_start  TEXT,
        period_end    TEXT,
        period_key    TEXT,
        location_filter TEXT DEFAULT 'all',
        vehicle_filter  TEXT DEFAULT 'all',
        recipients    TEXT,
        status        TEXT NOT NULL DEFAULT 'sent',
        error         TEXT,
        triggered_by  TEXT DEFAULT 'auto'
    )""")
    # Seed default roles
    now_iso = datetime.utcnow().isoformat()
    default_roles = [
        ("admin",    "Vollzugriff",                    1),
        ("user",     "Normaler Benutzer",               1),
        ("readonly", "Nur-Lese-Zugriff",               1),
    ]
    for name, desc, is_sys in default_roles:
        existing = con.execute("SELECT id FROM roles WHERE name=?", (name,)).fetchone()
        if not existing:
            con.execute(
                "INSERT INTO roles (name, description, is_system, created_at, updated_at) VALUES (?,?,?,?,?)",
                (name, desc, is_sys, now_iso, now_iso)
            )
            con.commit()
    # Seed default role permissions
    for role_name, perms in DEFAULT_ROLE_PERMISSIONS.items():
        role_row = con.execute("SELECT id FROM roles WHERE name=?", (role_name,)).fetchone()
        if role_row:
            role_id = role_row["id"]
            existing_perms = {r["permission_key"] for r in
                con.execute("SELECT permission_key FROM role_permissions WHERE role_id=?", (role_id,)).fetchall()}
            for pkey in perms:
                if pkey not in existing_perms and pkey in ALL_PERMISSIONS:
                    con.execute("INSERT OR IGNORE INTO role_permissions (role_id, permission_key) VALUES (?,?)",
                                (role_id, pkey))
            con.commit()
    # Migrate existing users to roles
    users_without_roles = con.execute("""
        SELECT u.id, u.role FROM users u
        WHERE NOT EXISTS (SELECT 1 FROM user_roles ur WHERE ur.user_id = u.id)
    """).fetchall()
    for u in users_without_roles:
        role_name = u["role"] if u["role"] in ("admin", "user", "readonly") else "user"
        role_row = con.execute("SELECT id FROM roles WHERE name=?", (role_name,)).fetchone()
        if role_row:
            con.execute("INSERT OR IGNORE INTO user_roles (user_id, role_id) VALUES (?,?)",
                        (u["id"], role_row["id"]))
    con.commit()
    con.close()

# ── Permission checking ───────────────────────────────────────────────────────

def _get_user_permissions(user_id: int) -> set:
    """Return set of permission keys for a user (via all assigned roles)."""
    con = _get_db()
    rows = con.execute("""
        SELECT DISTINCT rp.permission_key
        FROM user_roles ur
        JOIN role_permissions rp ON ur.role_id = rp.role_id
        WHERE ur.user_id = ?
    """, (user_id,)).fetchall()
    con.close()
    return {r["permission_key"] for r in rows}

def has_permission(user, permission_key: str) -> bool:
    """Check if user has a specific permission."""
    if not user:
        return False
    user_id = user["id"] if isinstance(user, dict) else int(user)
    perms = _get_user_permissions(user_id)
    return "admin:all" in perms or permission_key in perms

def require_permission(permission_key: str):
    """Decorator: require a specific permission, return 403 if missing."""
    def decorator(f):
        @functools.wraps(f)
        def wrapper(*args, **kwargs):
            user = _current_user()
            if not user:
                return jsonify({"ok": False, "error": "Nicht angemeldet"}), 401
            if not has_permission(user, permission_key):
                _audit("permission_denied",
                       f"perm={permission_key} endpoint={request.path}",
                       ip=request.remote_addr)
                return jsonify({"ok": False, "error": f"Keine Berechtigung: {permission_key}"}), 403
            return f(*args, **kwargs)
        return wrapper
    return decorator

def get_sessions(year=None, month=None, location=None, vehicle_id=None, limit=50):
    where = ["end_ts IS NOT NULL"]; params = []
    if year and month:
        where.append("start_ts LIKE ?"); params.append(f"{year:04d}-{month:02d}%")
    if location and location != "all":
        where.append("location = ?"); params.append(location)
    if vehicle_id and vehicle_id != "all":
        where.append("vehicle_id = ?"); params.append(vehicle_id)
    sql = f"SELECT * FROM sessions WHERE {' AND '.join(where)} ORDER BY start_ts DESC"
    if not (year and month): sql += f" LIMIT {limit}"
    con = sqlite3.connect(DB_PATH); con.row_factory = sqlite3.Row
    rows = con.execute(sql, params).fetchall(); con.close()
    return [dict(r) for r in rows]

def get_monthly_stats():
    con = sqlite3.connect(DB_PATH); con.row_factory = sqlite3.Row
    rows = con.execute("""
        SELECT strftime('%Y-%m', start_ts) AS month,
               COUNT(*) AS sessions,
               SUM(kwh_charged) AS total_kwh,
               SUM(cost_eur) AS total_cost,
               MAX(odo_end) - MIN(odo_start) AS km_driven,
               SUM(CASE WHEN location='home'   THEN cost_eur ELSE 0 END) AS home_cost,
               SUM(CASE WHEN location='extern' THEN cost_eur ELSE 0 END) AS ext_cost,
               SUM(CASE WHEN charger_type='dc' THEN kwh_charged ELSE 0 END) AS dc_kwh,
               SUM(CASE WHEN charger_type='ac' THEN kwh_charged ELSE 0 END) AS ac_kwh
        FROM sessions WHERE end_ts IS NOT NULL
        GROUP BY month ORDER BY month DESC LIMIT 12
    """).fetchall()
    con.close(); return [dict(r) for r in rows]

# ── ENTSO-E ───────────────────────────────────────────────────────────────────
_entsoe_cache = {"price": None, "ts": 0}
_forgot_pw_attempts: dict[str, list] = {}  # email -> [timestamp, ...]

def fetch_entsoe_spot(api_key: str):
    if not api_key: return None
    if _entsoe_cache["price"] is not None and time.time() - _entsoe_cache["ts"] < 3600:
        return _entsoe_cache["price"]
    try:
        now       = datetime.now(timezone.utc)
        day_start = now.replace(hour=0,minute=0,second=0,microsecond=0)
        day_end   = day_start + timedelta(days=1)
        url = (
            "https://web-api.tp.entsoe.eu/api"
            f"?securityToken={api_key}"
            "&documentType=A44"
            "&in_Domain=10Y1001A1001A83F"
            "&out_Domain=10Y1001A1001A83F"
            f"&periodStart={day_start.strftime('%Y%m%d%H%M')}"
            f"&periodEnd={day_end.strftime('%Y%m%d%H%M')}"
        )
        r    = requests.get(url, timeout=15); r.raise_for_status()
        root = ET.fromstring(r.text)
        ns   = "urn:iec62325.351:tc57wg16:451-3:publicationdocument:7:3"
        pts  = root.findall(f".//{{{ns}}}Point")
        if not pts: return None
        current_hour = now.hour
        pt = pts[min(current_hour, len(pts)-1)]
        price_el = pt.find(f"{{{ns}}}price.amount")
        if price_el is None: return None
        price_kwh = round(float(price_el.text) / 1000, 4)
        _entsoe_cache.update(price=price_kwh, ts=time.time())
        log.info("ENTSO-E: %.4f €/kWh", price_kwh)
        return price_kwh
    except Exception as e:
        log.warning("ENTSO-E error: %s", e); return None

def calc_extern_price(cfg, charger_type, spot):
    is_dc = charger_type == "dc"
    if spot is not None:
        markup = cfg.get("entsoe_dc_markup" if is_dc else "entsoe_ac_markup", 6.0 if is_dc else 3.0)
        return round(spot * markup, 4)
    return cfg.get("price_per_kwh_dc" if is_dc else "price_per_kwh_ac", 0.75 if is_dc else 0.45)

def ha_notify(cfg, title, message):
    service = cfg.get("notify_service","").strip()
    if not service or cfg.get("provider") != "ha": return
    try:
        parts = service.split(".",1)
        svc   = parts[1] if len(parts)==2 else parts[0]
        requests.post(
            f"{cfg['ha_url'].rstrip('/')}/api/services/notify/{svc}",
            headers={"Authorization":f"Bearer {cfg['ha_token']}","Content-Type":"application/json"},
            json={"title":title,"message":message}, timeout=8)
    except Exception as e:
        log.warning("Notify error: %s", e)

# ── Tracker ───────────────────────────────────────────────────────────────────
def _make_state(vehicle_id="v0", provider_id="ha"):
    return {
        "vehicle_id": vehicle_id,
        "running": False, "session_active": False, "session_id": None,
        "last_poll": None, "last_error": None, "soc_current": None,
        "odo_current": None, "charging": False, "location": "unknown",
        "charger_type": "unknown", "power_kw": None, "entsoe_spot": None,
        "provider": provider_id,
        "provider_name": PROVIDERS.get(provider_id, PROVIDERS["ha"]).PROVIDER_NAME,
    }

_vehicle_states: dict[str, dict] = {"v0": _make_state("v0")}
_vehicle_stops:  dict[str, threading.Event] = {"v0": threading.Event()}

# Backward compat alias
_state = _vehicle_states["v0"]
_stop  = _vehicle_stops["v0"]

def read_meter_value() -> Optional[float]:
    """Read current meter value. Returns kWh or None."""
    cfg = load_config()
    result = _read_meter_impl(cfg)
    if result.ok:
        return result.value
    if result.error:
        log.warning("Meter read error (%s): %s", result.source, result.error)
    return None

def tracker_loop(vehicle_id: str = "v0"):
    st   = _vehicle_states[vehicle_id]
    stop = _vehicle_stops[vehicle_id]
    st["running"] = True
    session_active = False; session_id = None
    soc_start = odo_start = peak_power = None

    log.info("Tracker gestartet: %s", vehicle_id)
    while not stop.is_set():
        cfg = load_config()
        # For v0 use flat config; for extra vehicles get their config merged with app config
        if vehicle_id == "v0":
            vcfg = cfg
        else:
            extras = cfg.get("extra_vehicles", [])
            vehicle = next((v for v in extras if v["id"] == vehicle_id), None)
            if not vehicle:
                log.warning("Fahrzeug %s nicht gefunden — Tracker stoppt", vehicle_id)
                break
            vcfg = build_vehicle_config(vehicle, cfg)
        provider_id = vcfg.get("provider", "ha")

        try:
            provider = get_provider(provider_id, vcfg)
            state    = provider.get_state()

            if state.error:
                st["last_error"] = state.error
                stop.wait(vcfg.get("poll_interval", 60)); continue

            charging     = state.charging or False
            soc          = state.soc
            odo          = state.odometer
            power_kw     = state.charge_power
            location     = state.location or "unknown"
            charger_type = state.charge_type or "unknown"

            st.update(
                charging=charging, soc_current=soc, odo_current=odo,
                location=location, charger_type=charger_type, power_kw=power_kw,
                session_active=session_active,
                last_poll=datetime.now().isoformat(timespec="seconds"),
                last_error=None, provider=provider_id,
                provider_name=PROVIDERS.get(provider_id, PROVIDERS["ha"]).PROVIDER_NAME,
            )

            con = sqlite3.connect(DB_PATH); cur = con.cursor()

            if charging and not session_active:
                soc_start = soc; odo_start = odo; peak_power = power_kw or 0
                _meter_start_res = _read_meter_impl(cfg)
                meter_start_val = _meter_start_res.value
                spot = fetch_entsoe_spot(cfg.get("entsoe_api_key","")) if location=="extern" else None
                st["entsoe_spot"] = spot
                price_kwh = (cfg["price_per_kwh_home"] if location=="home"
                             else calc_extern_price(cfg, charger_type, spot))
                cur.execute("""INSERT INTO sessions
                    (start_ts,odo_start,soc_start,location,charger_type,
                     max_power_kw,price_per_kwh,entsoe_spot,provider,meter_old,vehicle_id)
                    VALUES (?,?,?,?,?,?,?,?,?,?,?)""",
                    (datetime.now().isoformat(timespec="seconds"),
                     odo_start,soc_start,location,charger_type,power_kw,price_kwh,spot,
                     provider_id,meter_start_val,vehicle_id))
                con.commit(); session_id=cur.lastrowid; session_active=True
                st["session_id"]=session_id
                cur.execute("INSERT INTO session_points (session_id,ts,soc,power_kw) VALUES (?,?,?,?)",
                            (session_id,datetime.now().isoformat(timespec="seconds"),soc_start,power_kw))
                con.commit()
                log.info("⚡ [%s] Session #%d | %s | %s | %.2f €/kWh",
                         vehicle_id,session_id,location.upper(),charger_type.upper(),price_kwh)
                ha_notify(vcfg,f"⚡ {vcfg['car_name']} lädt",
                    f"{'🏠 Zuhause' if location=='home' else '⚡ Extern'} · "
                    f"{'DC' if charger_type=='dc' else 'AC'} · {price_kwh:.2f} €/kWh · SOC {soc_start or '?'}%")

            elif charging and session_active:
                cur.execute("INSERT INTO session_points (session_id,ts,soc,power_kw) VALUES (?,?,?,?)",
                            (session_id,datetime.now().isoformat(timespec="seconds"),soc,power_kw))
                con.commit()
                if power_kw and (peak_power is None or power_kw > peak_power):
                    peak_power = power_kw
                    new_type = "dc" if power_kw > float(vcfg.get("dc_threshold_kw",22)) else "ac"
                    if new_type != charger_type:
                        spot = st.get("entsoe_spot")
                        price = (cfg["price_per_kwh_home"] if location=="home"
                                 else calc_extern_price(cfg,new_type,spot))
                        cur.execute("UPDATE sessions SET charger_type=?,max_power_kw=?,price_per_kwh=? WHERE id=?",
                                    (new_type,peak_power,price,session_id))
                        con.commit(); charger_type=new_type

            elif not charging and session_active:
                row = cur.execute("SELECT price_per_kwh,cost_manual FROM sessions WHERE id=?",
                                  (session_id,)).fetchone()
                db_price=row[0] if row else None; cost_manual=row[1] if row else 0
                kwh=cost=None
                if soc is not None and soc_start is not None:
                    kwh  = round(max(0.0,soc-soc_start)/100.0*vcfg["battery_capacity_kwh"],2)
                    if not cost_manual:
                        cost = round(kwh*(db_price or cfg["price_per_kwh_home"]),2)
                _meter_end_res = _read_meter_impl(cfg)
                meter_end_val = _meter_end_res.value
                prefer_delta = cfg.get("meter_prefer_meter_delta", False)
                kwh_source = "soc"
                if (prefer_delta and meter_start_val is not None and meter_end_val is not None
                        and meter_end_val >= meter_start_val):
                    meter_delta = round(meter_end_val - meter_start_val, 3)
                    # plausibility: max 250 kWh per session
                    if 0 < meter_delta <= 250:
                        kwh = meter_delta
                        if not cost_manual:
                            cost = round(kwh * (db_price or cfg["price_per_kwh_home"]), 2)
                        kwh_source = "meter"
                cur.execute("""UPDATE sessions
                    SET end_ts=?,odo_end=?,soc_end=?,kwh_charged=?,
                    cost_eur=CASE WHEN cost_manual=1 THEN cost_eur ELSE ? END,
                    max_power_kw=?,meter_new=?,kwh_source=?,
                    meter_delta_kwh=CASE WHEN ? IS NOT NULL AND ? IS NOT NULL THEN ?-? ELSE NULL END
                    WHERE id=?""",
                    (datetime.now().isoformat(timespec="seconds"),odo,soc,kwh,cost,peak_power,
                     meter_end_val,kwh_source,
                     meter_start_val, meter_end_val, meter_end_val, meter_start_val,
                     session_id))
                con.commit(); session_active=False
                st.update(session_active=False,session_id=None)
                log.info("✅ [%s] Session #%d | %.2f kWh | %.2f €",vehicle_id,session_id,kwh or 0,cost or 0)
                ha_notify(vcfg,f"✅ {vcfg['car_name']} fertig",
                    f"{'🏠' if location=='home' else '⚡'} · {kwh or 0:.2f} kWh · {cost or 0:.2f} €")
                session_id=None; peak_power=None
            con.close()

        except Exception as e:
            log.warning("Tracker error [%s]: %s", vehicle_id, e); st["last_error"]=str(e)
        stop.wait(vcfg.get("poll_interval",60))
    st["running"]=False

def _start_vehicle_tracker(vehicle_id: str):
    if vehicle_id not in _vehicle_states:
        cfg = load_config()
        extras = cfg.get("extra_vehicles", [])
        v = next((x for x in extras if x["id"] == vehicle_id), {})
        _vehicle_states[vehicle_id] = _make_state(vehicle_id, v.get("provider","ha"))
    if vehicle_id not in _vehicle_stops:
        _vehicle_stops[vehicle_id] = threading.Event()
    _vehicle_stops[vehicle_id].clear()
    threading.Thread(target=tracker_loop, args=(vehicle_id,), daemon=True).start()

def _stop_vehicle_tracker(vehicle_id: str):
    if vehicle_id in _vehicle_stops:
        _vehicle_stops[vehicle_id].set()

def start_tracker():
    """Start trackers for all active vehicles."""
    _vehicle_stops["v0"].clear()
    threading.Thread(target=tracker_loop, args=("v0",), daemon=True).start()
    # Start extra vehicle trackers
    cfg = load_config()
    for v in cfg.get("extra_vehicles", []):
        if v.get("active", True):
            vid = v["id"]
            _vehicle_states[vid] = _make_state(vid, v.get("provider","ha"))
            _vehicle_stops[vid]  = threading.Event()
            threading.Thread(target=tracker_loop, args=(vid,), daemon=True).start()

# ── Routes ────────────────────────────────────────────────────────────────────

_AUTH_EXEMPT = {"/login", "/logout", "/setup",
                "/auth/google", "/auth/google/callback",
                "/auth/microsoft", "/auth/microsoft/callback",
                "/forgot-password",
                "/api/auth/passkey/login/begin",
                "/api/auth/passkey/login/complete"}

_AUTH_EXEMPT_PREFIXES = ("/reset-password", "/invite")

@app.before_request
def check_auth():
    if request.path.startswith("/static"):
        return
    if request.path in _AUTH_EXEMPT:
        # If users exist, /setup should redirect to index
        if request.path == "/setup" and _has_users():
            return redirect(url_for("index"))
        return
    if any(request.path.startswith(p) for p in _AUTH_EXEMPT_PREFIXES):
        return
    # If no users exist, everything redirects to setup
    if not _has_users():
        if request.path.startswith("/api/"):
            return jsonify({"error": "Setup erforderlich", "setup_required": True}), 503
        return redirect("/setup")
    # Check authentication
    if session.get("user_id"):
        # Ensure CSRF token is set
        if "csrf_token" not in session:
            session["csrf_token"] = secrets.token_hex(32)
        return
    if request.path.startswith("/api/"):
        return jsonify({"error": "Nicht eingeloggt", "login_required": True}), 401
    return redirect(url_for("login_page", next=request.path))

def _check_csrf():
    """Verify CSRF token for state-changing requests. Returns error response or None."""
    if request.method in ("GET", "HEAD", "OPTIONS"):
        return None
    _csrf_exempt_paths = {"/login", "/setup", "/forgot-password",
                          "/api/auth/passkey/login/begin",
                          "/api/auth/passkey/login/complete"}
    _csrf_exempt_prefixes = ("/reset-password", "/invite", "/auth/")
    if request.path in _csrf_exempt_paths:
        return None
    if any(request.path.startswith(p) for p in _csrf_exempt_prefixes):
        return None
    token = request.headers.get("X-CSRF-Token","")
    if not token or token != session.get("csrf_token",""):
        return jsonify({"error": "CSRF-Token ungültig"}), 403
    return None

@app.route("/login", methods=["GET","POST"])
def login_page():
    if session.get("user_id"):
        return redirect(url_for("index"))
    error = None
    if request.method == "POST":
        email = request.form.get("email","").strip().lower()
        pw    = request.form.get("password","")
        user  = _get_user_by_email(email)
        now_dt = datetime.utcnow()
        # Generic error to prevent user enumeration
        if not user or user.get("status") == "disabled":
            error = "Anmeldung fehlgeschlagen"
            _audit("login_failed", f"email={email}", ip=request.remote_addr)
        else:
            # Check account lockout
            locked_until = user.get("locked_until")
            if locked_until:
                try:
                    lu_dt = datetime.fromisoformat(locked_until)
                    if now_dt < lu_dt:
                        error = f"Konto gesperrt bis {lu_dt.strftime('%H:%M')} Uhr. Bitte später erneut versuchen."
                except Exception:
                    pass
            if not error and _hash_password(pw) != user.get("password_hash",""):
                # Wrong password — increment failed attempts
                con = _get_db()
                new_attempts = (user.get("failed_attempts") or 0) + 1
                new_locked = None
                if new_attempts >= 5:
                    new_locked = (now_dt + timedelta(minutes=15)).isoformat()
                con.execute("UPDATE users SET failed_attempts=?,locked_until=? WHERE id=?",
                            (new_attempts, new_locked, user["id"]))
                con.commit(); con.close()
                error = "Anmeldung fehlgeschlagen"
                _audit("login_failed", f"email={email} attempts={new_attempts}", ip=request.remote_addr)
                if new_locked:
                    # notify admins about account lockout
                    try:
                        con2 = _get_db()
                        admins = con2.execute("SELECT email,name FROM users WHERE role='admin' AND status='active'").fetchall()
                        con2.close()
                        for adm in admins:
                            body = _email_html(
                                "⚠️ Konto gesperrt",
                                f"Das Konto <b>{email}</b> wurde wegen zu vieler Fehlversuche für 15 Minuten gesperrt.",
                                f"IP-Adresse: {request.remote_addr}"
                            )
                            _send_email(adm["email"], f"EV Tracker — Konto gesperrt: {email}", body)
                    except Exception:
                        pass
            elif not error:
                # Check TOTP if enabled
                if user.get("totp_enabled") and user.get("totp_secret"):
                    code = request.form.get("totp","").strip().replace(" ","")
                    totp_ok = False
                    try:
                        import pyotp
                        totp_ok = pyotp.TOTP(user["totp_secret"]).verify(code, valid_window=1)
                    except Exception:
                        pass
                    if not totp_ok:
                        # Try backup code
                        code_hash = hashlib.sha256(code.encode()).hexdigest()
                        con = _get_db()
                        backup = con.execute(
                            "SELECT id FROM totp_backup_codes WHERE user_id=? AND code_hash=? AND used_at IS NULL",
                            (user["id"], code_hash)).fetchone()
                        if backup:
                            con.execute("UPDATE totp_backup_codes SET used_at=? WHERE id=?",
                                        (now_dt.isoformat(), backup["id"]))
                            con.commit(); con.close()
                            totp_ok = True
                        else:
                            con.close()
                    if not totp_ok:
                        error = "Ungültiger 2FA-Code"
                if not error:
                    # Reset failed attempts on success
                    con = _get_db()
                    con.execute("UPDATE users SET failed_attempts=0,locked_until=NULL,last_login_at=? WHERE id=?",
                                (now_dt.isoformat(), user["id"]))
                    con.commit(); con.close()
                    session["user_id"]    = user["id"]
                    session["user_email"] = user["email"]
                    session["user_role"]  = user["role"]
                    session["user_name"]  = user["name"]
                    session["csrf_token"] = secrets.token_hex(32)
                    session.permanent     = True
                    _audit("login", f"email={email} role={user['role']}", ip=request.remote_addr)
                    return redirect(request.args.get("next") or url_for("index"))
    cfg = load_config()
    return render_template("login.html", error=error,
                           totp_enabled=False,  # TOTP check done server-side now
                           google_enabled=bool(cfg.get("oauth_google_client_id","")),
                           microsoft_enabled=bool(cfg.get("oauth_microsoft_client_id","")))

@app.route("/logout")
def logout():
    session.clear()
    return redirect(url_for("login_page"))

@app.route("/setup", methods=["GET","POST"])
def setup_page():
    if _has_users():
        return redirect(url_for("index"))
    error = None
    cfg = load_config()
    has_old_auth = bool(cfg.get("auth_password_hash",""))
    if request.method == "POST":
        name  = request.form.get("name","").strip()
        email = request.form.get("email","").strip().lower()
        pw    = request.form.get("password","")
        pw2   = request.form.get("password2","")
        if not name or not email or not pw:
            error = "Alle Felder sind erforderlich"
        elif pw != pw2:
            error = "Passwörter stimmen nicht überein"
        elif _password_ok(pw):
            error = _password_ok(pw)
        else:
            now = datetime.utcnow().isoformat()
            con = _get_db()
            try:
                con.execute(
                    "INSERT INTO users (name,email,password_hash,role,status,created_at,updated_at) VALUES (?,?,?,?,?,?,?)",
                    (name, email, _hash_password(pw), "admin", "active", now, now))
                con.commit()
            except sqlite3.IntegrityError:
                error = "E-Mail-Adresse bereits vorhanden"
            finally:
                con.close()
            if not error:
                # Clear old single-user auth from config
                cfg["auth_password_hash"] = ""
                cfg["auth_totp_secret"]   = ""
                save_config(cfg)
                _audit("setup_complete", f"admin={email}", ip=request.remote_addr)
                return redirect(url_for("login_page"))
    return render_template("setup.html", error=error, has_old_auth=has_old_auth)

# ── Password Reset ────────────────────────────────────────────────────────────

@app.route("/forgot-password", methods=["GET","POST"])
def forgot_password_page():
    if request.method == "GET":
        sent = request.args.get("sent","")
        return render_template("forgot_password.html", sent=bool(sent))
    email = request.form.get("email","").strip().lower()
    # Rate limit: max 3 requests per email per hour
    now_ts = time.time()
    attempts = _forgot_pw_attempts.get(email, [])
    attempts = [t for t in attempts if now_ts - t < 3600]  # last hour
    if len(attempts) >= 3:
        # Still redirect to avoid timing oracle, just don't send
        return redirect(url_for("forgot_password_page", sent=1))
    attempts.append(now_ts)
    _forgot_pw_attempts[email] = attempts
    user  = _get_user_by_email(email)
    if user and user.get("status") != "disabled":
        token = secrets.token_urlsafe(32)
        token_hash = hashlib.sha256(token.encode()).hexdigest()
        now_dt = datetime.utcnow()
        expires = (now_dt + timedelta(hours=1)).isoformat()
        con = _get_db()
        # Invalidate old tokens for this user
        con.execute("UPDATE password_reset_tokens SET used_at=? WHERE user_id=? AND used_at IS NULL",
                    (now_dt.isoformat(), user["id"]))
        con.execute("INSERT INTO password_reset_tokens (user_id,token_hash,expires_at,created_at) VALUES (?,?,?,?)",
                    (user["id"], token_hash, expires, now_dt.isoformat()))
        con.commit(); con.close()
        reset_url = request.host_url.rstrip("/") + url_for("reset_password_page", token=token)
        body_html = _email_html(
            "Passwort zurücksetzen",
            f"Hallo {user['name']},",
            "du hast einen Passwort-Reset für deinen EV Tracker Account angefordert.",
            _email_btn(reset_url, "🔑 Passwort zurücksetzen"),
            "Dieser Link ist <b>1 Stunde</b> gültig. Falls du diese Anfrage nicht gestellt hast, kannst du diese E-Mail ignorieren."
        )
        _send_email(user["email"], "EV Tracker — Passwort zurücksetzen", body_html)
        _audit("password_reset_requested", f"email={email}", ip=request.remote_addr)
    return redirect(url_for("forgot_password_page", sent=1))

@app.route("/reset-password/<token>", methods=["GET","POST"])
def reset_password_page(token):
    token_hash = hashlib.sha256(token.encode()).hexdigest()
    now_iso = datetime.utcnow().isoformat()
    con = _get_db()
    row = con.execute(
        "SELECT * FROM password_reset_tokens WHERE token_hash=? AND used_at IS NULL AND expires_at > ?",
        (token_hash, now_iso)).fetchone()
    if not row:
        con.close()
        return render_template("reset_password.html", token=token, invalid=True)
    row = dict(row)
    if request.method == "GET":
        con.close()
        return render_template("reset_password.html", token=token, invalid=False)
    pw  = request.form.get("password","")
    pw2 = request.form.get("password2","")
    pw_err = _password_ok(pw)
    if pw_err:
        con.close()
        return render_template("reset_password.html", token=token, invalid=False,
                               error=pw_err)
    if pw != pw2:
        con.close()
        return render_template("reset_password.html", token=token, invalid=False,
                               error="Passwörter stimmen nicht überein")
    now_iso2 = datetime.utcnow().isoformat()
    con.execute("UPDATE users SET password_hash=?,updated_at=?,failed_attempts=0,locked_until=NULL WHERE id=?",
                (_hash_password(pw), now_iso2, row["user_id"]))
    con.execute("UPDATE password_reset_tokens SET used_at=? WHERE id=?", (now_iso2, row["id"]))
    con.commit(); con.close()
    _audit("password_reset_done", f"uid={row['user_id']}", ip=request.remote_addr)
    return redirect(url_for("login_page"))

# ── User Invitations ──────────────────────────────────────────────────────────

@app.route("/invite/<token>", methods=["GET","POST"])
def accept_invite_page(token):
    token_hash = hashlib.sha256(token.encode()).hexdigest()
    now_iso = datetime.utcnow().isoformat()
    con = _get_db()
    row = con.execute(
        "SELECT * FROM invite_tokens WHERE token_hash=? AND used_at IS NULL AND expires_at > ?",
        (token_hash, now_iso)).fetchone()
    if not row:
        con.close()
        return render_template("accept_invite.html", token=token, invalid=True)
    row = dict(row)
    user = _get_user_by_id(row["user_id"])
    if not user:
        con.close()
        return render_template("accept_invite.html", token=token, invalid=True)
    if request.method == "GET":
        con.close()
        return render_template("accept_invite.html", token=token, invalid=False,
                               user_name=user.get("name",""))
    pw  = request.form.get("password","")
    pw2 = request.form.get("password2","")
    pw_err = _password_ok(pw)
    if pw_err:
        con.close()
        return render_template("accept_invite.html", token=token, invalid=False,
                               user_name=user.get("name",""),
                               error=pw_err)
    if pw != pw2:
        con.close()
        return render_template("accept_invite.html", token=token, invalid=False,
                               user_name=user.get("name",""),
                               error="Passwörter stimmen nicht überein")
    now_iso2 = datetime.utcnow().isoformat()
    new_status = "active" if user.get("status") == "invited" else user.get("status","active")
    con.execute("UPDATE users SET password_hash=?,status=?,updated_at=? WHERE id=?",
                (_hash_password(pw), new_status, now_iso2, user["id"]))
    con.execute("UPDATE invite_tokens SET used_at=? WHERE id=?", (now_iso2, row["id"]))
    con.commit(); con.close()
    _audit("invite_accepted", f"uid={user['id']}", ip=request.remote_addr)
    return redirect(url_for("login_page"))

# ── SMTP helper ───────────────────────────────────────────────────────────────

def _email_html(title: str, *paragraphs: str) -> str:
    paras = "".join(f"<p style='margin:0 0 14px;line-height:1.6'>{p}</p>" for p in paragraphs)
    return f"""<!DOCTYPE html>
<html><head><meta charset='utf-8'></head>
<body style='margin:0;padding:0;background:#0f0f0f;font-family:system-ui,sans-serif'>
  <table width='100%' cellpadding='0' cellspacing='0'>
    <tr><td align='center' style='padding:40px 20px'>
      <table width='560' cellpadding='0' cellspacing='0' style='background:#1a1a1a;border-radius:12px;overflow:hidden;border:1px solid #2a2a2a'>
        <tr><td style='background:#1e3a5f;padding:24px 32px'>
          <div style='font-size:1.1rem;font-weight:700;color:#7eb8f7;letter-spacing:.05em'>⚡ EV Tracker</div>
        </td></tr>
        <tr><td style='padding:32px'>
          <h2 style='margin:0 0 20px;color:#e0e0e0;font-size:1.1rem;font-weight:600'>{title}</h2>
          <div style='color:#b0b0b0;font-size:.875rem'>{paras}</div>
        </td></tr>
        <tr><td style='padding:16px 32px;border-top:1px solid #2a2a2a'>
          <div style='color:#555;font-size:.75rem'>Diese E-Mail wurde automatisch von EV Tracker generiert.</div>
        </td></tr>
      </table>
    </td></tr>
  </table>
</body></html>"""

def _email_btn(url: str, label: str) -> str:
    return f"<a href='{url}' style='display:inline-block;background:#1e6fb5;color:#fff;text-decoration:none;padding:12px 28px;border-radius:8px;font-size:.875rem;font-weight:600;margin:8px 0'>{label}</a>"

def _send_email(to_addr: str, subject: str, body_html: str, body_text: str = None) -> tuple:
    import smtplib, ssl as _ssl
    from email.mime.multipart import MIMEMultipart
    from email.mime.text import MIMEText as _MIMEText
    cfg  = load_config()
    host = cfg.get("smtp_host","")
    port = int(cfg.get("smtp_port", 587))
    tls  = cfg.get("smtp_tls","starttls")
    user = cfg.get("smtp_user","")
    pw   = cfg.get("smtp_password","")
    frm  = cfg.get("smtp_from_email","")
    name = cfg.get("smtp_from_name","EV Tracker")
    if not host or not frm:
        return False, "SMTP nicht konfiguriert"
    try:
        msg = MIMEMultipart("alternative")
        msg["Subject"] = subject
        msg["From"]    = f"{name} <{frm}>"
        msg["To"]      = to_addr
        if body_text:
            msg.attach(_MIMEText(body_text, "plain", "utf-8"))
        msg.attach(_MIMEText(body_html, "html", "utf-8"))
        ctx = _ssl.create_default_context()
        if tls == "ssl":
            srv = smtplib.SMTP_SSL(host, port, context=ctx, timeout=10)
        else:
            srv = smtplib.SMTP(host, port, timeout=10)
            if tls == "starttls":
                srv.starttls(context=ctx)
        if user:
            srv.login(user, pw)
        srv.sendmail(frm, [to_addr], msg.as_string())
        srv.quit()
        return True, None
    except Exception as e:
        return False, str(e)

@app.route("/api/auth/setup", methods=["POST"])
def api_auth_setup():
    data = request.json or {}
    cfg  = load_config()
    if "password" in data and data["password"]:
        cfg["auth_password_hash"] = _hash_password(data["password"])
        _audit("password_set", ip=request.remote_addr)
    if data.get("disable_password"):
        cfg["auth_password_hash"] = ""
        cfg["auth_totp_secret"]   = ""
        _audit("password_disabled", ip=request.remote_addr)
    save_config(cfg)
    return jsonify({"ok": True})

@app.route("/api/auth/totp/setup", methods=["POST"])
def api_totp_setup():
    import pyotp
    secret = pyotp.random_base32()
    cfg = load_config()
    cfg["auth_totp_secret"] = secret
    save_config(cfg)
    car_name = cfg.get("car_name","EV Tracker")
    email    = cfg.get("ha_token","")[:6] or "user"
    uri = pyotp.TOTP(secret).provisioning_uri(name=email, issuer_name=f"EV Tracker ({car_name})")
    return jsonify({"ok": True, "secret": secret, "uri": uri})

@app.route("/api/auth/totp/disable", methods=["POST"])
def api_totp_disable():
    cfg = load_config()
    cfg["auth_totp_secret"] = ""
    save_config(cfg)
    _audit("totp_disabled", ip=request.remote_addr)
    return jsonify({"ok": True})

@app.route("/api/auth/status")
def api_auth_status():
    cfg = load_config()
    user = _current_user()
    return jsonify({
        "auth_enabled":       _has_users(),
        "user_id":            session.get("user_id"),
        "user_email":         session.get("user_email",""),
        "user_name":          session.get("user_name",""),
        "user_role":          session.get("user_role",""),
        "totp_enabled":       bool(user.get("totp_enabled")) if user else False,
        "google_enabled":     bool(cfg.get("oauth_google_client_id","")),
        "microsoft_enabled":  bool(cfg.get("oauth_microsoft_client_id","")),
        "has_users":          _has_users(),
    })

# ── WebAuthn / Passkey helpers ────────────────────────────────────────────────

def _webauthn_rp_id() -> str:
    """Get WebAuthn relying party ID (hostname without port)."""
    base_url = load_config().get("oauth_base_url", "").rstrip("/")
    if base_url:
        from urllib.parse import urlparse
        return urlparse(base_url).hostname or "localhost"
    host = request.host or "localhost"
    return host.split(":")[0]  # strip port

def _webauthn_rp_name() -> str:
    return "EV Tracker"

def _webauthn_origin() -> str:
    """Get expected origin for WebAuthn."""
    base_url = load_config().get("oauth_base_url", "").rstrip("/")
    if base_url:
        return base_url
    # Respect X-Forwarded-Proto from reverse proxies (nginx, Traefik, etc.)
    scheme = request.headers.get("X-Forwarded-Proto", request.scheme)
    return f"{scheme}://{request.host}"

# ── WebAuthn / Passkey routes ─────────────────────────────────────────────────

@app.route("/api/auth/passkey/register/begin", methods=["POST"])
@require_login
def api_passkey_register_begin():
    import webauthn
    from webauthn.helpers.structs import (
        AuthenticatorSelectionCriteria, ResidentKeyRequirement,
        UserVerificationRequirement, AttestationConveyancePreference
    )
    user = _current_user()
    if not user:
        return jsonify({"ok": False, "error": "Nicht angemeldet"}), 401

    con = _get_db()
    existing = con.execute(
        "SELECT credential_id FROM webauthn_credentials WHERE user_id=?",
        (user["id"],)
    ).fetchall()
    con.close()

    from webauthn.helpers.structs import PublicKeyCredentialDescriptor
    from webauthn.helpers import base64url_to_bytes
    exclude_creds = []
    for row in existing:
        try:
            exclude_creds.append(PublicKeyCredentialDescriptor(
                id=base64url_to_bytes(row["credential_id"])
            ))
        except Exception:
            pass

    try:
        options = webauthn.generate_registration_options(
            rp_id=_webauthn_rp_id(),
            rp_name=_webauthn_rp_name(),
            user_id=str(user["id"]).encode(),
            user_name=user["email"],
            user_display_name=user["name"],
            authenticator_selection=AuthenticatorSelectionCriteria(
                resident_key=ResidentKeyRequirement.PREFERRED,
                user_verification=UserVerificationRequirement.PREFERRED,
            ),
            attestation=AttestationConveyancePreference.NONE,
            exclude_credentials=exclude_creds,
        )
        session["webauthn_reg_challenge"] = webauthn.helpers.bytes_to_base64url(options.challenge)
        import json as _json
        from webauthn.helpers import options_to_json
        return jsonify({"ok": True, "options": _json.loads(options_to_json(options))})
    except Exception as e:
        log.exception("Passkey register begin failed")
        return jsonify({"ok": False, "error": str(e)}), 500


@app.route("/api/auth/passkey/register/complete", methods=["POST"])
@require_login
def api_passkey_register_complete():
    import webauthn
    user = _current_user()
    if not user:
        return jsonify({"ok": False, "error": "Nicht angemeldet"}), 401

    challenge_b64 = session.pop("webauthn_reg_challenge", None)
    if not challenge_b64:
        return jsonify({"ok": False, "error": "Keine Registrierungs-Challenge gefunden"}), 400

    body = request.get_json(force=True) or {}
    cred_name = body.pop("name", "Passkey")

    try:
        from webauthn.helpers import base64url_to_bytes, bytes_to_base64url
        expected_challenge = base64url_to_bytes(challenge_b64)

        verification = webauthn.verify_registration_response(
            credential=body,
            expected_challenge=expected_challenge,
            expected_rp_id=_webauthn_rp_id(),
            expected_origin=_webauthn_origin(),
            require_user_verification=False,
        )

        # Use body["id"] directly — same base64url the browser will send on login
        cred_id = body.get("id") or bytes_to_base64url(verification.credential_id)
        pub_key = bytes_to_base64url(verification.credential_public_key)

        con = _get_db()
        now = datetime.utcnow().isoformat()
        con.execute(
            "INSERT INTO webauthn_credentials (user_id, credential_id, public_key, sign_count, name, created_at) VALUES (?,?,?,?,?,?)",
            (user["id"], cred_id, pub_key, verification.sign_count, cred_name, now)
        )
        con.commit(); con.close()
        _audit("passkey_registered", f"user={user['email']} name={cred_name}", ip=request.remote_addr)
        return jsonify({"ok": True, "message": f"Passkey '{cred_name}' erfolgreich registriert"})
    except Exception as e:
        log.exception("Passkey register complete failed")
        return jsonify({"ok": False, "error": str(e)}), 400


@app.route("/api/auth/passkey/credentials")
@require_login
def api_passkey_credentials():
    user = _current_user()
    if not user:
        return jsonify({"ok": False, "error": "Nicht angemeldet"}), 401
    con = _get_db()
    rows = con.execute(
        "SELECT id, name, created_at, last_used_at FROM webauthn_credentials WHERE user_id=? ORDER BY created_at DESC",
        (user["id"],)
    ).fetchall()
    con.close()
    return jsonify({"ok": True, "credentials": [dict(r) for r in rows]})


@app.route("/api/auth/passkey/credentials/<int:cred_db_id>", methods=["DELETE"])
@require_login
def api_passkey_credential_delete(cred_db_id):
    user = _current_user()
    if not user:
        return jsonify({"ok": False, "error": "Nicht angemeldet"}), 401
    con = _get_db()
    con.execute("DELETE FROM webauthn_credentials WHERE id=? AND user_id=?", (cred_db_id, user["id"]))
    con.commit(); con.close()
    _audit("passkey_deleted", f"cred_id={cred_db_id}", ip=request.remote_addr)
    return jsonify({"ok": True})


@app.route("/api/auth/passkey/login/begin", methods=["POST"])
def api_passkey_login_begin():
    import webauthn
    from webauthn.helpers.structs import UserVerificationRequirement

    body = request.get_json(silent=True) or {}
    email = body.get("email", "").strip().lower()

    allow_creds = []
    if email:
        user = _get_user_by_email(email)
        if user:
            con = _get_db()
            rows = con.execute(
                "SELECT credential_id FROM webauthn_credentials WHERE user_id=?",
                (user["id"],)
            ).fetchall()
            con.close()
            from webauthn.helpers.structs import PublicKeyCredentialDescriptor
            from webauthn.helpers import base64url_to_bytes
            for row in rows:
                try:
                    allow_creds.append(PublicKeyCredentialDescriptor(
                        id=base64url_to_bytes(row["credential_id"])
                    ))
                except Exception:
                    pass

    try:
        options = webauthn.generate_authentication_options(
            rp_id=_webauthn_rp_id(),
            allow_credentials=allow_creds,
            user_verification=UserVerificationRequirement.PREFERRED,
        )
        from webauthn.helpers import bytes_to_base64url
        session["webauthn_auth_challenge"] = bytes_to_base64url(options.challenge)
        import json as _json
        from webauthn.helpers import options_to_json
        return jsonify({"ok": True, "options": _json.loads(options_to_json(options))})
    except Exception as e:
        log.exception("Passkey login begin failed")
        return jsonify({"ok": False, "error": str(e)}), 500


@app.route("/api/auth/passkey/login/complete", methods=["POST"])
def api_passkey_login_complete():
    import webauthn

    challenge_b64 = session.pop("webauthn_auth_challenge", None)
    if not challenge_b64:
        return jsonify({"ok": False, "error": "Keine Authentifizierungs-Challenge"}), 400

    body = request.get_json(force=True) or {}

    try:
        from webauthn.helpers import base64url_to_bytes, bytes_to_base64url
        expected_challenge = base64url_to_bytes(challenge_b64)

        # credential id is the base64url "id" field from the browser response
        cred_id_b64 = body.get("id", "")

        con = _get_db()
        row = con.execute(
            """SELECT wc.id as cred_id, wc.credential_id, wc.public_key, wc.sign_count, wc.name as cred_name,
                      u.id as user_id, u.email, u.name, u.role, u.status
               FROM webauthn_credentials wc
               JOIN users u ON wc.user_id = u.id
               WHERE wc.credential_id=?""",
            (cred_id_b64,)
        ).fetchone()

        if not row:
            con.close()
            return jsonify({"ok": False, "error": "Passkey nicht registriert. Bitte mit Passwort einloggen und den Passkey unter Einstellungen → Sicherheit neu hinzufügen."}), 400

        row = dict(row)
        if row.get("status") == "disabled":
            con.close()
            return jsonify({"ok": False, "error": "Konto deaktiviert"}), 403

        verification = webauthn.verify_authentication_response(
            credential=body,
            expected_challenge=expected_challenge,
            expected_rp_id=_webauthn_rp_id(),
            expected_origin=_webauthn_origin(),
            credential_public_key=base64url_to_bytes(row["public_key"]),
            credential_current_sign_count=row["sign_count"],
            require_user_verification=False,
        )

        now = datetime.utcnow().isoformat()
        con.execute(
            "UPDATE webauthn_credentials SET sign_count=?, last_used_at=? WHERE credential_id=?",
            (verification.new_sign_count, now, cred_id_b64)
        )
        con.execute(
            "UPDATE users SET last_login_at=?, failed_attempts=0, locked_until=NULL WHERE id=?",
            (now, row["user_id"])
        )
        con.commit(); con.close()

        session["user_id"]    = row["user_id"]
        session["user_email"] = row["email"]
        session["user_role"]  = row["role"]
        session["user_name"]  = row["name"]
        _audit("passkey_login", f"user={row['email']}", ip=request.remote_addr)

        return jsonify({"ok": True, "redirect": "/"})
    except Exception as e:
        log.exception("Passkey login complete failed")
        return jsonify({"ok": False, "error": str(e)}), 400


# ── Permissions API ───────────────────────────────────────────────────────────

@app.route("/api/me/permissions")
@require_login
def api_me_permissions():
    user = _current_user()
    if not user:
        return jsonify({"ok": False, "error": "Nicht angemeldet"}), 401
    con = _get_db()
    # Rollen des Users
    roles = con.execute("""
        SELECT r.name FROM user_roles ur JOIN roles r ON ur.role_id = r.id
        WHERE ur.user_id = ?
    """, (user["id"],)).fetchall()
    con.close()
    perms = _get_user_permissions(user["id"])
    # Wenn admin:all → alle Permissions zurückgeben
    if "admin:all" in perms:
        perms = set(ALL_PERMISSIONS.keys())
    return jsonify({
        "ok": True,
        "user": {"id": user["id"], "email": user.get("email",""), "name": user.get("name",""),
                 "roles": [r["name"] for r in roles]},
        "permissions": sorted(perms),
    })


@app.route("/api/admin/permissions")
@require_login
def api_admin_permissions_list():
    """Return all known permission keys with metadata."""
    user = _current_user()
    if not user or not has_permission(user, "users:manage_permissions"):
        return jsonify({"ok": False, "error": "Keine Berechtigung"}), 403
    groups = {}
    for key, meta in ALL_PERMISSIONS.items():
        g = meta["group"]
        if g not in groups:
            groups[g] = []
        groups[g].append({"key": key, "label": meta["label"]})
    return jsonify({"ok": True, "groups": [{"name": g, "permissions": p} for g,p in groups.items()]})


# ── Roles CRUD ────────────────────────────────────────────────────────────────

@app.route("/api/admin/roles")
@require_login
def api_admin_roles_list():
    user = _current_user()
    if not user or not has_permission(user, "users:manage_permissions"):
        return jsonify({"ok": False, "error": "Keine Berechtigung"}), 403
    con = _get_db()
    roles = [dict(r) for r in con.execute("SELECT * FROM roles ORDER BY is_system DESC, name").fetchall()]
    for role in roles:
        perms = [r["permission_key"] for r in
            con.execute("SELECT permission_key FROM role_permissions WHERE role_id=?",
                       (role["id"],)).fetchall()]
        role["permissions"] = perms
        user_count = con.execute("SELECT COUNT(*) as c FROM user_roles WHERE role_id=?",
                                  (role["id"],)).fetchone()["c"]
        role["user_count"] = user_count
    con.close()
    return jsonify({"ok": True, "roles": roles})


@app.route("/api/admin/roles", methods=["POST"])
@require_login
def api_admin_roles_create():
    user = _current_user()
    if not user or not has_permission(user, "users:manage_permissions"):
        return jsonify({"ok": False, "error": "Keine Berechtigung"}), 403
    body = request.get_json(force=True) or {}
    name = (body.get("name") or "").strip()
    if not name or len(name) < 2:
        return jsonify({"ok": False, "error": "Rollenname zu kurz"}), 400
    if not re.match(r'^[a-zA-Z0-9_\-äöüÄÖÜß ]+$', name):
        return jsonify({"ok": False, "error": "Ungültiger Rollenname"}), 400
    desc = (body.get("description") or "").strip()[:200]
    perms = [p for p in (body.get("permissions") or []) if p in ALL_PERMISSIONS]
    now_iso = datetime.utcnow().isoformat()
    con = _get_db()
    try:
        cur = con.execute(
            "INSERT INTO roles (name, description, is_system, created_at, updated_at) VALUES (?,?,0,?,?)",
            (name, desc, now_iso, now_iso)
        )
        role_id = cur.lastrowid
        for pkey in perms:
            con.execute("INSERT INTO role_permissions (role_id, permission_key) VALUES (?,?)",
                        (role_id, pkey))
        con.commit()
        _audit("role_created", f"name={name} perms={len(perms)}", ip=request.remote_addr)
        return jsonify({"ok": True, "role_id": role_id})
    except Exception as e:
        return jsonify({"ok": False, "error": str(e)}), 400
    finally:
        con.close()


@app.route("/api/admin/roles/<int:role_id>", methods=["PUT"])
@require_login
def api_admin_roles_update(role_id):
    user = _current_user()
    if not user or not has_permission(user, "users:manage_permissions"):
        return jsonify({"ok": False, "error": "Keine Berechtigung"}), 403
    body = request.get_json(force=True) or {}
    con = _get_db()
    role = con.execute("SELECT * FROM roles WHERE id=?", (role_id,)).fetchone()
    if not role:
        con.close()
        return jsonify({"ok": False, "error": "Rolle nicht gefunden"}), 404
    role = dict(role)
    name = (body.get("name") or role["name"]).strip()
    desc = (body.get("description") or role["description"] or "").strip()[:200]
    perms = [p for p in (body.get("permissions") or []) if p in ALL_PERMISSIONS]
    now_iso = datetime.utcnow().isoformat()
    try:
        con.execute("UPDATE roles SET name=?, description=?, updated_at=? WHERE id=?",
                    (name, desc, now_iso, role_id))
        con.execute("DELETE FROM role_permissions WHERE role_id=?", (role_id,))
        for pkey in perms:
            con.execute("INSERT INTO role_permissions (role_id, permission_key) VALUES (?,?)",
                        (role_id, pkey))
        con.commit()
        _audit("role_updated", f"id={role_id} name={name} perms={len(perms)}", ip=request.remote_addr)
        return jsonify({"ok": True})
    except Exception as e:
        return jsonify({"ok": False, "error": str(e)}), 400
    finally:
        con.close()


@app.route("/api/admin/roles/<int:role_id>", methods=["DELETE"])
@require_login
def api_admin_roles_delete(role_id):
    user = _current_user()
    if not user or not has_permission(user, "users:manage_permissions"):
        return jsonify({"ok": False, "error": "Keine Berechtigung"}), 403
    con = _get_db()
    role = con.execute("SELECT * FROM roles WHERE id=?", (role_id,)).fetchone()
    if not role:
        con.close()
        return jsonify({"ok": False, "error": "Rolle nicht gefunden"}), 404
    role = dict(role)
    if role["is_system"]:
        con.close()
        return jsonify({"ok": False, "error": "Systemrolle kann nicht gelöscht werden"}), 400
    # Sicherheit: mindestens ein User mit admin:all muss bleiben
    admin_role = con.execute("SELECT id FROM roles WHERE name='admin'").fetchone()
    if admin_role:
        remaining_admins = con.execute("""
            SELECT COUNT(*) as c FROM user_roles ur
            JOIN role_permissions rp ON ur.role_id = rp.role_id
            WHERE rp.permission_key = 'admin:all' AND ur.role_id != ?
        """, (role_id,)).fetchone()["c"]
        if remaining_admins == 0 and role.get("name") == "admin":
            con.close()
            return jsonify({"ok": False, "error": "Mindestens eine Admin-Rolle muss erhalten bleiben"}), 400
    con.execute("DELETE FROM role_permissions WHERE role_id=?", (role_id,))
    con.execute("DELETE FROM user_roles WHERE role_id=?", (role_id,))
    con.execute("DELETE FROM roles WHERE id=?", (role_id,))
    con.commit()
    con.close()
    _audit("role_deleted", f"id={role_id} name={role['name']}", ip=request.remote_addr)
    return jsonify({"ok": True})


# ── User Roles ────────────────────────────────────────────────────────────────

@app.route("/api/admin/users/<int:target_user_id>/roles")
@require_login
def api_admin_user_roles_get(target_user_id):
    user = _current_user()
    if not user or not has_permission(user, "users:manage_permissions"):
        return jsonify({"ok": False, "error": "Keine Berechtigung"}), 403
    con = _get_db()
    roles = [r["role_id"] for r in
        con.execute("SELECT role_id FROM user_roles WHERE user_id=?", (target_user_id,)).fetchall()]
    con.close()
    return jsonify({"ok": True, "role_ids": roles})


@app.route("/api/admin/users/<int:target_user_id>/roles", methods=["PUT"])
@require_login
def api_admin_user_roles_set(target_user_id):
    user = _current_user()
    if not user or not has_permission(user, "users:manage_permissions"):
        return jsonify({"ok": False, "error": "Keine Berechtigung"}), 403
    body = request.get_json(force=True) or {}
    role_ids = [int(r) for r in (body.get("role_ids") or [])]
    con = _get_db()
    # Sicherheit: mind. ein User mit admin:all
    current_user_has_admin = "admin:all" in _get_user_permissions(target_user_id)
    if current_user_has_admin and target_user_id == user["id"]:
        new_perms_check = set()
        for rid in role_ids:
            p = {r["permission_key"] for r in
                 con.execute("SELECT permission_key FROM role_permissions WHERE role_id=?", (rid,)).fetchall()}
            new_perms_check |= p
        if "admin:all" not in new_perms_check:
            # Check if other admins exist
            other_admins = con.execute("""
                SELECT COUNT(DISTINCT ur.user_id) as c FROM user_roles ur
                JOIN role_permissions rp ON ur.role_id = rp.role_id
                WHERE rp.permission_key = 'admin:all' AND ur.user_id != ?
            """, (target_user_id,)).fetchone()["c"]
            if other_admins == 0:
                con.close()
                return jsonify({"ok": False, "error": "Mindestens ein Admin muss admin:all behalten"}), 400
    # Validate role_ids
    valid_ids = {r["id"] for r in con.execute("SELECT id FROM roles").fetchall()}
    role_ids = [r for r in role_ids if r in valid_ids]
    con.execute("DELETE FROM user_roles WHERE user_id=?", (target_user_id,))
    for rid in role_ids:
        con.execute("INSERT INTO user_roles (user_id, role_id) VALUES (?,?)", (target_user_id, rid))
    con.commit()
    # Also update legacy users.role field for backwards compat
    if role_ids:
        first_role = con.execute("SELECT name FROM roles WHERE id=?", (role_ids[0],)).fetchone()
        if first_role:
            con.execute("UPDATE users SET role=? WHERE id=?", (first_role["name"], target_user_id))
            con.commit()
    con.close()
    _audit("user_roles_changed", f"user_id={target_user_id} roles={role_ids}", ip=request.remote_addr)
    return jsonify({"ok": True})


# ── OAuth2 helpers ────────────────────────────────────────────────────────────

def _oauth_redirect_base() -> str:
    cfg = load_config()
    base = cfg.get("oauth_base_url","").rstrip("/")
    if base:
        return base
    # Auto-detect: honour X-Forwarded-Proto for reverse proxy
    scheme = request.headers.get("X-Forwarded-Proto", request.scheme)
    host   = request.headers.get("X-Forwarded-Host", request.host)
    return f"{scheme}://{host}"

def _oauth_finish(email: str):
    """Called after successful OAuth2 — find or reject user, set session."""
    user = _get_user_by_email(email)
    if not user:
        # Auto-create user if no users exist yet (shouldn't happen post-setup, but just in case)
        if not _has_users():
            return redirect("/setup")
        # Deny if user not in DB
        return render_template("login.html", error=f"Kein Konto für {email} vorhanden. Bitte Admin kontaktieren.",
                               totp_enabled=False, google_enabled=False, microsoft_enabled=False)
    if user.get("status") == "disabled":
        return render_template("login.html", error="Konto deaktiviert.",
                               totp_enabled=False, google_enabled=False, microsoft_enabled=False)
    session["user_id"]    = user["id"]
    session["user_email"] = user["email"]
    session["user_role"]  = user["role"]
    session["user_name"]  = user["name"]
    session.permanent     = True
    now = datetime.utcnow().isoformat()
    con = _get_db()
    con.execute("UPDATE users SET last_login_at=? WHERE id=?", (now, user["id"]))
    con.commit(); con.close()
    _audit("login_oauth", f"email={email} role={user['role']}", ip=request.remote_addr)
    return redirect(request.args.get("next") or url_for("index"))

# ── Google OAuth2 ─────────────────────────────────────────────────────────────

@app.route("/auth/google")
def auth_google():
    cfg = load_config()
    if not cfg.get("oauth_google_client_id"):
        return "Google OAuth nicht konfiguriert", 400
    state = secrets.token_urlsafe(16)
    session["oauth_state"] = state
    session["oauth_next"]  = request.args.get("next","/")
    redirect_uri = _oauth_redirect_base() + "/auth/google/callback"
    params = {
        "client_id":     cfg["oauth_google_client_id"],
        "redirect_uri":  redirect_uri,
        "response_type": "code",
        "scope":         "openid email profile",
        "state":         state,
        "access_type":   "online",
    }
    from urllib.parse import urlencode
    return redirect("https://accounts.google.com/o/oauth2/v2/auth?" + urlencode(params))

@app.route("/auth/google/callback")
def auth_google_callback():
    cfg   = load_config()
    state = request.args.get("state","")
    code  = request.args.get("code","")
    if not code or state != session.pop("oauth_state",""):
        return render_template("login.html", error="OAuth-Fehler: ungültiger State", totp_enabled=False,
                               google_enabled=bool(cfg.get("oauth_google_client_id")),
                               microsoft_enabled=bool(cfg.get("oauth_microsoft_client_id")))
    redirect_uri = _oauth_redirect_base() + "/auth/google/callback"
    try:
        r = requests.post("https://oauth2.googleapis.com/token", data={
            "code":          code,
            "client_id":     cfg["oauth_google_client_id"],
            "client_secret": cfg["oauth_google_client_secret"],
            "redirect_uri":  redirect_uri,
            "grant_type":    "authorization_code",
        }, timeout=10)
        r.raise_for_status()
        token = r.json().get("access_token","")
        info  = requests.get("https://www.googleapis.com/oauth2/v3/userinfo",
                             headers={"Authorization":f"Bearer {token}"}, timeout=10).json()
        email = info.get("email","")
        if not email:
            raise RuntimeError("Keine E-Mail vom Google-Konto erhalten")
        return _oauth_finish(email)
    except Exception as e:
        return render_template("login.html", error=f"Google Login fehlgeschlagen: {e}",
                               totp_enabled=False,
                               google_enabled=True, microsoft_enabled=bool(cfg.get("oauth_microsoft_client_id")))

# ── Microsoft OAuth2 ──────────────────────────────────────────────────────────

@app.route("/auth/microsoft")
def auth_microsoft():
    cfg = load_config()
    if not cfg.get("oauth_microsoft_client_id"):
        return "Microsoft OAuth nicht konfiguriert", 400
    tenant = cfg.get("oauth_microsoft_tenant","common") or "common"
    state  = secrets.token_urlsafe(16)
    session["oauth_state"] = state
    session["oauth_next"]  = request.args.get("next","/")
    redirect_uri = _oauth_redirect_base() + "/auth/microsoft/callback"
    from urllib.parse import urlencode
    params = {
        "client_id":     cfg["oauth_microsoft_client_id"],
        "redirect_uri":  redirect_uri,
        "response_type": "code",
        "scope":         "openid email profile User.Read",
        "state":         state,
        "response_mode": "query",
    }
    return redirect(f"https://login.microsoftonline.com/{tenant}/oauth2/v2.0/authorize?" + urlencode(params))

@app.route("/auth/microsoft/callback")
def auth_microsoft_callback():
    cfg    = load_config()
    tenant = cfg.get("oauth_microsoft_tenant","common") or "common"
    state  = request.args.get("state","")
    code   = request.args.get("code","")
    if not code or state != session.pop("oauth_state",""):
        return render_template("login.html", error="OAuth-Fehler: ungültiger State", totp_enabled=False,
                               google_enabled=bool(cfg.get("oauth_google_client_id")),
                               microsoft_enabled=bool(cfg.get("oauth_microsoft_client_id")))
    redirect_uri = _oauth_redirect_base() + "/auth/microsoft/callback"
    try:
        r = requests.post(
            f"https://login.microsoftonline.com/{tenant}/oauth2/v2.0/token",
            data={
                "code":          code,
                "client_id":     cfg["oauth_microsoft_client_id"],
                "client_secret": cfg["oauth_microsoft_client_secret"],
                "redirect_uri":  redirect_uri,
                "grant_type":    "authorization_code",
            }, timeout=10)
        r.raise_for_status()
        token = r.json().get("access_token","")
        info  = requests.get("https://graph.microsoft.com/v1.0/me",
                             headers={"Authorization":f"Bearer {token}"}, timeout=10).json()
        email = info.get("mail") or info.get("userPrincipalName","")
        if not email:
            raise RuntimeError("Keine E-Mail vom Microsoft-Konto erhalten")
        return _oauth_finish(email)
    except Exception as e:
        return render_template("login.html", error=f"Microsoft Login fehlgeschlagen: {e}",
                               totp_enabled=False,
                               google_enabled=bool(cfg.get("oauth_google_client_id")),
                               microsoft_enabled=True)

@app.route("/")
@require_login
def index():
    cfg = load_config()
    caps = get_all_capabilities()
    provider_fields = get_config_fields(cfg.get("provider","ha"))
    resp = make_response(render_template("index.html", cfg=cfg, state=_state,
                           has_template=TEMPLATE_PATH.exists(),
                           all_providers=caps,
                           provider_fields=provider_fields,
                           provider_names={k:v.PROVIDER_NAME for k,v in PROVIDERS.items()},
                           app_version=APP_VERSION,
                           changelog=CHANGELOG,
                           all_vehicles=get_all_vehicles(cfg),
                           current_user=_current_user()))
    resp.headers["Cache-Control"] = "no-cache, no-store, must-revalidate"
    resp.headers["Pragma"]        = "no-cache"
    resp.headers["Expires"]       = "0"
    return resp

@app.route("/api/config", methods=["GET"])
@require_login
def api_get_config(): return jsonify(load_config())

@app.route("/api/config", methods=["POST"])
@require_admin
def api_save_config():
    data=request.json; cfg=load_config()
    floats={"battery_capacity_kwh","price_per_kwh_home","price_per_kwh_ac","price_per_kwh_dc",
            "dc_threshold_kw","entsoe_ac_markup","entsoe_dc_markup","home_radius_m"}
    ints={"poll_interval"}
    for key in DEFAULT_CONFIG:
        if key in data:
            v=data[key]
            if key in floats and v!="": v=float(v)
            elif key in ints: v=int(v)
            cfg[key]=v
    save_config(cfg); return jsonify({"ok":True})

@app.route("/api/providers")
def api_providers(): return jsonify(get_all_capabilities())

@app.route("/api/providers/<provider_id>/fields")
def api_provider_fields(provider_id):
    return jsonify(get_config_fields(provider_id))

@app.route("/api/status")
@require_login
def api_status():
    vid = request.args.get("vehicle_id","v0")
    st  = _vehicle_states.get(vid, _vehicle_states.get("v0", {}))
    result = dict(st)
    result["all_vehicles"] = [
        {"vehicle_id": k, "name": v.get("name",k), "running": v.get("running",False),
         "charging": v.get("charging",False), "session_active": v.get("session_active",False)}
        for k, v in _vehicle_states.items()
    ]
    return jsonify(result)

@app.route("/api/vehicles", methods=["GET"])
def api_get_vehicles():
    return jsonify(get_all_vehicles())

@app.route("/api/vehicles", methods=["POST"])
@require_admin
def api_add_vehicle():
    data = request.json or {}
    cfg  = load_config()
    extras = list(cfg.get("extra_vehicles", []))
    vid = f"v{int(time.time())}"
    data["id"] = vid
    data.setdefault("active", True)
    extras.append(data)
    cfg["extra_vehicles"] = extras
    save_config(cfg)
    if data.get("active", True):
        _start_vehicle_tracker(vid)
    return jsonify({"ok": True, "id": vid})

@app.route("/api/vehicles/<vid>", methods=["PUT"])
@require_admin
def api_update_vehicle(vid):
    if vid == "v0":
        # Update primary vehicle = update flat config fields
        data = request.json or {}
        cfg  = load_config()
        for k, val in data.items():
            if k in VEHICLE_SPECIFIC_KEYS or k == "car_name":
                cfg[k] = val
        save_config(cfg)
        return jsonify({"ok": True})
    data   = request.json or {}
    cfg    = load_config()
    extras = list(cfg.get("extra_vehicles", []))
    for i, v in enumerate(extras):
        if v["id"] == vid:
            was_active = v.get("active", True)
            extras[i] = {**v, **data, "id": vid}
            cfg["extra_vehicles"] = extras
            save_config(cfg)
            now_active = extras[i].get("active", True)
            if was_active:
                _stop_vehicle_tracker(vid)
            if now_active:
                _start_vehicle_tracker(vid)
            return jsonify({"ok": True})
    return jsonify({"error": "Fahrzeug nicht gefunden"}), 404

@app.route("/api/vehicles/<vid>", methods=["DELETE"])
@require_admin
def api_delete_vehicle(vid):
    if vid == "v0":
        return jsonify({"error": "Primärfahrzeug kann nicht gelöscht werden"}), 400
    cfg    = load_config()
    extras = [v for v in cfg.get("extra_vehicles", []) if v["id"] != vid]
    cfg["extra_vehicles"] = extras
    save_config(cfg)
    _stop_vehicle_tracker(vid)
    return jsonify({"ok": True})

@app.route("/api/sessions")
@require_login
def api_sessions():
    return jsonify(get_sessions(
        request.args.get("year",type=int),
        request.args.get("month",type=int),
        request.args.get("location",default="all"),
        request.args.get("vehicle_id",default=None),
    ))

@app.route("/api/sessions/<int:sid>", methods=["DELETE"])
def api_delete_session(sid):
    con=sqlite3.connect(DB_PATH)
    con.execute("DELETE FROM sessions WHERE id=?",(sid,))
    con.execute("DELETE FROM session_points WHERE session_id=?",(sid,))
    con.commit(); con.close()
    return jsonify({"ok":True})

@app.route("/api/sessions/<int:sid>/points")
def api_session_points(sid):
    con=sqlite3.connect(DB_PATH); con.row_factory=sqlite3.Row
    rows=con.execute("SELECT ts,soc,power_kw FROM session_points WHERE session_id=? ORDER BY ts",(sid,)).fetchall()
    con.close(); return jsonify([dict(r) for r in rows])

@app.route("/api/sessions/<int:sid>/location", methods=["POST"])
def api_update_location(sid):
    loc = request.json.get("location","unknown")
    if loc not in ("home","extern","unknown"):
        return jsonify({"ok":False,"error":"Ungültiger Standort"}), 400
    con = sqlite3.connect(DB_PATH)
    con.execute("UPDATE sessions SET location=? WHERE id=?", (loc, sid))
    con.commit(); con.close()
    log.info("Session #%d Standort → %s", sid, loc)
    return jsonify({"ok":True})

@app.route("/api/sessions/<int:sid>/cost", methods=["POST"])
def api_update_cost(sid):
    data=request.json; cost=float(data["cost_eur"]); price_kwh=data.get("price_per_kwh")
    con=sqlite3.connect(DB_PATH); cur=con.cursor()
    if price_kwh is not None:
        row=cur.execute("SELECT kwh_charged FROM sessions WHERE id=?",(sid,)).fetchone()
        if row and row[0]: cost=round(float(row[0])*float(price_kwh),2)
        cur.execute("UPDATE sessions SET cost_eur=?,price_per_kwh=?,cost_manual=1 WHERE id=?",(cost,float(price_kwh),sid))
    else:
        cur.execute("UPDATE sessions SET cost_eur=?,cost_manual=1 WHERE id=?",(cost,sid))
    con.commit(); con.close()
    return jsonify({"ok":True,"cost_eur":cost})

@app.route("/api/stats/monthly")
def api_monthly_stats(): return jsonify(get_monthly_stats())

@app.route("/api/test-connection", methods=["POST"])
def api_test():
    data=request.json; cfg=load_config()
    # merge submitted fields into config for test
    test_cfg={**cfg,**data}
    # use saved token if empty
    if not test_cfg.get("ha_token"): test_cfg["ha_token"]=cfg.get("ha_token","")
    try:
        provider=get_provider(test_cfg.get("provider","ha"), test_cfg)
        result=provider.test_connection()
        return jsonify(result)
    except Exception as e:
        return jsonify({"ok":False,"message":str(e)})

import re as _re_sanitize_url

def _sanitize_url(url):
    if not url:
        return url
    # Remove user:pass@ from URL
    return _re_sanitize_url.sub(r'://([^@]+)@', '://', url)

@app.route("/api/meter/test", methods=["POST"])
def api_meter_test():
    # Load saved config as base
    cfg = load_config()
    # Body-first: merge body values with priority over stored config (without saving)
    body = request.get_json(silent=True) or {}
    for k, v in body.items():
        if k.startswith("meter_") or k in ("ha_url", "ha_token"):
            cfg[k] = v

    result = _read_meter_impl(cfg)

    msg = (f"Zählerstand: {result.value:.3f} kWh" if result.ok
           else result.error or "Kein Wert erhalten")

    return jsonify({
        "ok":             result.ok,
        "value_kwh":      result.value,
        "value":          result.value,   # backward compat
        "message":        msg,
        "provider":       cfg.get("meter_source", ""),
        "endpoint":       _sanitize_url(result.endpoint),
        "raw_value":      result.raw_value,
        "unit":           result.unit,
        "normalized_from": result.normalized_from,
        "debug":          result.debug,
        "suggestions":    result.suggestions or [],
    })

@app.route("/api/entsoe/test", methods=["POST"])
def api_entsoe_test():
    key=request.json.get("entsoe_api_key","").strip()
    if not key: return jsonify({"ok":False,"error":"Kein API Key"})
    _entsoe_cache["price"]=None
    price=fetch_entsoe_spot(key)
    if price: return jsonify({"ok":True,"price_kwh":price,"price_mwh":round(price*1000,2)})
    return jsonify({"ok":False,"error":"Kein Preis erhalten"})

@app.route("/api/template", methods=["POST"])
def api_upload_template():
    if "file" not in request.files: return jsonify({"ok":False,"error":"Keine Datei"}),400
    f=request.files["file"]
    if not f.filename.endswith(".xlsx"): return jsonify({"ok":False,"error":"Nur .xlsx"}),400
    DATA_DIR.mkdir(parents=True,exist_ok=True); f.save(TEMPLATE_PATH)
    cfg = load_config()
    cfg["active_template"] = {"source": "upload", "template_id": None, "name": f.filename}
    save_config(cfg)
    return jsonify({"ok":True,"filename":f.filename,"size":TEMPLATE_PATH.stat().st_size})

@app.route("/api/template", methods=["DELETE"])
def api_delete_template():
    if TEMPLATE_PATH.exists(): TEMPLATE_PATH.unlink()
    cfg = load_config()
    cfg["active_template"] = {"source": None, "template_id": None, "name": None}
    save_config(cfg)
    return jsonify({"ok":True})

@app.route("/api/template/info")
@require_login
def api_template_info():
    cfg = load_config()
    if TEMPLATE_PATH.exists():
        return jsonify({
            "exists": True,
            "size": TEMPLATE_PATH.stat().st_size,
            "modified": datetime.fromtimestamp(TEMPLATE_PATH.stat().st_mtime).isoformat(timespec="seconds"),
            "active_template": cfg.get("active_template") or {},
            "has_signature": SIGNATURE_PATH.exists(),
        })
    return jsonify({"exists": False, "active_template": cfg.get("active_template") or {}, "has_signature": SIGNATURE_PATH.exists()})

@app.route("/api/template/gallery")
@require_login
def api_template_gallery():
    import json as _json
    from pathlib import Path as _Path
    gallery_dir = _Path(__file__).parent / "builtin_templates"
    templates = []
    if gallery_dir.exists():
        for tpl_dir in sorted(gallery_dir.iterdir()):
            manifest_path = tpl_dir / "manifest.json"
            if not manifest_path.exists():
                continue
            try:
                m = _json.loads(manifest_path.read_text(encoding="utf-8"))
                templates.append({
                    "id":              m["id"],
                    "name":            m["name"],
                    "description":     m.get("description", ""),
                    "category":        m.get("category", ""),
                    "recommended_for": m.get("recommended_for", []),
                    "preview_url":     f"/api/template/gallery/{m['id']}/preview",
                })
            except Exception:
                pass
    # indicate which is active
    cfg = load_config()
    active_id = (cfg.get("active_template") or {}).get("template_id")
    for t in templates:
        t["active"] = (t["id"] == active_id)
    return jsonify({"ok": True, "templates": templates})

@app.route("/api/template/gallery/<template_id>/preview")
def api_template_gallery_preview(template_id):
    from pathlib import Path as _Path
    import re
    if not re.match(r'^[a-z0-9_]+$', template_id):
        return "Not found", 404
    svg_path = _Path(__file__).parent / "builtin_templates" / template_id / "preview.svg"
    if not svg_path.exists():
        return "Not found", 404
    return send_file(str(svg_path), mimetype="image/svg+xml")

@app.route("/api/template/gallery/<template_id>/use", methods=["POST"])
@require_admin
def api_template_gallery_use(template_id):
    import json as _json, re
    from pathlib import Path as _Path
    if not re.match(r'^[a-z0-9_]+$', template_id):
        return jsonify({"ok": False, "error": "Ungültige Template-ID"}), 400
    gallery_dir = _Path(__file__).parent / "builtin_templates"
    manifest_path = gallery_dir / template_id / "manifest.json"
    if not manifest_path.exists():
        return jsonify({"ok": False, "error": "Template nicht gefunden"}), 404
    try:
        m = _json.loads(manifest_path.read_text(encoding="utf-8"))
    except Exception as e:
        return jsonify({"ok": False, "error": f"Manifest-Fehler: {e}"}), 500

    # Generate the .xlsx
    try:
        from builtin_template_gen import generate_builtin_template
        ok = generate_builtin_template(template_id, TEMPLATE_PATH)
        if not ok:
            return jsonify({"ok": False, "error": "Template konnte nicht generiert werden"}), 500
    except Exception as e:
        return jsonify({"ok": False, "error": str(e)}), 500

    # Save mapping config
    dm = m.get("default_mapping", {})
    cfg = load_config()
    cfg["template_column_mapping"] = dm.get("column_mapping", {})
    cfg["template_cell_mapping"]   = dm.get("cell_mapping", {})
    cfg["template_start_row"]      = dm.get("start_row")
    cfg["template_sheet"]          = dm.get("sheet", "")
    cfg["signature_mapping"]       = dm.get("signature_mapping", {})
    cfg["active_template"] = {
        "source":      "builtin",
        "template_id": template_id,
        "name":        m["name"],
    }
    save_config(cfg)
    _audit("template_gallery_use", f"id={template_id}", ip=request.remote_addr)
    return jsonify({
        "ok": True,
        "active_template": cfg["active_template"],
        "mapping": {
            "column_mapping": cfg["template_column_mapping"],
            "cell_mapping":   cfg["template_cell_mapping"],
            "start_row":      cfg["template_start_row"],
            "sheet":          cfg["template_sheet"],
        }
    })

@app.route("/api/template/preview")
def api_template_preview():
    if not TEMPLATE_PATH.exists(): return jsonify({"ok":False,"error":"Kein Template"})
    try:
        import openpyxl
        from openpyxl.utils import get_column_letter
        from export_excel import match_column
        wb  = openpyxl.load_workbook(TEMPLATE_PATH, read_only=True, data_only=True)
        ws  = wb.active
        all_rows = []
        max_col  = 0
        raw_rows = list(ws.iter_rows(values_only=True, max_row=40))
        for row in raw_rows:
            last = max((i for i,v in enumerate(row) if v is not None), default=-1)
            if last >= 0: max_col = max(max_col, last + 1)
        max_col = max(max_col, 1)
        for ri, row in enumerate(raw_rows):
            cells = []
            for ci in range(max_col):
                v = row[ci] if ci < len(row) else None
                cells.append(str(v) if v is not None else "")
            all_rows.append({"row": ri + 1, "cells": cells})
        # auto-detect header row (first row with >= 2 filled cells)
        auto_header = None
        for r in all_rows:
            if sum(1 for c in r["cells"] if c.strip()) >= 2:
                auto_header = r["row"]; break
        # build column letters
        col_letters = [get_column_letter(i+1) for i in range(max_col)]
        # build column mapping suggestions from auto-header row
        col_suggestions = {}
        if auto_header:
            hrow = next(r for r in all_rows if r["row"] == auto_header)
            for ci, val in enumerate(hrow["cells"], 1):
                col_suggestions[ci] = {
                    "header": val,
                    "mapped_to": match_column(val) if val else None,
                    "col_letter": get_column_letter(ci),
                }
        wb.close()
        return jsonify({"ok": True, "rows": all_rows, "col_letters": col_letters,
                        "auto_header": auto_header, "col_suggestions": col_suggestions})
    except Exception as e:
        return jsonify({"ok": False, "error": str(e)})

@app.route("/api/template/render")
def api_template_render():
    if not TEMPLATE_PATH.exists(): return jsonify({"ok": False, "error": "Kein Template"})
    try:
        import openpyxl
        from openpyxl.utils import get_column_letter
        wb = openpyxl.load_workbook(TEMPLATE_PATH, data_only=True)
        ws = wb.active

        # merged cell lookup
        merged_topleft = {}   # (r,c) -> (rowspan, colspan)
        merged_skip    = set()
        for rng in ws.merged_cells.ranges:
            merged_topleft[(rng.min_row, rng.min_col)] = (
                rng.max_row - rng.min_row + 1,
                rng.max_col - rng.min_col + 1,
            )
            for r in range(rng.min_row, rng.max_row + 1):
                for c in range(rng.min_col, rng.max_col + 1):
                    if (r, c) != (rng.min_row, rng.min_col):
                        merged_skip.add((r, c))

        max_row = min(ws.max_row or 1, 40)
        max_col = ws.max_column or 1

        def hex_color(color_obj):
            if color_obj is None: return None
            try:
                if color_obj.type == "rgb":
                    rgb = color_obj.rgb
                    if rgb and rgb not in ("00000000", "FF000000", "FFFFFFFF"):
                        return "#" + rgb[-6:]
                if color_obj.type == "theme":
                    return None
            except Exception:
                pass
            return None

        rows = []
        for ri in range(1, max_row + 1):
            cells = []
            for ci in range(1, max_col + 1):
                if (ri, ci) in merged_skip:
                    cells.append(None)
                    continue
                cell = ws.cell(row=ri, column=ci)
                rs, cs = merged_topleft.get((ri, ci), (1, 1))
                bg = bold = fg = None
                try:
                    if cell.fill and cell.fill.patternType not in (None, "none"):
                        bg = hex_color(cell.fill.fgColor)
                except Exception: pass
                try:
                    if cell.font:
                        bold = bool(cell.font.bold)
                        fg   = hex_color(cell.font.color)
                except Exception: pass
                cells.append({
                    "v": str(cell.value) if cell.value is not None else "",
                    "bg": bg, "fg": fg, "bold": bold,
                    "rs": rs, "cs": cs, "r": ri, "c": ci,
                })
            rows.append(rows.__class__() or None)  # placeholder replaced below
            rows[-1] = {"row": ri, "cells": cells}

        # column widths (approximate em)
        col_widths = []
        for ci in range(1, max_col + 1):
            letter = get_column_letter(ci)
            dim = ws.column_dimensions.get(letter)
            w = dim.width if dim and dim.width else 8
            col_widths.append(round(min(max(w, 4), 25), 1))

        col_letters = [get_column_letter(i) for i in range(1, max_col + 1)]
        wb.close()
        return jsonify({"ok": True, "rows": rows, "col_letters": col_letters,
                        "col_widths": col_widths, "max_col": max_col})
    except Exception as e:
        return jsonify({"ok": False, "error": str(e)})

@app.route("/api/sessions/sample")
def api_sessions_sample():
    """Return one recent completed session for live export preview."""
    try:
        con = sqlite3.connect(DB_PATH); con.row_factory = sqlite3.Row
        row = con.execute(
            "SELECT * FROM sessions WHERE end_ts IS NOT NULL ORDER BY start_ts DESC LIMIT 1"
        ).fetchone()
        con.close()
        return jsonify(dict(row) if row else {})
    except Exception:
        return jsonify({})

@app.route("/api/template/analyze")
@require_login
def api_template_analyze():
    if not TEMPLATE_PATH.exists():
        return jsonify({"ok": False, "error": "Kein Template hochgeladen"})
    try:
        from template_analyzer import analyze_template
        result = analyze_template(TEMPLATE_PATH)
        return jsonify(result)
    except Exception as e:
        return jsonify({"ok": False, "error": str(e)})

@app.route("/api/template/mapping", methods=["GET","POST"])
@require_login
def api_template_mapping():
    cfg = load_config()
    if request.method == "POST":
        body = request.get_json(force=True)
        # backward compat: "mapping" → column_mapping
        cfg["template_column_mapping"] = body.get("column_mapping") or body.get("mapping") or {}
        cfg["template_cell_mapping"]   = body.get("cell_mapping", {})
        cfg["template_start_row"]      = body.get("start_row")
        cfg["template_header_row"]     = body.get("header_row")
        cfg["template_sheet"]          = body.get("sheet") or ""
        cfg["signature_mapping"]       = body.get("signature_mapping") or {}
        save_config(cfg)
        return jsonify({"ok": True})
    return jsonify({
        "column_mapping":  cfg.get("template_column_mapping") or cfg.get("template_mapping") or {},
        "cell_mapping":    cfg.get("template_cell_mapping", {}),
        "start_row":       cfg.get("template_start_row"),
        "header_row":      cfg.get("template_header_row"),
        "sheet":           cfg.get("template_sheet", ""),
        "signature_mapping": cfg.get("signature_mapping") or {},
    })


def _normalize_signature_image(img, padding=None):
    """Crop to visible content bbox, then add transparent padding."""
    from PIL import Image as _PILImage, ImageOps as _ImageOps
    cfg = load_config()
    if padding is None:
        padding = int(cfg.get("signature_padding_px", 24))
    img = img.convert("RGBA")
    alpha = img.getchannel("A")
    bbox = alpha.getbbox()
    if bbox:
        img = img.crop(bbox)
    img = _ImageOps.expand(img, border=padding, fill=(255, 255, 255, 0))
    return img


# ── Signature routes ──────────────────────────────────────────────────────────

@app.route("/api/signature")
@require_login
def api_signature_info():
    cfg = load_config()
    sig = cfg.get("signature") or {}
    exists = SIGNATURE_PATH.exists()
    return jsonify({
        "ok": True,
        "has_signature": exists,
        "signature_url": "/api/signature/image" if exists else None,
        "source": sig.get("source"),
        "created_at": sig.get("created_at"),
    })


@app.route("/api/signature/image")
@require_login
def api_signature_image():
    if not SIGNATURE_PATH.exists():
        return jsonify({"error": "Keine Unterschrift"}), 404
    return send_file(str(SIGNATURE_PATH), mimetype="image/png")


@app.route("/api/signature/upload", methods=["POST"])
@require_login
def api_signature_upload():
    if "file" not in request.files:
        return jsonify({"ok": False, "error": "Keine Datei"}), 400
    f = request.files["file"]
    ext = (f.filename or "").rsplit(".", 1)[-1].lower()
    if ext not in ("png", "jpg", "jpeg", "webp"):
        return jsonify({"ok": False, "error": "Nur PNG, JPG oder WebP erlaubt"}), 400
    data = f.read()
    if len(data) > 2 * 1024 * 1024:
        return jsonify({"ok": False, "error": "Datei zu groß (max 2 MB)"}), 400
    try:
        from PIL import Image as _PILImage
        import io as _io
        img = _PILImage.open(_io.BytesIO(data))
        img = _normalize_signature_image(img)
        SIGNATURE_DIR.mkdir(parents=True, exist_ok=True)
        img.save(str(SIGNATURE_PATH), "PNG")
    except Exception as e:
        return jsonify({"ok": False, "error": f"Bildverarbeitung fehlgeschlagen: {e}"}), 500
    cfg = load_config()
    cfg["signature"] = {"source": "upload", "created_at": datetime.utcnow().isoformat()}
    save_config(cfg)
    _audit("signature_upload", ip=request.remote_addr)
    return jsonify({"ok": True})


@app.route("/api/signature/draw", methods=["POST"])
@require_login
def api_signature_draw():
    import base64 as _b64, io as _io
    body = request.get_json(force=True) or {}
    image_data = body.get("image_data", "")
    if not image_data or "base64," not in image_data:
        return jsonify({"ok": False, "error": "Ungültige Bilddaten"}), 400
    try:
        raw = _b64.b64decode(image_data.split("base64,", 1)[1])
        from PIL import Image as _PILImage
        img = _PILImage.open(_io.BytesIO(raw))
        img = _normalize_signature_image(img)
        SIGNATURE_DIR.mkdir(parents=True, exist_ok=True)
        img.save(str(SIGNATURE_PATH), "PNG")
    except Exception as e:
        return jsonify({"ok": False, "error": f"Fehler beim Speichern: {e}"}), 500
    cfg = load_config()
    cfg["signature"] = {"source": "draw", "created_at": datetime.utcnow().isoformat()}
    save_config(cfg)
    _audit("signature_draw", ip=request.remote_addr)
    return jsonify({"ok": True})


@app.route("/api/signature", methods=["DELETE"])
@require_login
def api_signature_delete():
    if SIGNATURE_PATH.exists():
        SIGNATURE_PATH.unlink()
    cfg = load_config()
    cfg["signature"] = {"source": None, "created_at": None}
    save_config(cfg)
    _audit("signature_delete", ip=request.remote_addr)
    return jsonify({"ok": True})

@app.route("/api/export")
@require_login
def api_export():
    user = _current_user()
    if not has_permission(user, "export:create"):
        return jsonify({"ok": False, "error": "Keine Berechtigung: export:create"}), 403
    import io as _io_exp
    from export_excel import export
    y=request.args.get("year",datetime.now().year,type=int)
    m=request.args.get("month",datetime.now().month,type=int)
    loc=request.args.get("location","all")
    override=json.loads(request.args.get("col_override","null") or "null")
    cfg = load_config()
    if override is None:
        saved = cfg.get("template_column_mapping") or cfg.get("template_mapping") or {}
        if isinstance(saved, dict) and saved:
            override = {k: v for k, v in saved.items() if v}
        else:
            override = None
    start_row    = cfg.get("template_start_row")
    header_row   = cfg.get("template_header_row")
    # Backward compat: template_start_row without header_row
    if start_row and not header_row:
        try:
            header_row = int(start_row) - 1
        except (ValueError, TypeError):
            pass
    lang         = request.args.get("lang") or cfg.get("export_language", "de")
    _raw_cm      = cfg.get("template_cell_mapping") or {}
    cell_mapping = _raw_cm if isinstance(_raw_cm, dict) else {}
    sheet        = cfg.get("template_sheet") or None
    header_info = {
        "fahrer":            cfg.get("template_fahrer", ""),
        "kennzeichen":       cfg.get("template_kennzeichen", ""),
        "abteilung":         cfg.get("template_abteilung", ""),
        "kostenstelle":      cfg.get("template_kostenstelle", ""),
        "price_per_kwh":     cfg.get("price_per_kwh_home", 0.30),
        "meter_start_value": cfg.get("template_meter_start", 0.0),
    }
    include_sig_param = request.args.get("include_signature")
    if include_sig_param is not None:
        include_signature = include_sig_param.lower() == "true"
    else:
        include_signature = bool(cfg.get("export_include_signature", False))
    sig_mapping = cfg.get("signature_mapping") or {}
    # Backward compat: "cell" → "anchor_cell" in signature_mapping
    if sig_mapping and "cell" in sig_mapping and "anchor_cell" not in sig_mapping:
        sig_mapping = dict(sig_mapping)
        sig_mapping["anchor_cell"] = sig_mapping["cell"]
    try:
        xlsx_bytes = export(y, m, loc, col_override=override, start_row=start_row, header_row=header_row,
                      header_info=header_info,
                      cell_mapping=cell_mapping, sheet=sheet,
                      include_signature=include_signature,
                      signature_path=str(SIGNATURE_PATH) if SIGNATURE_PATH.exists() else None,
                      signature_mapping=sig_mapping, lang=lang)
        filename = f"EV_Ladeprotokoll_{y:04d}-{m:02d}.xlsx"
        return send_file(_io_exp.BytesIO(xlsx_bytes), as_attachment=True,
                         download_name=filename,
                         mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
    except Exception as e:
        log.exception("Export fehlgeschlagen")
        return jsonify({"ok": False, "error": str(e)}), 500

# ── Export token store ────────────────────────────────────────────────────────
_export_tokens: dict = {}  # token -> {"path": str, "expires": float}

def _cleanup_export_tokens():
    """Delete expired token entries and their temp files."""
    now = time.time()
    expired = [k for k, v in list(_export_tokens.items()) if v["expires"] < now]
    for k in expired:
        info = _export_tokens.pop(k, None)
        if info:
            try:
                Path(info["path"]).unlink(missing_ok=True)
            except Exception:
                pass
    # Also clean up orphaned /tmp/ev_export_*.xlsx
    import glob as _glob
    for fp in _glob.glob("/tmp/ev_export_*.xlsx"):
        # only delete if not referenced by any token
        if not any(v["path"] == fp for v in _export_tokens.values()):
            try:
                if Path(fp).stat().st_mtime < now - 3600:
                    Path(fp).unlink(missing_ok=True)
            except Exception:
                pass

@app.route("/api/export/preview", methods=["POST"])
@require_login
def api_export_preview():
    user = _current_user()
    if not has_permission(user, "export:preview"):
        return jsonify({"ok": False, "error": "Keine Berechtigung: export:preview"}), 403
    import io as _io_prev
    import openpyxl as _opxl_prev
    from export_excel import export as _export_func
    body = request.get_json(silent=True) or {}
    y    = int(body.get("year",  datetime.now().year))
    m    = int(body.get("month", datetime.now().month))
    loc  = body.get("location", "all")
    cfg  = load_config()
    lang = body.get("lang") or cfg.get("export_language", "de")

    _raw_override = body.get("col_override") or cfg.get("template_column_mapping") or {}
    override     = _raw_override if isinstance(_raw_override, dict) else {}
    start_row    = body.get("start_row") or cfg.get("template_start_row")
    header_row   = body.get("header_row") or cfg.get("template_header_row")
    if start_row and not header_row:
        try:
            header_row = int(start_row) - 1
        except (ValueError, TypeError):
            pass
    _raw_cm      = body.get("cell_mapping") or cfg.get("template_cell_mapping") or {}
    cell_mapping = _raw_cm if isinstance(_raw_cm, dict) else {}
    sheet        = body.get("sheet") or cfg.get("template_sheet")
    header_info  = {
        "fahrer":            cfg.get("template_fahrer", ""),
        "kennzeichen":       cfg.get("template_kennzeichen", ""),
        "abteilung":         cfg.get("template_abteilung", ""),
        "kostenstelle":      cfg.get("template_kostenstelle", ""),
        "price_per_kwh":     cfg.get("price_per_kwh_home", 0.30),
        "meter_start_value": cfg.get("template_meter_start", 0.0),
    }
    include_signature = bool(body.get("include_signature", False))
    sig_mapping       = cfg.get("signature_mapping") or {}
    if sig_mapping and "cell" in sig_mapping and "anchor_cell" not in sig_mapping:
        sig_mapping = dict(sig_mapping)
        sig_mapping["anchor_cell"] = sig_mapping["cell"]

    # Cleanup old tokens first
    _cleanup_export_tokens()

    try:
        xlsx_bytes, warnings = _export_func(
            y, m, loc,
            col_override=override, start_row=start_row, header_row=header_row,
            header_info=header_info, cell_mapping=cell_mapping, sheet=sheet,
            include_signature=include_signature,
            signature_path=str(SIGNATURE_PATH) if SIGNATURE_PATH.exists() else None,
            signature_mapping=sig_mapping, lang=lang,
            return_warnings=True,
        )
    except Exception as e:
        log.exception("Export-Vorschau fehlgeschlagen")
        return jsonify({"ok": False, "error": str(e)}), 500

    # Save to temp file and issue download token
    token = secrets.token_urlsafe(16)
    tmp_path = f"/tmp/ev_export_{token}.xlsx"
    try:
        with open(tmp_path, "wb") as _tf:
            _tf.write(xlsx_bytes)
        _export_tokens[token] = {"path": tmp_path, "expires": time.time() + 1800}
    except Exception as e:
        log.warning(f"Konnte Token-Datei nicht speichern: {e}")
        token = None

    # Determine data_start_row from config
    _template_config = cfg.get("template_config") or {}
    _data_start_row = int(start_row) if start_row else int(_template_config.get("start_row", 1))

    # Build grid from xlsx_bytes
    sheets_out = []
    try:
        import datetime as _dt_prev
        wb_prev = _opxl_prev.load_workbook(_io_prev.BytesIO(xlsx_bytes), data_only=True)
        for ws_p in wb_prev.worksheets:
            rows_out = []
            for row_idx_0, row_p in enumerate(ws_p.iter_rows(max_row=200, max_col=30, values_only=True)):
                row_idx = row_idx_0 + 1  # 1-based
                cells = []
                for val in row_p:
                    if val is None:
                        cells.append("")
                    elif isinstance(val, (_dt_prev.datetime, _dt_prev.date)):
                        cells.append(val.strftime("%d.%m.%Y %H:%M") if isinstance(val, _dt_prev.datetime) else val.strftime("%d.%m.%Y"))
                    else:
                        cells.append(str(val))
                rows_out.append({
                    "row": row_idx,
                    "is_data": row_idx >= _data_start_row,
                    "cells": cells,
                })
            sheets_out.append({
                "name": ws_p.title,
                "data_start_row": _data_start_row,
                "rows": rows_out,
            })
    except Exception as e:
        warnings.append(f"Grid-Erzeugung fehlgeschlagen: {e}")

    result = {
        "ok":             True,
        "sheets":         sheets_out,
        "warnings":       warnings,
        "download_token": token,
    }
    return jsonify(result)


@app.route("/api/export/download/<token>")
@require_login
def api_export_download_token(token):
    """Download a previously generated export XLSX by token."""
    import re as _re_tok
    # Validate token format (URL-safe base64)
    if not _re_tok.match(r'^[A-Za-z0-9_-]{10,64}$', token):
        return jsonify({"error": "Ungültiger Token"}), 404
    info = _export_tokens.get(token)
    if not info:
        return jsonify({"error": "Token nicht gefunden oder abgelaufen"}), 404
    if info["expires"] < time.time():
        _export_tokens.pop(token, None)
        try:
            Path(info["path"]).unlink(missing_ok=True)
        except Exception:
            pass
        return jsonify({"error": "Token abgelaufen"}), 404
    file_path = Path(info["path"])
    if not file_path.exists():
        return jsonify({"error": "Datei nicht mehr vorhanden"}), 404
    return send_file(
        str(file_path),
        as_attachment=True,
        download_name=file_path.name.replace(f"ev_export_{token}", "EV_Export"),
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )

import calendar as _calendar

# ── Backup ────────────────────────────────────────────────────────────────────
import zipfile, subprocess
from threading import Timer

_backup_timer = None

def create_backup(label="manual"):
    BACKUP_DIR.mkdir(parents=True,exist_ok=True)
    ts=datetime.now().strftime("%Y%m%d_%H%M%S")
    out=BACKUP_DIR/f"ev-tracker_backup_{label}_{ts}.zip"
    with zipfile.ZipFile(out,"w",zipfile.ZIP_DEFLATED) as zf:
        for item in DATA_DIR.rglob("*"):
            if BACKUP_DIR in item.parents or item==out: continue
            zf.write(item,item.relative_to(DATA_DIR))
    all_backups=sorted(BACKUP_DIR.glob("*.zip"),key=lambda p:p.stat().st_mtime)
    while len(all_backups)>10: all_backups.pop(0).unlink()
    return out

def restore_backup(zip_path):
    with zipfile.ZipFile(zip_path,"r") as zf:
        for member in zf.namelist():
            if member.startswith("backups/"): continue
            zf.extract(member,DATA_DIR)

def parse_cron_next(cron_expr):
    now=datetime.now(); expr=cron_expr.strip().lower()
    if expr in ("@daily","0 0 * * *"):
        nxt=now.replace(hour=0,minute=0,second=0,microsecond=0)+timedelta(days=1)
    elif expr in ("@weekly","0 0 * * 0"):
        days=(6-now.weekday())%7 or 7
        nxt=now.replace(hour=0,minute=0,second=0,microsecond=0)+timedelta(days=days)
    elif expr in ("@monthly","0 0 1 * *"):
        nxt=now.replace(month=now.month%12+1,day=1,hour=0,minute=0,second=0,microsecond=0)
    else:
        try:
            parts=expr.split()
            if len(parts)>=2:
                nxt=now.replace(hour=int(parts[1]),minute=int(parts[0]),second=0,microsecond=0)
                if nxt<=now: nxt+=timedelta(days=1)
            else: return None
        except: return None
    return (nxt-now).total_seconds()

def schedule_backup():
    global _backup_timer
    cfg=load_config(); cron=cfg.get("backup_cron","").strip()
    if not cron: return
    secs=parse_cron_next(cron)
    if not secs or secs<=0: return
    def run():
        try: create_backup("auto"); log.info("Auto-Backup OK")
        except Exception as e: log.warning("Auto-Backup Fehler: %s",e)
        schedule_backup()
    _backup_timer=Timer(secs,run); _backup_timer.daemon=True; _backup_timer.start()
    log.info("Nächstes Backup: %s",(datetime.now()+timedelta(seconds=secs)).strftime("%d.%m.%Y %H:%M"))

@app.route("/api/backup/list")
def api_backup_list():
    BACKUP_DIR.mkdir(parents=True,exist_ok=True)
    backups=[{"name":f.name,"size":f.stat().st_size,
              "modified":datetime.fromtimestamp(f.stat().st_mtime).isoformat(timespec="seconds")}
             for f in sorted(BACKUP_DIR.glob("*.zip"),key=lambda p:p.stat().st_mtime,reverse=True)]
    cfg=load_config(); cron=cfg.get("backup_cron","")
    next_backup=None
    if cron:
        secs=parse_cron_next(cron)
        if secs: next_backup=(datetime.now()+timedelta(seconds=secs)).strftime("%d.%m.%Y %H:%M")
    return jsonify({"backups":backups,"next_backup":next_backup,"cron":cron})

@app.route("/api/backup/create",methods=["POST"])
def api_backup_create():
    try: out=create_backup("manual"); return jsonify({"ok":True,"name":out.name,"size":out.stat().st_size})
    except Exception as e: return jsonify({"ok":False,"error":str(e)})

@app.route("/api/backup/download/<filename>")
def api_backup_download(filename):
    if ".." in filename or "/" in filename: return jsonify({"error":"ungültig"}),400
    path=BACKUP_DIR/filename
    if not path.exists(): return jsonify({"error":"nicht gefunden"}),404
    return send_file(path,as_attachment=True)

@app.route("/api/backup/restore",methods=["POST"])
def api_backup_restore():
    user = _current_user()
    if not has_permission(user, "backup:restore"):
        return jsonify({"ok": False, "error": "Keine Berechtigung: backup:restore"}), 403
    name=request.json.get("name","")
    if ".." in name or "/" in name: return jsonify({"ok":False,"error":"ungültig"}),400
    path=BACKUP_DIR/name
    if not path.exists(): return jsonify({"ok":False,"error":"nicht gefunden"}),404
    try: restore_backup(path); return jsonify({"ok":True})
    except Exception as e: return jsonify({"ok":False,"error":str(e)})

@app.route("/api/backup/upload",methods=["POST"])
def api_backup_upload():
    if "file" not in request.files: return jsonify({"ok":False,"error":"Keine Datei"}),400
    f=request.files["file"]
    if not f.filename.endswith(".zip"): return jsonify({"ok":False,"error":"Nur .zip"}),400
    BACKUP_DIR.mkdir(parents=True,exist_ok=True)
    tmp=BACKUP_DIR/f"upload_{datetime.now().strftime('%Y%m%d_%H%M%S')}.zip"; f.save(tmp)
    try: restore_backup(tmp); return jsonify({"ok":True,"restored":tmp.name})
    except Exception as e: tmp.unlink(missing_ok=True); return jsonify({"ok":False,"error":str(e)})

@app.route("/api/backup/cron",methods=["POST"])
def api_backup_cron():
    global _backup_timer
    cron=request.json.get("cron","").strip()
    cfg=load_config(); cfg["backup_cron"]=cron; save_config(cfg)
    if _backup_timer: _backup_timer.cancel()
    if cron:
        schedule_backup()
        secs=parse_cron_next(cron)
        nxt=(datetime.now()+timedelta(seconds=secs)).strftime("%d.%m.%Y %H:%M") if secs else "?"
        return jsonify({"ok":True,"next":nxt})
    return jsonify({"ok":True,"next":None})

@app.route("/api/backup/delete/<filename>",methods=["DELETE"])
def api_backup_delete(filename):
    if ".." in filename or "/" in filename: return jsonify({"ok":False}),400
    path=BACKUP_DIR/filename
    if path.exists(): path.unlink()
    return jsonify({"ok":True})

# ── Update via Docker Hub ────────────────────────────────────────────────────
DOCKER_HUB_REPO = "19121412/ev-tracker"
DOCKER_SOCKET   = "/var/run/docker.sock"

def get_dockerhub_digest(tag: str) -> str | None:
    """Fetch latest digest from Docker Hub for given tag."""
    try:
        # Get token
        r = requests.get(
            f"https://auth.docker.io/token?service=registry.docker.io&scope=repository:{DOCKER_HUB_REPO}:pull",
            timeout=10)
        token = r.json().get("token")
        # Get manifest digest
        r2 = requests.get(
            f"https://registry-1.docker.io/v2/{DOCKER_HUB_REPO}/manifests/{tag}",
            headers={"Authorization": f"Bearer {token}",
                     "Accept": "application/vnd.docker.distribution.manifest.v2+json"},
            timeout=10)
        return r2.headers.get("Docker-Content-Digest","")
    except Exception as e:
        log.warning("Docker Hub check error: %s", e)
        return None

def get_local_digest(tag: str) -> str | None:
    """Get local image digest via Docker socket."""
    if not os.path.exists(DOCKER_SOCKET):
        return None
    try:
        import socket, json as _json
        sock = socket.socket(socket.AF_UNIX, socket.SOCK_STREAM)
        sock.connect(DOCKER_SOCKET)
        req = f"GET /images/{DOCKER_HUB_REPO}:{tag}/json HTTP/1.0\r\nHost: localhost\r\n\r\n"
        sock.send(req.encode())
        resp = b""
        while chunk := sock.recv(4096): resp += chunk
        sock.close()
        body = resp.split(b"\r\n\r\n", 1)[-1]
        data = _json.loads(body)
        digests = data.get("RepoDigests", [])
        for d in digests:
            if "@" in d: return d.split("@")[1]
        return None
    except Exception as e:
        log.warning("Docker socket error: %s", e)
        return None

def docker_socket_request(method: str, path: str, body: dict = None) -> dict:
    """Make HTTP request to Docker socket using HTTP/1.0 so connection closes after response."""
    import socket as _socket, json as _json
    sock = _socket.socket(_socket.AF_UNIX, _socket.SOCK_STREAM)
    sock.settimeout(30)
    sock.connect(DOCKER_SOCKET)
    body_bytes = _json.dumps(body).encode() if body else b""
    headers = (
        f"{method} {path} HTTP/1.0\r\n"
        f"Host: localhost\r\n"
        f"Content-Type: application/json\r\n"
        f"Content-Length: {len(body_bytes)}\r\n"
        f"\r\n"
    )
    sock.send(headers.encode() + body_bytes)
    resp = b""
    while True:
        chunk = sock.recv(4096)
        if not chunk: break
        resp += chunk
    sock.close()
    header, _, body_raw = resp.partition(b"\r\n\r\n")
    status = int(header.split(b" ")[1])
    try: data = _json.loads(body_raw)
    except: data = {}
    return {"status": status, "data": data}

_update_log: list[str] = []
_update_running = False

def _ulog(msg: str):
    _update_log.append(msg)
    if len(_update_log) > 200:
        _update_log.pop(0)
    log.info("Update: %s", msg)

def docker_pull_and_restart(tag: str):
    """Pull new image in background, then stop/remove/recreate container."""
    global _update_running
    if not os.path.exists(DOCKER_SOCKET):
        return False, "Docker Socket nicht gefunden"

    # Find container by image name (works regardless of container name on Unraid)
    try:
        containers = docker_socket_request("GET", "/containers/json?all=1")
        container_id = None
        container_name = None
        clist = containers.get("data", [])
        if not isinstance(clist, list):
            clist = []
        for c in clist:
            img = c.get("Image", "")
            if DOCKER_HUB_REPO in img:
                container_id = c["Id"]
                names = c.get("Names", [])
                container_name = names[0].lstrip("/") if names else "ev-tracker"
                break
        if not container_id:
            # Fallback: search by container name keywords
            for c in clist:
                names = c.get("Names", [])
                if any("ev" in n.lower() and "track" in n.lower() for n in names):
                    container_id = c["Id"]
                    container_name = names[0].lstrip("/") if names else "ev-tracker"
                    break
        if not container_id:
            return False, f"Container für Image '{DOCKER_HUB_REPO}' nicht gefunden — {len(clist)} Container auf dem Host"

        inspect = docker_socket_request("GET", f"/containers/{container_id}/json")
        if inspect["status"] != 200:
            return False, "Container-Konfiguration nicht lesbar"

        old_cfg     = inspect["data"]
        host_config = old_cfg.get("HostConfig", {})
        env         = old_cfg.get("Config", {}).get("Env", [])
        labels      = old_cfg.get("Config", {}).get("Labels", {})
    except Exception as e:
        return False, str(e)

    _update_running = True
    _update_log.clear()

    def pull_and_recreate():
        global _update_running
        import time as _time
        import socket as _socket, json as _json
        _ulog(f"Starte Pull für {DOCKER_HUB_REPO}:{tag}")
        try:
            # Use HTTP/1.0 so Docker closes connection after response (avoids keep-alive hang)
            sock = _socket.socket(_socket.AF_UNIX, _socket.SOCK_STREAM)
            sock.settimeout(600)
            sock.connect(DOCKER_SOCKET)
            req = (f"POST /images/create?fromImage={DOCKER_HUB_REPO}&tag={tag} HTTP/1.0\r\n"
                   f"Host: localhost\r\nContent-Length: 0\r\n\r\n")
            sock.send(req.encode())
            buf = b""
            while True:
                chunk = sock.recv(8192)
                if not chunk: break
                buf += chunk
            sock.close()
            # parse status from first line: "HTTP/1.0 200 OK"
            first_line = buf.split(b"\r\n")[0]
            parts = first_line.split(b" ")
            status = int(parts[1]) if len(parts) > 1 else 0
            _ulog(f"Pull HTTP-Status {status}")
            if status not in (200, 204):
                _ulog(f"FEHLER: Pull fehlgeschlagen — HTTP {status}")
                _update_running = False
                return
            # Check response body for error JSON lines
            body_part = buf.split(b"\r\n\r\n", 1)[-1].decode(errors="replace")
            for line in body_part.splitlines():
                line = line.strip()
                if not line:
                    continue
                try:
                    j = _json.loads(line)
                    if "error" in j:
                        _ulog(f"FEHLER vom Docker-Daemon: {j['error']}")
                        _update_running = False
                        return
                    if "status" in j:
                        _ulog(j["status"])
                except Exception:
                    pass
        except Exception as e:
            _ulog(f"FEHLER beim Pull: {e}")
            _update_running = False
            return

        _ulog("Pull abgeschlossen — baue Recreate-Konfiguration")

        # Build full create body from inspect so ports/volumes/networks are preserved
        old_config = old_cfg.get("Config", {})
        nets       = old_cfg.get("NetworkSettings", {}).get("Networks", {})
        clean_nets = {}
        for net_name, net_cfg in nets.items():
            clean_nets[net_name] = {
                k: v for k, v in net_cfg.items()
                if k in ("IPAMConfig", "Links", "Aliases", "NetworkID", "EndpointID",
                         "Gateway", "IPAddress", "IPPrefixLen", "IPv6Gateway",
                         "GlobalIPv6Address", "GlobalIPv6PrefixLen", "MacAddress",
                         "DriverOpts")
            }
        create_body = {
            "Image":        f"{DOCKER_HUB_REPO}:{tag}",
            "Env":          old_config.get("Env", []),
            "Labels":       old_config.get("Labels", {}),
            "ExposedPorts": old_config.get("ExposedPorts", {}),
            "Volumes":      old_config.get("Volumes", {}),
            "WorkingDir":   old_config.get("WorkingDir", ""),
            "HostConfig":   host_config,
            "NetworkingConfig": {"EndpointsConfig": clean_nets},
        }
        if old_config.get("Cmd"):        create_body["Cmd"]        = old_config["Cmd"]
        if old_config.get("Entrypoint"): create_body["Entrypoint"] = old_config["Entrypoint"]

        # --- Helper-Container-Ansatz ---
        # Wir können Stop/Remove/Recreate NICHT aus unserem eigenen Thread heraus
        # ausführen, da der Thread stirbt sobald Docker unseren Container stoppt.
        # Lösung: Helper-Container aus dem neu gezogenen Image starten, der die
        # Sequenz unabhängig von unserem Prozess durchführt.
        import base64 as _b64
        create_b64 = _b64.b64encode(_json.dumps(create_body).encode()).decode()

        helper_py = (
            "import socket,json,time,os,base64\n"
            "SOCK='/var/run/docker.sock'\n"
            "def req(m,p,b=None):\n"
            "  s=socket.socket(socket.AF_UNIX,socket.SOCK_STREAM);s.connect(SOCK)\n"
            "  d=json.dumps(b).encode() if b else b''\n"
            "  h=f'{m} {p} HTTP/1.0\\r\\nHost: localhost\\r\\n'\n"
            "  if d: h+=f'Content-Type: application/json\\r\\nContent-Length: {len(d)}\\r\\n'\n"
            "  s.send((h+'\\r\\n').encode()+d)\n"
            "  r=b''\n"
            "  while True:\n"
            "    c=s.recv(8192)\n"
            "    if not c: break\n"
            "    r+=c\n"
            "  s.close()\n"
            "  st=int(r.split(b' ')[1]);body=r.split(b'\\r\\n\\r\\n',1)[-1]\n"
            "  try: return st,json.loads(body)\n"
            "  except: return st,{}\n"
            "time.sleep(5)\n"
            "cid=os.environ['OLD_CID'];cname=os.environ['NEW_NAME']\n"
            "cb=json.loads(base64.b64decode(os.environ['CB64']).decode())\n"
            "req('POST',f'/containers/{cid}/stop?t=10')\n"
            "time.sleep(4)\n"
            "req('DELETE',f'/containers/{cid}?force=1')\n"
            "time.sleep(1)\n"
            "st,resp=req('POST',f'/containers/create?name={cname}',cb)\n"
            "nid=resp.get('Id') if isinstance(resp,dict) else None\n"
            "if nid: req('POST',f'/containers/{nid}/start')\n"
            "req('DELETE','/containers/ev-tracker-updater?force=1')\n"
        )

        # Vorhandenen alten Helper-Container bereinigen
        docker_socket_request("DELETE", "/containers/ev-tracker-updater?force=1")

        helper_resp = docker_socket_request("POST", "/containers/create?name=ev-tracker-updater", {
            "Image":      f"{DOCKER_HUB_REPO}:{tag}",
            "Entrypoint": [],
            "Cmd":        ["python3", "-c", helper_py],
            "Env":        [f"OLD_CID={container_id}",
                           f"NEW_NAME={container_name}",
                           f"CB64={create_b64}"],
            "HostConfig": {"Binds": [f"{DOCKER_SOCKET}:{DOCKER_SOCKET}"],
                           "AutoRemove": True},
        })
        helper_id = (helper_resp.get("data") or {}).get("Id")
        if helper_id:
            start = docker_socket_request("POST", f"/containers/{helper_id}/start")
            if start["status"] in (200, 204, 304):
                _ulog(f"Helper-Container gestartet ({helper_id[:12]}) — "
                      f"Container '{container_name}' wird in ~10s auf {tag} neu gestartet")
            else:
                _ulog(f"FEHLER: Helper-Container konnte nicht gestartet werden: "
                      f"HTTP {start['status']} — {start.get('data','')}")
        else:
            err = (helper_resp.get("data") or {}).get("message") or str(helper_resp.get("data",""))
            _ulog(f"FEHLER: Helper-Container konnte nicht erstellt werden: {err}")
        _update_running = False

    threading.Thread(target=pull_and_recreate, daemon=True).start()
    return True, "Update läuft im Hintergrund · Seite lädt neu sobald der Container bereit ist"

def get_update_info():
    cfg = load_config()
    tag = cfg.get("update_channel", "latest")
    remote = get_dockerhub_digest(tag)
    local  = get_local_digest(tag)
    if not remote:
        return {"ok": False, "error": "Docker Hub nicht erreichbar"}
    up_to_date = (remote == local) if local else False
    has_socket = os.path.exists(DOCKER_SOCKET)
    return {
        "ok":          True,
        "up_to_date":  up_to_date,
        "local_digest": (local or "unbekannt")[:19],
        "remote_digest": remote[:19],
        "tag":         tag,
        "has_socket":  has_socket,
        "update_count": 0 if up_to_date else 1,
    }

def fetch_remote_version(tag: str) -> dict:
    """Fetch version.json from GitHub to get remote version number and changelog."""
    try:
        url = f"https://raw.githubusercontent.com/fdreckmann/ev_tracker/{'main' if tag == 'latest' else 'dev'}/version.json"
        r = requests.get(url, timeout=8)
        if r.status_code == 200:
            return r.json()
    except Exception:
        pass
    return {}

@app.route("/api/update/check")
def api_update_check():
    info = get_update_info()
    if info.get("ok") and not info.get("up_to_date"):
        tag = info.get("tag", "latest")
        remote_ver = fetch_remote_version(tag)
        info["remote_version"] = remote_ver.get("version", "")
        info["remote_changelog"] = remote_ver.get("changelog", [])
    return jsonify(info)

@app.route("/api/update/pull", methods=["POST"])
@require_admin
def api_update_pull():
    user = _current_user()
    if not has_permission(user, "updates:start"):
        return jsonify({"ok": False, "error": "Keine Berechtigung: updates:start"}), 403
    cfg = load_config()
    tag = cfg.get("update_channel","latest")
    ok, msg = docker_pull_and_restart(tag)
    return jsonify({"ok": ok, "output": msg, "restarting": ok})

@app.route("/api/update/log")
def api_update_log():
    return jsonify({"running": _update_running, "lines": list(_update_log)})

# ── User Management ───────────────────────────────────────────────────────────
@app.route("/api/users", methods=["GET"])
@require_admin
def api_get_users():
    con = _get_db()
    rows = con.execute(
        "SELECT id,name,email,role,status,totp_enabled,created_at,updated_at,last_login_at FROM users ORDER BY id"
    ).fetchall()
    con.close()
    return jsonify([dict(r) for r in rows])

@app.route("/api/users", methods=["POST"])
@require_admin
def api_create_user():
    data  = request.json or {}
    name  = (data.get("name") or "").strip()
    email = (data.get("email") or "").strip().lower()
    pw    = data.get("password") or ""
    role  = data.get("role","user")
    invite_mode = not pw  # no password → create as invited
    if not name or not email:
        return jsonify({"ok": False, "error": "Name und E-Mail erforderlich"})
    now = datetime.utcnow().isoformat()
    try:
        con = _get_db()
        if invite_mode:
            con.execute(
                "INSERT INTO users (name,email,password_hash,role,status,created_at,updated_at) VALUES (?,?,?,?,?,?,?)",
                (name, email, "", role, "invited", now, now))
        else:
            pw_err = _password_ok(pw)
            if pw_err:
                return jsonify({"ok": False, "error": pw_err})
            con.execute(
                "INSERT INTO users (name,email,password_hash,role,status,created_at,updated_at) VALUES (?,?,?,?,?,?,?)",
                (name, email, _hash_password(pw), role, "active", now, now))
        con.commit(); con.close()
    except sqlite3.IntegrityError:
        return jsonify({"ok": False, "error": "E-Mail bereits vorhanden"})
    _audit("user_created", f"email={email} role={role} invited={invite_mode}", ip=request.remote_addr)
    return jsonify({"ok": True, "invited": invite_mode})

@app.route("/api/users/<int:uid>", methods=["PUT"])
@require_admin
def api_update_user(uid):
    data = request.json or {}
    user = _get_user_by_id(uid)
    if not user:
        return jsonify({"ok": False, "error": "Nicht gefunden"}), 404
    now = datetime.utcnow().isoformat()
    con = _get_db()
    for field in ("name","role","status"):
        if field in data:
            con.execute(f"UPDATE users SET {field}=?,updated_at=? WHERE id=?", (data[field], now, uid))
    if data.get("password") and len(data["password"]) >= 8:
        con.execute("UPDATE users SET password_hash=?,updated_at=? WHERE id=?",
                    (_hash_password(data["password"]), now, uid))
    con.commit(); con.close()
    _audit("user_updated", f"uid={uid}", ip=request.remote_addr)
    return jsonify({"ok": True})

@app.route("/api/users/<int:uid>", methods=["DELETE"])
@require_admin
def api_delete_user(uid):
    if uid == session.get("user_id"):
        return jsonify({"ok": False, "error": "Eigenen Account nicht löschbar"})
    con = _get_db()
    con.execute("DELETE FROM users WHERE id=?", (uid,))
    con.commit(); con.close()
    _audit("user_deleted", f"uid={uid}", ip=request.remote_addr)
    return jsonify({"ok": True})

@app.route("/api/users/<int:uid>/reset-2fa", methods=["POST"])
@require_admin
def api_admin_reset_2fa(uid):
    now = datetime.utcnow().isoformat()
    con = _get_db()
    con.execute("UPDATE users SET totp_secret='',totp_enabled=0,updated_at=? WHERE id=?", (now, uid))
    con.commit(); con.close()
    _audit("totp_reset", f"uid={uid}", ip=request.remote_addr)
    return jsonify({"ok": True})

@app.route("/api/users/<int:uid>/invite", methods=["POST"])
@require_admin
def api_invite_user(uid):
    user = _get_user_by_id(uid)
    if not user:
        return jsonify({"ok": False, "error": "Benutzer nicht gefunden"}), 404
    token = secrets.token_urlsafe(32)
    token_hash = hashlib.sha256(token.encode()).hexdigest()
    now_dt = datetime.utcnow()
    expires = (now_dt + timedelta(hours=48)).isoformat()
    con = _get_db()
    # Invalidate old invite tokens
    con.execute("UPDATE invite_tokens SET used_at=? WHERE user_id=? AND used_at IS NULL",
                (now_dt.isoformat(), uid))
    con.execute("INSERT INTO invite_tokens (user_id,token_hash,expires_at,created_at) VALUES (?,?,?,?)",
                (uid, token_hash, expires, now_dt.isoformat()))
    con.commit(); con.close()
    invite_url = request.host_url.rstrip("/") + url_for("accept_invite_page", token=token)
    body_html = _email_html(
        "Einladung zu EV Tracker",
        f"Hallo {user['name']},",
        "du wurdest zu EV Tracker eingeladen. Klicke auf den Button, um dein Passwort festzulegen und dein Konto zu aktivieren.",
        _email_btn(invite_url, "✉ Einladung annehmen"),
        "Dieser Link ist <b>48 Stunden</b> gültig."
    )
    ok, err = _send_email(user["email"], "EV Tracker — Einladung", body_html)
    _audit("user_invited", f"uid={uid} email={user['email']} smtp_ok={ok}", ip=request.remote_addr)
    if ok:
        return jsonify({"ok": True})
    return jsonify({"ok": False, "error": err or "E-Mail konnte nicht gesendet werden"})

# ── Profile (own account) ─────────────────────────────────────────────────────
@app.route("/api/users/me")
def api_get_me():
    user = _current_user()
    if not user:
        return jsonify({"error": "Nicht eingeloggt"}), 401
    return jsonify({
        "id":           user["id"],
        "name":         user["name"],
        "email":        user["email"],
        "role":         user["role"],
        "totp_enabled": bool(user.get("totp_enabled")),
    })

@app.route("/api/users/me/password", methods=["POST"])
def api_change_password():
    user = _current_user()
    if not user:
        return jsonify({"error": "Nicht eingeloggt"}), 401
    data    = request.json or {}
    current = data.get("current","")
    new_pw  = data.get("new","")
    if _hash_password(current) != user["password_hash"]:
        return jsonify({"ok": False, "error": "Aktuelles Passwort falsch"})
    pw_err = _password_ok(new_pw)
    if pw_err:
        return jsonify({"ok": False, "error": pw_err})
    now = datetime.utcnow().isoformat()
    con = _get_db()
    con.execute("UPDATE users SET password_hash=?,updated_at=? WHERE id=?",
                (_hash_password(new_pw), now, user["id"]))
    con.commit(); con.close()
    _audit("password_changed", f"uid={user['id']}", ip=request.remote_addr)
    return jsonify({"ok": True})

@app.route("/api/users/me/totp/setup", methods=["POST"])
def api_my_totp_setup():
    user = _current_user()
    if not user:
        return jsonify({"error": "Nicht eingeloggt"}), 401
    import pyotp
    secret = pyotp.random_base32()
    session["pending_totp"] = secret
    uri = pyotp.TOTP(secret).provisioning_uri(name=user["email"], issuer_name="EV Tracker")
    return jsonify({"ok": True, "secret": secret, "uri": uri})

@app.route("/api/users/me/totp/confirm", methods=["POST"])
def api_my_totp_confirm():
    user   = _current_user()
    secret = session.get("pending_totp","")
    if not user or not secret:
        return jsonify({"ok": False, "error": "Kein ausstehender TOTP"})
    code = (request.json or {}).get("code","").strip().replace(" ","")
    import pyotp
    if not pyotp.TOTP(secret).verify(code, valid_window=1):
        return jsonify({"ok": False, "error": "Ungültiger Code — bitte erneut versuchen"})
    now = datetime.utcnow().isoformat()
    con = _get_db()
    con.execute("UPDATE users SET totp_secret=?,totp_enabled=1,updated_at=? WHERE id=?",
                (secret, now, user["id"]))
    con.commit(); con.close()
    session.pop("pending_totp", None)
    _audit("totp_enabled", f"uid={user['id']}", ip=request.remote_addr)
    return jsonify({"ok": True})

@app.route("/api/users/me/totp/disable", methods=["POST"])
def api_my_totp_disable():
    user = _current_user()
    if not user:
        return jsonify({"error": "Nicht eingeloggt"}), 401
    now = datetime.utcnow().isoformat()
    con = _get_db()
    con.execute("UPDATE users SET totp_secret='',totp_enabled=0,updated_at=? WHERE id=?",
                (now, user["id"]))
    con.commit(); con.close()
    _audit("totp_disabled", f"uid={user['id']}", ip=request.remote_addr)
    return jsonify({"ok": True})

@app.route("/api/users/me/totp/backup-codes", methods=["POST"])
@require_login
def api_generate_backup_codes():
    user = _current_user()
    if not user:
        return jsonify({"error": "Nicht eingeloggt"}), 401
    raw_codes = [secrets.token_hex(4).upper() for _ in range(8)]
    formatted = [f"{c[:4]}-{c[4:]}" for c in raw_codes]
    now_iso = datetime.utcnow().isoformat()
    con = _get_db()
    con.execute("DELETE FROM totp_backup_codes WHERE user_id=?", (user["id"],))
    for code in raw_codes:
        code_hash = hashlib.sha256(code.encode()).hexdigest()
        con.execute("INSERT INTO totp_backup_codes (user_id, code_hash) VALUES (?,?)",
                    (user["id"], code_hash))
    con.commit(); con.close()
    _audit("backup_codes_generated", f"uid={user['id']}", ip=request.remote_addr)
    return jsonify({"ok": True, "codes": formatted})

@app.route("/api/users/me/totp/backup-codes/count", methods=["GET"])
@require_login
def api_backup_codes_count():
    user = _current_user()
    if not user:
        return jsonify({"error": "Nicht eingeloggt"}), 401
    con = _get_db()
    count = con.execute(
        "SELECT COUNT(*) FROM totp_backup_codes WHERE user_id=? AND used_at IS NULL",
        (user["id"],)).fetchone()[0]
    con.close()
    return jsonify({"count": count})

@app.route("/api/csrf-token", methods=["GET"])
@require_login
def api_csrf_token():
    if "csrf_token" not in session:
        session["csrf_token"] = secrets.token_hex(32)
    return jsonify({"token": session["csrf_token"]})

# ── SMTP ─────────────────────────────────────────────────────────────────────
@app.route("/api/smtp/test", methods=["POST"])
def smtp_test():
    import smtplib, ssl as _ssl
    data = request.json or {}
    cfg = load_config()
    host = data.get("smtp_host") or cfg.get("smtp_host","")
    port = int(data.get("smtp_port") or cfg.get("smtp_port", 587))
    tls  = data.get("smtp_tls") or cfg.get("smtp_tls","starttls")
    user = data.get("smtp_user") or cfg.get("smtp_user","")
    pw   = data.get("smtp_password") or cfg.get("smtp_password","")
    if not host:
        return jsonify({"ok": False, "error": "Kein SMTP-Server konfiguriert"})
    try:
        ctx = _ssl.create_default_context()
        if tls == "ssl":
            srv = smtplib.SMTP_SSL(host, port, context=ctx, timeout=10)
        else:
            srv = smtplib.SMTP(host, port, timeout=10)
            if tls == "starttls":
                srv.starttls(context=ctx)
        if user:
            srv.login(user, pw)
        srv.quit()
        _audit("smtp_test", f"host={host}:{port} tls={tls}")
        return jsonify({"ok": True, "message": f"✅ Verbindung zu {host}:{port} erfolgreich"})
    except Exception as e:
        return jsonify({"ok": False, "error": str(e)})

@app.route("/api/smtp/send-test", methods=["POST"])
def smtp_send_test():
    data = request.json or {}
    cfg = load_config()
    to  = data.get("to") or cfg.get("smtp_from_email","")
    if not cfg.get("smtp_host","") or not cfg.get("smtp_from_email",""):
        return jsonify({"ok": False, "error": "SMTP nicht konfiguriert"})
    body_html = _email_html(
        "SMTP Testmail",
        "Diese E-Mail bestätigt, dass deine SMTP-Konfiguration in EV Tracker korrekt funktioniert.",
        f"Gesendet an: <b>{to}</b>",
        "Falls du diese E-Mail erhalten hast, ist alles richtig eingestellt. ✅"
    )
    ok, err = _send_email(to, "EV Tracker — SMTP Test", body_html)
    if ok:
        _audit("smtp_send_test", f"to={to}")
        return jsonify({"ok": True, "message": f"Testmail an {to} versendet"})
    return jsonify({"ok": False, "error": err or "Unbekannter Fehler"})

# ── Export Templates ──────────────────────────────────────────────────────────
@app.route("/api/export/templates", methods=["GET"])
def get_export_templates():
    cfg = load_config()
    return jsonify(cfg.get("export_templates", []))

@app.route("/api/export/templates", methods=["POST"])
def create_export_template():
    data = request.json or {}
    cfg  = load_config()
    templates = cfg.get("export_templates", [])
    tid = secrets.token_hex(6)
    tpl = {
        "id":         tid,
        "name":       data.get("name", "Neue Vorlage"),
        "mapping":    data.get("mapping", {}),
        "start_row":  data.get("start_row"),
        "is_default": data.get("is_default", False),
    }
    if tpl["is_default"]:
        for t in templates: t["is_default"] = False
    templates.append(tpl)
    cfg["export_templates"] = templates
    save_config(cfg)
    _audit("export_template_create", f"name={tpl['name']}")
    return jsonify({"ok": True, "template": tpl})

@app.route("/api/export/templates/<tid>", methods=["PUT"])
def update_export_template(tid):
    data = request.json or {}
    cfg  = load_config()
    templates = cfg.get("export_templates", [])
    tpl = next((t for t in templates if t["id"]==tid), None)
    if not tpl:
        return jsonify({"ok": False, "error": "Nicht gefunden"}), 404
    if data.get("is_default"):
        for t in templates: t["is_default"] = False
    tpl.update({k:v for k,v in data.items() if k != "id"})
    cfg["export_templates"] = templates
    save_config(cfg)
    _audit("export_template_update", f"id={tid} name={tpl['name']}")
    return jsonify({"ok": True, "template": tpl})

@app.route("/api/export/templates/<tid>", methods=["DELETE"])
def delete_export_template(tid):
    cfg  = load_config()
    templates = cfg.get("export_templates", [])
    cfg["export_templates"] = [t for t in templates if t["id"]!=tid]
    save_config(cfg)
    _audit("export_template_delete", f"id={tid}")
    return jsonify({"ok": True})

# ── Admin Dashboard ──────────────────────────────────────────────────────────
@app.route("/api/admin/dashboard")
@require_admin
def api_admin_dashboard():
    con = _get_db()
    total   = con.execute("SELECT COUNT(*) FROM users").fetchone()[0]
    active  = con.execute("SELECT COUNT(*) FROM users WHERE status='active'").fetchone()[0]
    invited = con.execute("SELECT COUNT(*) FROM users WHERE status='invited'").fetchone()[0]
    locked  = con.execute("SELECT COUNT(*) FROM users WHERE locked_until IS NOT NULL AND locked_until > ?",
                          (datetime.utcnow().isoformat(),)).fetchone()[0]
    since24 = (datetime.utcnow() - timedelta(hours=24)).isoformat()
    failures = con.execute(
        "SELECT COUNT(*) FROM audit_log WHERE action='login_failed' AND ts > ?", (since24,)).fetchone()[0]
    lockouts = con.execute(
        "SELECT COUNT(*) FROM audit_log WHERE action='account_locked' AND ts > ?", (since24,)).fetchone()[0]
    con.close()
    return jsonify({
        "total_users":    total,
        "active_users":   active,
        "invited_users":  invited,
        "locked_users":   locked,
        "recent_failures": failures,
        "recent_lockouts": lockouts,
    })

# ── Audit Log ────────────────────────────────────────────────────────────────
@app.route("/api/audit-log")
@require_admin
def get_audit_log():
    limit = int(request.args.get("limit", 200))
    con = sqlite3.connect(DB_PATH)
    con.row_factory = sqlite3.Row
    rows = con.execute("""
        SELECT a.*, u.name as user_name, u.email as user_email
        FROM audit_log a
        LEFT JOIN users u ON a.user_id = u.id
        ORDER BY a.id DESC LIMIT ?
    """, (limit,)).fetchall()
    con.close()
    return jsonify([dict(r) for r in rows])

# ── Email Reports ─────────────────────────────────────────────────────────────

_report_timer = None

_DE_MONTHS_FULL = ["Januar","Februar","März","April","Mai","Juni",
                    "Juli","August","September","Oktober","November","Dezember"]
_EN_MONTHS_FULL = ["January","February","March","April","May","June",
                    "July","August","September","October","November","December"]


def calculate_report_period(schedule_type, period_mode, now, config):
    """Return dict with start, end (date objects), label_de, label_en, period_key."""
    from datetime import date, timedelta
    today = now.date() if hasattr(now, 'date') else now

    if period_mode == "custom_range":
        s = config.get("report_email_custom_start_date", "")
        e = config.get("report_email_custom_end_date", "")
        try:
            start = date.fromisoformat(s); end = date.fromisoformat(e)
        except Exception:
            start = end = today
        return {"start": start, "end": end,
                "label_de": f"{start.strftime('%d.%m.%Y')} – {end.strftime('%d.%m.%Y')}",
                "label_en": f"{start.isoformat()} – {end.isoformat()}",
                "period_key": f"custom:{start}:{end}"}

    if schedule_type == "daily":
        d = (today - timedelta(days=1)) if period_mode == "previous_period" else today
        return {"start": d, "end": d, "label_de": d.strftime("%d.%m.%Y"),
                "label_en": d.isoformat(), "period_key": f"daily:{d}"}

    if schedule_type == "weekly":
        if period_mode == "previous_period":
            this_mon = today - timedelta(days=today.weekday())
            start = this_mon - timedelta(weeks=1)
        else:
            start = today - timedelta(days=today.weekday())
        end = start + timedelta(days=6)
        year, week, _ = start.isocalendar()
        return {"start": start, "end": end,
                "label_de": f"KW {week:02d} / {year}", "label_en": f"Week {week:02d} / {year}",
                "period_key": f"weekly:{year}-W{week:02d}"}

    if schedule_type == "monthly":
        if period_mode == "previous_period":
            first_this = today.replace(day=1)
            end = first_this - timedelta(days=1); start = end.replace(day=1)
        else:
            start = today.replace(day=1)
            last = _calendar.monthrange(today.year, today.month)[1]
            end = today.replace(day=last)
        return {"start": start, "end": end,
                "label_de": f"{_DE_MONTHS_FULL[start.month-1]} {start.year}",
                "label_en": f"{_EN_MONTHS_FULL[start.month-1]} {start.year}",
                "period_key": f"monthly:{start.year}-{start.month:02d}"}

    if schedule_type == "quarterly":
        q = (today.month - 1) // 3 + 1; year = today.year
        if period_mode == "previous_period":
            q -= 1
            if q == 0: q = 4; year -= 1
        sm = (q - 1) * 3 + 1; em = sm + 2
        start = date(year, sm, 1)
        end   = date(year, em, _calendar.monthrange(year, em)[1])
        return {"start": start, "end": end,
                "label_de": f"Q{q} {year}", "label_en": f"Q{q} {year}",
                "period_key": f"quarterly:{year}-Q{q}"}

    if schedule_type == "yearly":
        year = (today.year - 1) if period_mode == "previous_period" else today.year
        return {"start": date(year, 1, 1), "end": date(year, 12, 31),
                "label_de": str(year), "label_en": str(year),
                "period_key": f"yearly:{year}"}

    if schedule_type == "custom_days":
        x = int(config.get("report_email_custom_days", 14))
        end = today; start = today - timedelta(days=x)
        return {"start": start, "end": end,
                "label_de": f"Letzte {x} Tage", "label_en": f"Last {x} days",
                "period_key": f"custom_days:{x}:{today}"}

    return calculate_report_period("monthly", period_mode, now, config)


def _get_report_sessions(start_date, end_date, location_filter="all", vehicle_filter="all"):
    from datetime import timedelta
    where  = ["end_ts IS NOT NULL", "start_ts >= ?", "start_ts < ?"]
    params = [start_date.isoformat(), (end_date + timedelta(days=1)).isoformat()]
    if location_filter == "home":
        where.append("location = 'home'")
    elif location_filter == "external":
        where.append("location = 'extern'")
    if vehicle_filter and vehicle_filter != "all":
        where.append("vehicle_id = ?"); params.append(vehicle_filter)
    sql = f"SELECT * FROM sessions WHERE {' AND '.join(where)} ORDER BY start_ts ASC"
    con = sqlite3.connect(DB_PATH); con.row_factory = sqlite3.Row
    rows = con.execute(sql, params).fetchall(); con.close()
    return [dict(r) for r in rows]


def _report_filter_labels(cfg, is_de):
    loc = cfg.get("report_email_location_filter", "all")
    veh = cfg.get("report_email_vehicle_filter", "all")
    if is_de:
        loc_lbl = {"all": "Alle Ladevorgänge", "home": "Nur Zuhause / Intern",
                   "external": "Nur Extern"}.get(loc, loc)
        veh_lbl = "Alle Fahrzeuge" if veh == "all" else veh
    else:
        loc_lbl = {"all": "All charging sessions", "home": "Home only",
                   "external": "External only"}.get(loc, loc)
        veh_lbl = "All vehicles" if veh == "all" else veh
    return loc_lbl, veh_lbl


def _build_report_html(sessions, period_info, cfg, lang="de"):
    is_de      = lang != "en"
    plabel     = period_info.get("label_de" if is_de else "label_en", "")
    loc_filter = cfg.get("report_email_location_filter", "all")
    loc_lbl, veh_lbl = _report_filter_labels(cfg, is_de)
    total_kwh  = sum(s.get("kwh_charged") or 0 for s in sessions)
    total_cost = sum(s.get("cost_eur")    or 0 for s in sessions)
    total_secs = sum(s.get("duration_sec") or 0 for s in sessions)
    home_kwh   = sum((s.get("kwh_charged") or 0) for s in sessions if s.get("location") == "home")
    ext_kwh    = sum((s.get("kwh_charged") or 0) for s in sessions if s.get("location") == "extern")
    total_h    = total_secs / 3600
    avg_price  = total_cost / total_kwh if total_kwh else 0
    avg_power  = total_kwh / total_h if total_h else 0
    n          = len(sessions)
    if is_de:
        title = "EV Tracker — Lade-Report"
        rows  = [("Zeitraum", plabel), ("Filter", loc_lbl), ("Fahrzeug", veh_lbl),
                 ("Ladevorgänge", str(n)),
                 ("Geladene kWh", f"{total_kwh:.2f} kWh"),
                 ("Gesamtkosten", f"{total_cost:.2f} €"),
                 ("Ø Preis/kWh", f"{avg_price:.4f} €"),
                 ("Ladezeit gesamt", f"{total_h:.1f} h"),
                 ("Ø Ladeleistung", f"{avg_power:.1f} kW")]
        if loc_filter == "all" and (home_kwh or ext_kwh):
            rows.append(("Zuhause / Extern", f"{home_kwh:.1f} / {ext_kwh:.1f} kWh"))
        empty_txt = "Keine Ladevorgänge im gewählten Zeitraum."
    else:
        title = "EV Tracker — Charging Report"
        rows  = [("Period", plabel), ("Sessions", str(n)),
                 ("Energy charged", f"{total_kwh:.2f} kWh"),
                 ("Total cost", f"{total_cost:.2f} €"),
                 ("Avg price/kWh", f"{avg_price:.4f} €"),
                 ("Total charge time", f"{total_h:.1f} h"),
                 ("Avg charge power", f"{avg_power:.1f} kW")]
        if home_kwh or ext_kwh:
            rows.append(("Home / External", f"{home_kwh:.1f} / {ext_kwh:.1f} kWh"))
        empty_txt = "No charging sessions in the selected period."
    trows = "".join(
        f'<tr><td style="padding:6px 14px;color:#888;white-space:nowrap">{k}</td>'
        f'<td style="padding:6px 14px;font-weight:600;color:#fff">{v}</td></tr>'
        for k, v in rows)
    empty = f'<p style="color:#f59e0b;margin:16px 0">{empty_txt}</p>' if not sessions else ""
    return (f'<!DOCTYPE html><html><head><meta charset="utf-8"></head>'
            f'<body style="background:#0f1117;color:#e8e8f0;font-family:sans-serif;margin:0;padding:24px">'
            f'<div style="max-width:560px;margin:0 auto"><div style="background:#1e2030;border-radius:12px;padding:28px">'
            f'<h1 style="color:#6ee7b7;font-size:1.3rem;margin:0 0 6px">⚡ {title}</h1>'
            f'<hr style="border:none;border-top:1px solid #2d3147;margin:16px 0">'
            f'{empty}<table style="width:100%;border-collapse:collapse">{trows}</table>'
            f'<hr style="border:none;border-top:1px solid #2d3147;margin:16px 0">'
            f'<p style="color:#555;font-size:.75rem;margin:0">EV Tracker v{APP_VERSION}</p>'
            f'</div></div></body></html>')


def _send_email_with_attachments(to_addr, subject, body_html, attachments=None):
    import smtplib as _smtp, ssl as _ssl
    from email.mime.multipart import MIMEMultipart
    from email.mime.text import MIMEText
    from email.mime.base import MIMEBase
    from email import encoders as _enc
    cfg  = load_config()
    host = cfg.get("smtp_host",""); port = int(cfg.get("smtp_port", 587))
    tls  = cfg.get("smtp_tls","starttls"); user = cfg.get("smtp_user","")
    pw   = cfg.get("smtp_password",""); frm  = cfg.get("smtp_from_email","")
    name = cfg.get("smtp_from_name","EV Tracker")
    if not host or not frm:
        return False, "SMTP nicht konfiguriert"
    try:
        msg = MIMEMultipart(); msg["From"] = f"{name} <{frm}>"; msg["To"] = to_addr
        msg["Subject"] = subject; msg.attach(MIMEText(body_html, "html", "utf-8"))
        for fname, data, mime_type in (attachments or []):
            part = MIMEBase(*mime_type.split("/", 1)); part.set_payload(data)
            _enc.encode_base64(part)
            part.add_header("Content-Disposition", "attachment", filename=fname)
            msg.attach(part)
        ctx = _ssl.create_default_context()
        if tls == "ssl":
            srv = _smtp.SMTP_SSL(host, port, context=ctx, timeout=15)
        else:
            srv = _smtp.SMTP(host, port, timeout=15)
            if tls == "starttls": srv.starttls(context=ctx)
        if user: srv.login(user, pw)
        srv.sendmail(frm, to_addr, msg.as_string()); srv.quit()
        return True, None
    except Exception as e:
        return False, str(e)


def _log_report_history(period_info, cfg, status, error, triggered_by):
    try:
        con = _get_db()
        con.execute("""INSERT INTO email_report_history
            (sent_at,schedule_type,period_start,period_end,period_key,
             location_filter,vehicle_filter,recipients,status,error,triggered_by)
            VALUES (?,?,?,?,?,?,?,?,?,?,?)""",
            (datetime.utcnow().isoformat(),
             cfg.get("report_email_schedule_type","monthly"),
             period_info["start"].isoformat(), period_info["end"].isoformat(),
             period_info["period_key"],
             cfg.get("report_email_location_filter","all"),
             cfg.get("report_email_vehicle_filter","all"),
             json.dumps(cfg.get("report_email_recipients",[])),
             status, error, triggered_by))
        con.commit(); con.close()
    except Exception as e:
        log.warning("Report-History-Log fehlgeschlagen: %s", e)


def _send_report_email(cfg=None, triggered_by="auto"):
    if cfg is None: cfg = load_config()
    if not cfg.get("report_email_enabled"):
        return False, "Reports nicht aktiviert"
    if not cfg.get("smtp_host","") or not cfg.get("smtp_from_email",""):
        return False, "SMTP nicht konfiguriert"
    recipients = cfg.get("report_email_recipients", [])
    if not recipients:
        return False, "Keine Empfänger konfiguriert"
    stype       = cfg.get("report_email_schedule_type", "monthly")
    period_mode = cfg.get("report_email_period_mode", "previous_period")
    period_info = calculate_report_period(stype, period_mode, datetime.now(), cfg)
    period_key  = period_info["period_key"]
    if triggered_by == "auto" and cfg.get("report_email_last_sent_key","") == period_key:
        log.info("Report bereits gesendet für %s — übersprungen", period_key)
        _log_report_history(period_info, cfg, "skipped", None, triggered_by)
        return True, None
    loc_filter = cfg.get("report_email_location_filter", "all")
    veh_filter = cfg.get("report_email_vehicle_filter", "all")
    lang       = cfg.get("report_email_language", "auto")
    if lang == "auto": lang = "de"
    sessions   = _get_report_sessions(period_info["start"], period_info["end"], loc_filter, veh_filter)
    if cfg.get("report_email_include_summary", True):
        html = _build_report_html(sessions, period_info, cfg, lang)
    else:
        html = f"<p>EV Tracker Report — {period_info.get('label_de','')}</p>"
    plabel  = period_info.get("label_de" if lang != "en" else "label_en", period_key)
    subject = f"EV Tracker — Report {plabel}"
    attachments = []
    if cfg.get("report_email_include_excel") and sessions:
        try:
            from export_excel import export as _export_func
            sig_path = str(SIGNATURE_PATH) if SIGNATURE_PATH.exists() and cfg.get("report_email_include_signature") else None
            sig_map  = cfg.get("signature_mapping", {}) if sig_path else {}
            xl_loc   = "extern" if loc_filter == "external" else loc_filter
            xl_bytes, _ = _export_func(
                sessions=sessions, year=period_info["start"].year,
                month=period_info["start"].month, location=xl_loc,
                config=cfg, lang=lang, include_signature=bool(sig_path),
                signature_path=sig_path, signature_mapping=sig_map, return_warnings=True)
            attachments.append(("Ladeprotokoll.xlsx", xl_bytes,
                                 "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"))
        except Exception as e:
            log.warning("Report-Excel-Anhang fehlgeschlagen: %s", e)
    errors = []
    for to in recipients:
        ok, err = _send_email_with_attachments(to, subject, html, attachments)
        if not ok: errors.append(f"{to}: {err}")
    if errors:
        _log_report_history(period_info, cfg, "error", "; ".join(errors), triggered_by)
        return False, "; ".join(errors)
    cfg["report_email_last_sent_key"] = period_key
    save_config(cfg)
    _log_report_history(period_info, cfg, "sent", None, triggered_by)
    log.info("Report gesendet: %s → %s", period_key, recipients)
    return True, None


def _next_report_seconds(cfg, now=None):
    if not cfg.get("report_email_enabled"): return None
    stype  = cfg.get("report_email_schedule_type", "monthly")
    t_str  = cfg.get("report_email_time", "08:00")
    if now is None: now = datetime.now()
    try: t_hour, t_min = [int(x) for x in t_str.split(":")]
    except Exception: t_hour, t_min = 8, 0
    today = now.date()
    from datetime import date, timedelta

    if stype == "daily":
        fire = datetime.combine(today, datetime.min.time()).replace(hour=t_hour, minute=t_min)
        if fire <= now: fire += timedelta(days=1)
        return (fire - now).total_seconds()

    if stype == "weekly":
        wd = int(cfg.get("report_email_weekday", 1)) - 1
        da = (wd - today.weekday()) % 7 or 7
        fire = datetime.combine(today + timedelta(days=da), datetime.min.time()).replace(hour=t_hour, minute=t_min)
        if fire <= now: fire += timedelta(weeks=1)
        return (fire - now).total_seconds()

    if stype == "monthly":
        dom = int(cfg.get("report_email_day_of_month", 1))
        try:
            fire = datetime.combine(today.replace(day=dom), datetime.min.time()).replace(hour=t_hour, minute=t_min)
            if fire > now: return (fire - now).total_seconds()
        except ValueError: pass
        nm = today.month % 12 + 1; ny = today.year + (1 if today.month == 12 else 0)
        try: fire = datetime(ny, nm, dom, t_hour, t_min)
        except ValueError: fire = datetime(ny, nm, _calendar.monthrange(ny, nm)[1], t_hour, t_min)
        return (fire - now).total_seconds()

    if stype == "quarterly":
        dom = int(cfg.get("report_email_day_of_month", 1))
        for yo in range(2):
            for qm in [1, 4, 7, 10]:
                try:
                    fire = datetime(today.year + yo, qm, dom, t_hour, t_min)
                    if fire > now: return (fire - now).total_seconds()
                except ValueError: continue
        return 90 * 86400

    if stype == "yearly":
        mo  = int(cfg.get("report_email_month", 1))
        dom = int(cfg.get("report_email_day_of_month", 1))
        for year in [today.year, today.year + 1]:
            try:
                fire = datetime(year, mo, dom, t_hour, t_min)
                if fire > now: return (fire - now).total_seconds()
            except ValueError: continue
        return 365 * 86400

    if stype == "custom_days":
        x    = int(cfg.get("report_email_custom_days", 14))
        fire = datetime.combine(today, datetime.min.time()).replace(hour=t_hour, minute=t_min)
        if fire <= now: fire += timedelta(days=x)
        return (fire - now).total_seconds()

    if stype == "custom_cron":
        cron = cfg.get("report_email_cron", "").strip()
        if not cron: return None
        try:
            parts = cron.split()
            if len(parts) >= 2:
                c_min, c_hour = int(parts[0]), int(parts[1])
                fire = datetime.combine(today, datetime.min.time()).replace(hour=c_hour, minute=c_min)
                if fire <= now: fire += timedelta(days=1)
                return (fire - now).total_seconds()
        except Exception: return None

    return None


def schedule_report():
    global _report_timer
    cfg  = load_config()
    secs = _next_report_seconds(cfg)
    if not secs or secs <= 0: return
    def run():
        try:
            ok, err = _send_report_email(triggered_by="auto")
            if not ok: log.warning("Auto-Report Fehler: %s", err)
        except Exception as e:
            log.warning("Auto-Report Exception: %s", e)
        schedule_report()
    _report_timer = Timer(secs, run); _report_timer.daemon = True; _report_timer.start()
    log.info("Nächster Auto-Report: %s", (datetime.now() + timedelta(seconds=secs)).strftime("%d.%m.%Y %H:%M"))


@app.route("/api/report/config", methods=["GET"])
@require_login
def api_report_config_get():
    if not has_permission(_current_user(), "reports:view"):
        return jsonify({"error": "Keine Berechtigung"}), 403
    cfg  = load_config()
    keys = [k for k in DEFAULT_CONFIG if k.startswith("report_email_")]
    return jsonify({k: cfg.get(k, DEFAULT_CONFIG[k]) for k in keys})


@app.route("/api/report/config", methods=["POST"])
@require_admin
def api_report_config_save():
    if not has_permission(_current_user(), "reports:configure"):
        return jsonify({"error": "Keine Berechtigung"}), 403
    data = request.get_json(force=True) or {}
    cfg  = load_config()
    allowed = [k for k in DEFAULT_CONFIG if k.startswith("report_email_")]
    for k in allowed:
        if k in data:
            cfg[k] = data[k]
    save_config(cfg)
    schedule_report()
    _audit("report_config_saved", ip=request.remote_addr)
    return jsonify({"ok": True})


@app.route("/api/report/send-now", methods=["POST"])
@require_login
def api_report_send_now():
    if not has_permission(_current_user(), "reports:send"):
        return jsonify({"error": "Keine Berechtigung"}), 403
    data = request.get_json(force=True) or {}
    cfg  = load_config()
    # Allow overriding config for this send
    for k in ["report_email_location_filter","report_email_vehicle_filter",
              "report_email_period_mode","report_email_schedule_type",
              "report_email_recipients","report_email_language"]:
        if k in data: cfg[k] = data[k]
    cfg["report_email_enabled"] = True
    ok, err = _send_report_email(cfg=cfg, triggered_by="manual")
    _audit("report_send_now", f"ok={ok} err={err}", ip=request.remote_addr)
    return jsonify({"ok": ok, "error": err})


@app.route("/api/report/history")
@require_login
def api_report_history():
    if not has_permission(_current_user(), "reports:history"):
        return jsonify({"error": "Keine Berechtigung"}), 403
    limit = int(request.args.get("limit", 50))
    con = _get_db()
    rows = con.execute(
        "SELECT * FROM email_report_history ORDER BY id DESC LIMIT ?", (limit,)
    ).fetchall()
    con.close()
    return jsonify([dict(r) for r in rows])

if __name__=="__main__":
    DATA_DIR.mkdir(parents=True, exist_ok=True)
    app.secret_key = _get_secret_key()
    init_db(); start_tracker(); schedule_backup(); schedule_report()
    app.run(host="0.0.0.0",port=8080,debug=False)
