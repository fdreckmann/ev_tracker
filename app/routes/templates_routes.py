"""
Export template (XLSX) routes: upload, gallery, preview, render, analyze, mapping.
"""
import json
from io import BytesIO

from flask import Blueprint, jsonify, request, send_file

from core.db import _get_db, close_db_if_owned, DATA_DIR
from core.config import load_config, save_config
from core.security import require_login, has_permission, _current_user, _audit

templates_routes_bp = Blueprint("templates_routes", __name__)

# ── Path constants (mirrors server.py) ───────────────────────────────────────
TEMPLATE_PATH  = DATA_DIR / "template.xlsx"
SIGNATURE_DIR  = DATA_DIR / "signatures"
SIGNATURE_PATH = SIGNATURE_DIR / "default_signature.png"


# ── POST /api/template ────────────────────────────────────────────────────────
@templates_routes_bp.route("/api/template", methods=["POST"])
@require_login
def api_upload_template():
    if not has_permission(_current_user(), "templates:upload"):
        return jsonify({"ok": False, "error": "Keine Berechtigung: templates:upload"}), 403
    if "file" not in request.files:
        return jsonify({"ok": False, "error": "Keine Datei"}), 400
    f = request.files["file"]
    if not f.filename.endswith(".xlsx"):
        return jsonify({"ok": False, "error": "Nur .xlsx"}), 400
    f.seek(0, 2)
    if f.tell() > 10 * 1024 * 1024:
        return jsonify({"ok": False, "error": "Template zu groß (max 10 MB)"}), 400
    f.seek(0)
    DATA_DIR.mkdir(parents=True, exist_ok=True)
    f.save(TEMPLATE_PATH)
    cfg = load_config()
    cfg["active_template"] = {"source": "upload", "template_id": None, "name": f.filename}
    save_config(cfg)
    return jsonify({"ok": True, "filename": f.filename, "size": TEMPLATE_PATH.stat().st_size})


# ── DELETE /api/template ──────────────────────────────────────────────────────
@templates_routes_bp.route("/api/template", methods=["DELETE"])
@require_login
def api_delete_template():
    if not has_permission(_current_user(), "templates:delete"):
        return jsonify({"ok": False, "error": "Keine Berechtigung: templates:delete"}), 403
    if TEMPLATE_PATH.exists():
        TEMPLATE_PATH.unlink()
    cfg = load_config()
    cfg["active_template"] = {"source": None, "template_id": None, "name": None}
    save_config(cfg)
    return jsonify({"ok": True})


# ── GET /api/template/info ────────────────────────────────────────────────────
@templates_routes_bp.route("/api/template/info")
@require_login
def api_template_info():
    from datetime import datetime
    cfg = load_config()
    if TEMPLATE_PATH.exists():
        return jsonify({
            "exists": True,
            "size": TEMPLATE_PATH.stat().st_size,
            "modified": datetime.fromtimestamp(TEMPLATE_PATH.stat().st_mtime).isoformat(timespec="seconds"),
            "active_template": cfg.get("active_template") or {},
            "has_signature": SIGNATURE_PATH.exists(),
        })
    return jsonify({
        "exists": False,
        "active_template": cfg.get("active_template") or {},
        "has_signature": SIGNATURE_PATH.exists(),
    })


# ── GET /api/template/gallery ─────────────────────────────────────────────────
@templates_routes_bp.route("/api/template/gallery")
@require_login
def api_template_gallery():
    import json as _json
    from pathlib import Path as _Path
    gallery_dir = _Path(__file__).parent.parent / "builtin_templates"
    templates = []
    if gallery_dir.exists():
        for tpl_dir in sorted(gallery_dir.iterdir()):
            manifest_path = tpl_dir / "manifest.json"
            if not manifest_path.exists():
                continue
            try:
                m = _json.loads(manifest_path.read_text(encoding="utf-8"))
                templates.append({
                    "id":              m["id"],
                    "name":            m["name"],
                    "description":     m.get("description", ""),
                    "category":        m.get("category", ""),
                    "recommended_for": m.get("recommended_for", []),
                    "preview_url":     f"/api/template/gallery/{m['id']}/preview",
                })
            except Exception:
                pass
    cfg = load_config()
    active_id = (cfg.get("active_template") or {}).get("template_id")
    for t in templates:
        t["active"] = (t["id"] == active_id)
    return jsonify({"ok": True, "templates": templates})


# ── GET /api/template/gallery/<template_id>/preview ───────────────────────────
@templates_routes_bp.route("/api/template/gallery/<template_id>/preview")
def api_template_gallery_preview(template_id):
    from pathlib import Path as _Path
    import re
    if not re.match(r'^[a-z0-9_]+$', template_id):
        return "Not found", 404
    svg_path = _Path(__file__).parent.parent / "builtin_templates" / template_id / "preview.svg"
    if not svg_path.exists():
        return "Not found", 404
    return send_file(str(svg_path), mimetype="image/svg+xml")


# ── POST /api/template/gallery/<template_id>/use ──────────────────────────────
@templates_routes_bp.route("/api/template/gallery/<template_id>/use", methods=["POST"])
@require_login
def api_template_gallery_use(template_id):
    if not has_permission(_current_user(), "templates:gallery_use"):
        return jsonify({"ok": False, "error": "Keine Berechtigung: templates:gallery_use"}), 403
    import json as _json
    import re
    from pathlib import Path as _Path
    if not re.match(r'^[a-z0-9_]+$', template_id):
        return jsonify({"ok": False, "error": "Ungültige Template-ID"}), 400
    gallery_dir = _Path(__file__).parent.parent / "builtin_templates"
    manifest_path = gallery_dir / template_id / "manifest.json"
    if not manifest_path.exists():
        return jsonify({"ok": False, "error": "Template nicht gefunden"}), 404
    try:
        m = _json.loads(manifest_path.read_text(encoding="utf-8"))
    except Exception as e:
        return jsonify({"ok": False, "error": f"Manifest-Fehler: {e}"}), 500

    try:
        from builtin_template_gen import generate_builtin_template
        ok = generate_builtin_template(template_id, TEMPLATE_PATH)
        if not ok:
            return jsonify({"ok": False, "error": "Template konnte nicht generiert werden"}), 500
    except Exception as e:
        return jsonify({"ok": False, "error": str(e)}), 500

    dm = m.get("default_mapping", {})
    cfg = load_config()
    cfg["template_column_mapping"] = dm.get("column_mapping", {})
    cfg["template_cell_mapping"]   = dm.get("cell_mapping", {})
    cfg["template_start_row"]      = dm.get("start_row")
    cfg["template_sheet"]          = dm.get("sheet", "")
    cfg["signature_mapping"]       = dm.get("signature_mapping", {})
    cfg["active_template"] = {
        "source":      "builtin",
        "template_id": template_id,
        "name":        m["name"],
    }
    save_config(cfg)
    _audit("template_gallery_use", f"id={template_id}", ip=request.remote_addr)
    return jsonify({
        "ok": True,
        "active_template": cfg["active_template"],
        "mapping": {
            "column_mapping": cfg["template_column_mapping"],
            "cell_mapping":   cfg["template_cell_mapping"],
            "start_row":      cfg["template_start_row"],
            "sheet":          cfg["template_sheet"],
        },
    })


# ── POST /api/template/preview ────────────────────────────────────────────────
@templates_routes_bp.route("/api/template/preview")
@require_login
def api_template_preview():
    if not TEMPLATE_PATH.exists():
        return jsonify({"ok": False, "error": "Kein Template"})
    try:
        import openpyxl
        from openpyxl.utils import get_column_letter
        from export_excel import match_column
        wb = openpyxl.load_workbook(TEMPLATE_PATH, read_only=True, data_only=True)
        ws = wb.active
        all_rows = []
        max_col  = 0
        raw_rows = list(ws.iter_rows(values_only=True, max_row=40))
        for row in raw_rows:
            last = max((i for i, v in enumerate(row) if v is not None), default=-1)
            if last >= 0:
                max_col = max(max_col, last + 1)
        max_col = max(max_col, 1)
        for ri, row in enumerate(raw_rows):
            cells = []
            for ci in range(max_col):
                v = row[ci] if ci < len(row) else None
                cells.append(str(v) if v is not None else "")
            all_rows.append({"row": ri + 1, "cells": cells})
        auto_header = None
        for r in all_rows:
            if sum(1 for c in r["cells"] if c.strip()) >= 2:
                auto_header = r["row"]
                break
        col_letters = [get_column_letter(i + 1) for i in range(max_col)]
        col_suggestions = {}
        if auto_header:
            hrow = next(r for r in all_rows if r["row"] == auto_header)
            for ci, val in enumerate(hrow["cells"], 1):
                col_suggestions[ci] = {
                    "header":     val,
                    "mapped_to":  match_column(val) if val else None,
                    "col_letter": get_column_letter(ci),
                }
        wb.close()
        return jsonify({
            "ok": True,
            "rows": all_rows,
            "col_letters": col_letters,
            "auto_header": auto_header,
            "col_suggestions": col_suggestions,
        })
    except Exception as e:
        return jsonify({"ok": False, "error": str(e)})


# ── GET /api/template/render ──────────────────────────────────────────────────
@templates_routes_bp.route("/api/template/render")
@require_login
def api_template_render():
    if not TEMPLATE_PATH.exists():
        return jsonify({"ok": False, "error": "Kein Template"})
    try:
        import openpyxl
        from openpyxl.utils import get_column_letter
        wb = openpyxl.load_workbook(TEMPLATE_PATH, data_only=True)
        ws = wb.active

        merged_topleft = {}
        merged_skip    = set()
        for rng in ws.merged_cells.ranges:
            merged_topleft[(rng.min_row, rng.min_col)] = (
                rng.max_row - rng.min_row + 1,
                rng.max_col - rng.min_col + 1,
            )
            for r in range(rng.min_row, rng.max_row + 1):
                for c in range(rng.min_col, rng.max_col + 1):
                    if (r, c) != (rng.min_row, rng.min_col):
                        merged_skip.add((r, c))

        max_row = min(ws.max_row or 1, 40)
        max_col = ws.max_column or 1

        def hex_color(color_obj):
            if color_obj is None:
                return None
            try:
                if color_obj.type == "rgb":
                    rgb = color_obj.rgb
                    if rgb and rgb not in ("00000000", "FF000000", "FFFFFFFF"):
                        return "#" + rgb[-6:]
                if color_obj.type == "theme":
                    return None
            except Exception:
                pass
            return None

        rows = []
        for ri in range(1, max_row + 1):
            cells = []
            for ci in range(1, max_col + 1):
                if (ri, ci) in merged_skip:
                    cells.append(None)
                    continue
                cell = ws.cell(row=ri, column=ci)
                rs, cs = merged_topleft.get((ri, ci), (1, 1))
                bg = bold = fg = None
                try:
                    if cell.fill and cell.fill.patternType not in (None, "none"):
                        bg = hex_color(cell.fill.fgColor)
                except Exception:
                    pass
                try:
                    if cell.font:
                        bold = bool(cell.font.bold)
                        fg   = hex_color(cell.font.color)
                except Exception:
                    pass
                cells.append({
                    "v":    str(cell.value) if cell.value is not None else "",
                    "bg":   bg,
                    "fg":   fg,
                    "bold": bold,
                    "rs":   rs,
                    "cs":   cs,
                    "r":    ri,
                    "c":    ci,
                })
            rows.append({"row": ri, "cells": cells})

        col_widths = []
        for ci in range(1, max_col + 1):
            letter = get_column_letter(ci)
            dim = ws.column_dimensions.get(letter)
            w = dim.width if dim and dim.width else 8
            col_widths.append(round(min(max(w, 4), 25), 1))

        col_letters = [get_column_letter(i) for i in range(1, max_col + 1)]
        wb.close()
        return jsonify({
            "ok": True,
            "rows": rows,
            "col_letters": col_letters,
            "col_widths": col_widths,
            "max_col": max_col,
        })
    except Exception as e:
        return jsonify({"ok": False, "error": str(e)})


# ── GET /api/sessions/sample ──────────────────────────────────────────────────
@templates_routes_bp.route("/api/sessions/sample")
@require_login
def api_sessions_sample():
    """Return one recent completed session for live export preview."""
    import sqlite3
    from core.db import DB_PATH
    try:
        con = sqlite3.connect(DB_PATH)
        con.row_factory = sqlite3.Row
        row = con.execute(
            "SELECT * FROM sessions WHERE end_ts IS NOT NULL ORDER BY start_ts DESC LIMIT 1"
        ).fetchone()
        close_db_if_owned(con)
        return jsonify(dict(row) if row else {})
    except Exception:
        return jsonify({})


# ── GET /api/template/analyze ─────────────────────────────────────────────────
@templates_routes_bp.route("/api/template/analyze")
@require_login
def api_template_analyze():
    if not TEMPLATE_PATH.exists():
        return jsonify({"ok": False, "error": "Kein Template hochgeladen"})
    try:
        from template_analyzer import analyze_template
        result = analyze_template(TEMPLATE_PATH)
        return jsonify(result)
    except Exception as e:
        return jsonify({"ok": False, "error": str(e)})


# ── GET + POST /api/template/mapping ─────────────────────────────────────────
@templates_routes_bp.route("/api/template/mapping", methods=["GET", "POST"])
@require_login
def api_template_mapping():
    cfg = load_config()
    if request.method == "POST":
        if not has_permission(_current_user(), "templates:edit_mapping"):
            return jsonify({"ok": False, "error": "Keine Berechtigung: templates:edit_mapping"}), 403
        body = request.get_json(force=True)
        # backward compat: "mapping" → column_mapping
        cfg["template_column_mapping"] = body.get("column_mapping") or body.get("mapping") or {}
        cfg["template_cell_mapping"]   = body.get("cell_mapping", {})
        cfg["template_start_row"]      = body.get("start_row")
        cfg["template_header_row"]     = body.get("header_row")
        cfg["template_sheet"]          = body.get("sheet") or ""
        cfg["signature_mapping"]       = body.get("signature_mapping") or {}
        save_config(cfg)
        return jsonify({"ok": True})
    return jsonify({
        "column_mapping":    cfg.get("template_column_mapping") or cfg.get("template_mapping") or {},
        "cell_mapping":      cfg.get("template_cell_mapping", {}),
        "start_row":         cfg.get("template_start_row"),
        "header_row":        cfg.get("template_header_row"),
        "sheet":             cfg.get("template_sheet", ""),
        "signature_mapping": cfg.get("signature_mapping") or {},
    })
