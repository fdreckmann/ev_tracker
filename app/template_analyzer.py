"""
Template analyzer for EV Tracker Excel templates.
Scans up to 80 rows x 40 columns to detect header rows, column mappings,
cell mappings and {{placeholder}} fields.
"""
import re
from pathlib import Path

try:
    import openpyxl
    from openpyxl.cell.cell import MergedCell
except ImportError:
    openpyxl = None
    MergedCell = None

from template_fields import TABLE_FIELDS, HEADER_FIELDS, CHARGER_POWER_PRIORITY_KEYWORDS

MAX_ROWS = 80
MAX_COLS = 40

_SIGNATURE_KEYWORDS = [
    "unterschrift", "signatur", "signature",
    "unterschrift fahrer", "unterschrift mitarbeiter",
    "ort datum unterschrift", "bestätigung", "genehmigt",
    "freigabe", "gezeichnet", "unterzeichner",
]


def _cell_addr(row, col):
    """Convert (row, col) to Excel address like 'B4'."""
    letters = ""
    c = col
    while c > 0:
        c, rem = divmod(c - 1, 26)
        letters = chr(65 + rem) + letters
    return f"{letters}{row}"


def _get_cell_value(ws, row, col):
    """Get cell value, handling MergedCell gracefully."""
    try:
        cell = ws.cell(row=row, column=col)
        if isinstance(cell, MergedCell):
            return None
        return cell.value
    except Exception:
        return None


def _normalize(text):
    """Lowercase, strip, remove trailing colon."""
    if text is None:
        return ""
    return str(text).lower().strip().rstrip(":").strip()


def _match_table_field(text):
    """Match normalized text against TABLE_FIELDS synonyms.
    Returns (field, confidence, reason) or (None, 0, '')."""
    norm = _normalize(text)
    if not norm:
        return None, 0.0, ""

    best_field = None
    best_conf = 0.0
    best_reason = ""

    for field, info in TABLE_FIELDS.items():
        synonyms = info.get("synonyms", [])
        for syn in synonyms:
            syn_lower = syn.lower()
            if norm == syn_lower:
                conf = 1.0
                reason = f"Exakter Treffer '{text}'"
                if conf > best_conf:
                    best_field, best_conf, best_reason = field, conf, reason
            elif syn_lower in norm or norm in syn_lower:
                conf = 0.8
                reason = f"Enthält '{syn}'"
                if conf > best_conf:
                    best_field, best_conf, best_reason = field, conf, reason
            else:
                # Partial: check if any keyword word appears in cell text
                words = syn_lower.split()
                if len(words) > 1 and all(w in norm for w in words):
                    conf = 0.7
                    reason = f"Teilweise Übereinstimmung '{syn}'"
                    if conf > best_conf:
                        best_field, best_conf, best_reason = field, conf, reason

    # Tiebreaker: if both charge_power_kw and charger_power_kw match with equal confidence,
    # and the header contains a priority keyword (kw/stunde, wallbox, 11kw …) → charger_power_kw wins.
    if best_field == "charge_power_kw" and best_conf < 1.0:
        norm_lower = norm
        if any(kw in norm_lower for kw in CHARGER_POWER_PRIORITY_KEYWORDS):
            best_field = "charger_power_kw"
            best_reason += " [priority: charger]"

    return best_field, best_conf, best_reason


def _match_header_field(text):
    """Match normalized text against HEADER_FIELDS synonyms.
    Returns (field, confidence, reason) or (None, 0, '')."""
    norm = _normalize(text)
    if not norm:
        return None, 0.0, ""

    best_field = None
    best_conf = 0.0
    best_reason = ""

    for field, info in HEADER_FIELDS.items():
        synonyms = info.get("synonyms", [])
        for syn in synonyms:
            syn_lower = syn.lower()
            if norm == syn_lower:
                conf = 1.0
                reason = f"Exakter Treffer '{text}'"
                if conf > best_conf:
                    best_field, best_conf, best_reason = field, conf, reason
            elif syn_lower in norm or norm in syn_lower:
                conf = 0.85
                reason = f"Enthält '{syn}'"
                if conf > best_conf:
                    best_field, best_conf, best_reason = field, conf, reason

    return best_field, best_conf, best_reason


def _find_merged_topleft(ws, row, col):
    """If (row, col) is within a merged range, return top-left (row, col), else same."""
    for mr in ws.merged_cells.ranges:
        if mr.min_row <= row <= mr.max_row and mr.min_col <= col <= mr.max_col:
            return mr.min_row, mr.min_col
    return row, col


def _is_merged_continuation(ws, row, col):
    """Check if cell is a MergedCell (continuation, not top-left)."""
    cell = ws.cell(row=row, column=col)
    return isinstance(cell, MergedCell)


def _merge_span(ws, row, col):
    """Return column span of a merged cell starting at (row, col)."""
    for mr in ws.merged_cells.ranges:
        if mr.min_row == row and mr.min_col == col:
            return mr.max_col - mr.min_col + 1
    return 1


def analyze_template(path: Path) -> dict:
    """Analyze an Excel template and return field mapping suggestions."""
    if openpyxl is None:
        return {"ok": False, "error": "openpyxl nicht verfügbar"}

    try:
        wb = openpyxl.load_workbook(str(path), data_only=True)
    except Exception as e:
        return {"ok": False, "error": str(e)}

    try:
        all_sheets = wb.sheetnames
        ws = wb.active
        sheet_name = ws.title

        warnings = []
        placeholders = []
        column_mapping = {}
        cell_mapping = {}

        # ── Step 1: Placeholder detection (scan all cells) ───────────────────
        ph_pattern = re.compile(r'\{\{(\w+)\}\}')
        for row in ws.iter_rows(min_row=1, max_row=min(ws.max_row or MAX_ROWS, MAX_ROWS),
                                 max_col=min(ws.max_column or MAX_COLS, MAX_COLS)):
            for cell in row:
                if isinstance(cell, MergedCell):
                    continue
                val = cell.value
                if val is None:
                    continue
                val_str = str(val)
                matches = ph_pattern.findall(val_str)
                for m in matches:
                    field = None
                    if m in HEADER_FIELDS:
                        field = m
                    elif m in TABLE_FIELDS:
                        field = m
                    if field:
                        addr = _cell_addr(cell.row, cell.column)
                        placeholders.append({
                            "cell": addr,
                            "field": field,
                            "confidence": 1.0,
                        })
                        # Add to cell_mapping with confidence 1.0
                        cell_mapping[addr] = {
                            "field": field,
                            "confidence": 1.0,
                            "reason": f"Platzhalter {{{{{field}}}}}",
                        }

        # ── Step 2: Header row detection (rows 1..40) ────────────────────────
        row_scores = {}
        scan_max_row = min(ws.max_row or MAX_ROWS, 40)

        for rnum in range(1, scan_max_row + 1):
            score = 0
            filled = 0
            merged_wide = False
            matched_fields = 0

            for cnum in range(1, min(ws.max_column or MAX_COLS, MAX_COLS) + 1):
                cell = ws.cell(row=rnum, column=cnum)
                if isinstance(cell, MergedCell):
                    continue
                val = cell.value
                if val is None or str(val).strip() == "":
                    continue

                filled += 1
                score += 2  # non-empty, non-merged-continuation

                # Check if it's a wide merged cell (title row indicator)
                span = _merge_span(ws, rnum, cnum)
                if span > 3:
                    merged_wide = True

                # Check if it matches a TABLE_FIELDS synonym
                field, conf, _ = _match_table_field(val)
                if field and conf >= 0.65:
                    score += 3
                    matched_fields += 1

            if merged_wide:
                score -= 2
            if filled < 2:
                score -= 1
            if filled == 0:
                score = -999  # skip empty rows

            row_scores[rnum] = score

        # Find best scoring row
        if row_scores:
            header_row = max(row_scores, key=lambda r: row_scores[r])
            max_score = row_scores[header_row]
        else:
            header_row = 1
            max_score = 0

        header_confidence = 1.0
        if max_score < 4:
            header_confidence = 0.4
            warnings.append("Keine klare Tabellenkopfzeile erkannt")

        start_row = header_row + 1

        # ── Step 3: Column mapping from header row ───────────────────────────
        total_header_cols = 0
        matched_columns = 0

        for cnum in range(1, min(ws.max_column or MAX_COLS, MAX_COLS) + 1):
            cell = ws.cell(row=header_row, column=cnum)
            if isinstance(cell, MergedCell):
                continue
            val = cell.value
            if val is None or str(val).strip() == "":
                continue

            total_header_cols += 1
            field, conf, reason = _match_table_field(val)

            if field and conf >= 0.65:
                col_str = str(cnum)
                # Only assign if not already assigned with higher confidence
                if col_str not in column_mapping or column_mapping[col_str]["confidence"] < conf:
                    column_mapping[col_str] = {
                        "field": field,
                        "confidence": round(conf, 2),
                        "reason": reason,
                    }
                matched_columns += 1
            elif val is not None and str(val).strip():
                warnings.append(f"Unbekannte Spalte: '{str(val).strip()}'")

        # ── Step 4: Cell mapping (header/footer area) ────────────────────────
        # Scan rows before header_row and after start_row + some data rows
        scan_ranges = list(range(1, header_row)) + list(range(
            start_row + 1,
            min((ws.max_row or MAX_ROWS) + 1, MAX_ROWS + 1)
        ))

        for rnum in scan_ranges:
            if rnum > (ws.max_row or MAX_ROWS):
                break
            for cnum in range(1, min(ws.max_column or MAX_COLS, MAX_COLS) + 1):
                cell = ws.cell(row=rnum, column=cnum)
                if isinstance(cell, MergedCell):
                    continue
                val = cell.value
                if val is None or str(val).strip() == "":
                    continue

                # Skip placeholder cells (already handled)
                if ph_pattern.search(str(val)):
                    continue

                field, conf, reason = _match_header_field(val)
                if field and conf >= 0.65:
                    # Look for target cell: prefer cell to the right
                    target_row, target_col = rnum, cnum + 1

                    # Try right neighbor first
                    right_cell = ws.cell(row=rnum, column=cnum + 1)
                    if isinstance(right_cell, MergedCell):
                        # Find merge top-left
                        tl_row, tl_col = _find_merged_topleft(ws, rnum, cnum + 1)
                        target_row, target_col = tl_row, tl_col
                    else:
                        right_val = right_cell.value
                        # Use right if empty or placeholder
                        if right_val is None or str(right_val).strip() == "" or (
                            isinstance(right_val, str) and ph_pattern.search(right_val)
                        ):
                            target_row, target_col = rnum, cnum + 1
                        else:
                            # Try cell below
                            below_cell = ws.cell(row=rnum + 1, column=cnum)
                            if not isinstance(below_cell, MergedCell) and (
                                below_cell.value is None or str(below_cell.value).strip() == ""
                            ):
                                target_row, target_col = rnum + 1, cnum
                            else:
                                # Fall back to right anyway
                                target_row, target_col = rnum, cnum + 1

                    addr = _cell_addr(target_row, target_col)
                    label_addr = _cell_addr(rnum, cnum)

                    # Check if target cell has a formula
                    target_cell = ws.cell(row=target_row, column=target_col)
                    target_val = target_cell.value if not isinstance(target_cell, MergedCell) else None
                    if isinstance(target_val, str) and target_val.startswith("="):
                        warnings.append(f"Formel in Zielzelle {addr}")
                        continue

                    # Only add if not already in cell_mapping with higher confidence
                    if addr not in cell_mapping or cell_mapping[addr]["confidence"] < conf:
                        cell_mapping[addr] = {
                            "field": field,
                            "confidence": round(conf, 2),
                            "reason": f"Label '{str(val).strip()}' in {label_addr}",
                        }

        # ── Step 5: Signature field detection ────────────────────────────────
        # Build merged_skip set for signature detection
        merged_skip = set()
        for mr in ws.merged_cells.ranges:
            for r2 in range(mr.min_row, mr.max_row + 1):
                for c2 in range(mr.min_col, mr.max_col + 1):
                    if r2 != mr.min_row or c2 != mr.min_col:
                        merged_skip.add((r2, c2))

        max_row = ws.max_row or MAX_ROWS
        max_col = ws.max_column or MAX_COLS

        signature_suggestion = None
        sig_confidence = 0.0
        for ri in range(1, min(max_row + 1, 81)):
            for ci in range(1, min(max_col + 1, 41)):
                if (ri, ci) in merged_skip:
                    continue
                try:
                    cell = ws.cell(row=ri, column=ci)
                except Exception:
                    continue
                if not cell.value:
                    continue
                txt = str(cell.value).lower().strip()
                # Check for {{signature}} placeholder
                if "{{signature}}" in txt:
                    addr = _cell_addr(ri, ci)
                    signature_suggestion = {
                        "cell": addr,
                        "width": 220, "height": 80,
                        "offset_x": 0, "offset_y": 0,
                        "confidence": 1.0,
                        "reason": f"Platzhalter {{{{signature}}}} in {addr}",
                    }
                    sig_confidence = 1.0
                    break
                # Check against keywords
                matched = any(kw in txt for kw in _SIGNATURE_KEYWORDS)
                if matched:
                    conf = 0.95 if any(kw == txt.rstrip(':').strip() for kw in _SIGNATURE_KEYWORDS) else 0.8
                    # Find target cell: prefer right neighbor, then below
                    target_addr = None
                    # Check right (ci+1)
                    if ci + 1 <= max_col:
                        r_cell = ws.cell(row=ri, column=ci + 1)
                        if (ri, ci + 1) not in merged_skip and (not r_cell.value or r_cell.value == ""):
                            target_addr = _cell_addr(ri, ci + 1)
                    # Check below (ri+1) if no right target
                    if not target_addr and ri + 1 <= max_row:
                        b_cell = ws.cell(row=ri + 1, column=ci)
                        if (ri + 1, ci) not in merged_skip and (not b_cell.value or b_cell.value == ""):
                            target_addr = _cell_addr(ri + 1, ci)
                    if target_addr and conf > sig_confidence:
                        sig_confidence = conf
                        reason = f"Label '{cell.value}' in {_cell_addr(ri, ci)}"
                        signature_suggestion = {
                            "cell": target_addr,
                            "width": 220, "height": 80,
                            "offset_x": 0, "offset_y": 0,
                            "confidence": conf,
                            "reason": reason,
                        }
            if sig_confidence >= 1.0:
                break

        # ── Step 6: Overall confidence ───────────────────────────────────────
        base = (matched_columns / max(total_header_cols, 1)) * 0.6 + header_confidence * 0.4
        confidence = base - 0.1 * len(warnings)
        confidence = max(0.1, min(1.0, confidence))

        return {
            "ok": True,
            "sheet": sheet_name,
            "all_sheets": all_sheets,
            "header_row": header_row,
            "start_row": start_row,
            "column_mapping": column_mapping,
            "cell_mapping": cell_mapping,
            "placeholders": placeholders,
            "warnings": warnings,
            "confidence": round(confidence, 2),
            "signature_suggestion": signature_suggestion,
        }

    except Exception as e:
        return {"ok": False, "error": str(e)}
