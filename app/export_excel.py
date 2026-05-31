import os, sys, sqlite3, shutil, json, re, copy
import re as _re_formula

_RE_RANGE_FORMULA = re.compile(r'[A-Za-z]+\d+:[A-Za-z]+\d+')

def _is_range_formula(formula):
    """Return True only for formulas referencing a cell range (e.g. SUM(C3:C12))."""
    if not isinstance(formula, str) or not formula.startswith("="):
        return False
    return bool(_RE_RANGE_FORMULA.search(formula))


def _detect_footer_start_row(ws, ds, max_row, col_map):
    """Detect footer_start_row heuristically when not explicitly configured.

    Phase 1 – range formulas: any row with a cell-range formula (SUM/SUMIF/…)
    is the start of the footer section.

    Phase 2 – plain-text footer: scan bottom-up from max_row.  A row qualifies
    as a footer row when ALL mapped columns are either empty or contain only a
    per-row formula AND at least one unmapped column carries a non-empty,
    non-formula text value.  We stop collecting as soon as we hit a row that
    is completely empty (a blank separator between data area and footer) or
    that has actual data in a mapped column.
    """
    if max_row < ds:
        return max_row + 1

    all_cols = range(1, (ws.max_column or 20) + 1)

    # Phase 1: range formulas are definitive footer indicators
    for r in range(ds, max_row + 1):
        for c in all_cols:
            try:
                if _is_range_formula(ws.cell(row=r, column=c).value):
                    return r
            except Exception:
                pass

    # Phase 2: bottom-up plain-text footer detection.
    # Real data rows contain NUMERIC values in mapped columns (kwh, cost, etc.).
    # Footer rows contain text labels (e.g. "Gesamt:", "Total") in any column.
    # We stop as soon as we hit a row with a numeric value in a mapped column,
    # or a completely blank row (blank separator between data and footer).
    footer_rows = []
    for r in range(max_row, ds - 1, -1):
        row_empty = True
        mapped_has_numeric = False   # numeric → real data row → stop
        has_any_text = False         # text label → footer candidate
        for c in all_cols:
            try:
                v = ws.cell(row=r, column=c).value
                if v is None:
                    continue
                row_empty = False
                if c in col_map:
                    if isinstance(v, (int, float)):
                        mapped_has_numeric = True   # real numeric data
                    elif isinstance(v, str) and v.startswith("="):
                        pass                         # formula — neutral
                    elif isinstance(v, str) and v.strip():
                        has_any_text = True          # label in mapped col → footer
                else:
                    if isinstance(v, str) and v.strip() and not v.startswith("="):
                        has_any_text = True
            except Exception:
                pass

        if row_empty:
            break               # blank separator → stop collecting footer rows
        if mapped_has_numeric:
            break               # found a real numeric data row → stop
        if has_any_text:
            footer_rows.append(r)
        else:
            break               # non-text, non-numeric content → stop

    if footer_rows:
        return min(footer_rows)
    return max_row + 1     # no footer detected
from datetime import datetime, timezone
from pathlib import Path
import openpyxl
from openpyxl.cell.cell import MergedCell
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter


def _adjust_formula_rows(formula, delta):
    """Adjust relative row references in an Excel formula by delta rows."""
    if not isinstance(formula, str) or not formula.startswith("=") or delta == 0:
        return formula
    def _sub(m):
        col_abs, col_let, row_abs, row_num = m.group(1), m.group(2), m.group(3), int(m.group(4))
        if not row_abs:
            row_num += delta
        return f"{col_abs}{col_let}{row_abs}{row_num}"
    return _re_formula.sub(r'(\$?)([A-Za-z]+)(\$?)(\d+)', _sub, formula)

try:
    from template_fields import TABLE_FIELDS, HEADER_FIELDS
except ImportError:
    TABLE_FIELDS = {}
    HEADER_FIELDS = {}

_PH_PATTERN   = re.compile(r'\{\{(\w+)\}\}')
_COL_ADDR_RE  = re.compile(r'^([A-Za-z]+)(\d+)$')


def format_meter_value_kwh(value):
    """Round meter reading to whole kWh for display/export. Returns int or None."""
    if value is None:
        return None
    try:
        return int(round(float(value)))
    except (ValueError, TypeError):
        return None

# ── Localization ──────────────────────────────────────────────────────────────
MONTH_NAMES = {
    "de": {1:"Januar",2:"Februar",3:"März",4:"April",5:"Mai",6:"Juni",
           7:"Juli",8:"August",9:"September",10:"Oktober",11:"November",12:"Dezember"},
    "en": {1:"January",2:"February",3:"March",4:"April",5:"May",6:"June",
           7:"July",8:"August",9:"September",10:"October",11:"November",12:"December"},
}

DATE_FORMATS = {
    "de": "%d.%m.%Y",
    "en": "%Y-%m-%d",
}

EXPORT_LABELS = {
    "de": {
        "charging_log": "Ladeprotokoll",
        "summary": "Zusammenfassung",
        "date": "Datum",
        "start": "Start",
        "end": "Ende",
        "duration": "Dauer",
        "duration_hours": "Ladedauer (h)",
        "charge_power_kw": "Ø Ladeleistung berech. (kW)",
        "charger_power_kw": "Ladepunkt-/Wallbox-Leistung (kW)",
        "kwh": "kWh",
        "cost": "Kosten",
        "location": "Standort",
        "charger_type": "Ladeart",
        "driver": "Fahrer",
        "license_plate": "Kennzeichen",
        "cost_center": "Kostenstelle",
        "total_cost": "Gesamtkosten",
        "total_kwh": "Gesamt kWh",
        "avg_charge_power_kw": "Durchschnittliche Ladeleistung (kW)",
        "total_charging_hours": "Gesamte Ladezeit (h)",
        "home": "Zuhause",
        "external": "Extern",
        "all_locations": "Alle Standorte",
        "month_name": "Monat",
        "month_year": "Monat/Jahr",
        "export_period": "Zeitraum",
        "price_per_kwh": "Preis/kWh",
        "row_num": "Nr.",
        "soc_start": "SOC Start (%)",
        "soc_end": "SOC Ende (%)",
        "odo_start": "KM-Stand Start",
        "odo_end": "KM-Stand Ende",
        "meter_old": "Zählerstand Alt",
        "meter_new": "Zählerstand Neu",
        "total_sessions": "Anzahl Ladevorgänge",
        "total_km": "Gesamte KM",
        # legacy keys kept for compatibility
        "department": "Abteilung",
        "export_date": "Exportdatum",
        "avg_charge_power": "Ø Ladeleistung (kW)",
        "total_charging_time": "Ladezeit gesamt",
        "signature": "Unterschrift",
        "period": "Zeitraum",
    },
    "en": {
        "charging_log": "Charging log",
        "summary": "Summary",
        "date": "Date",
        "start": "Start",
        "end": "End",
        "duration": "Duration",
        "duration_hours": "Charging duration (h)",
        "charge_power_kw": "Avg. charging power calc. (kW)",
        "charger_power_kw": "Charger/Wallbox power (kW)",
        "kwh": "kWh",
        "cost": "Cost",
        "location": "Location",
        "charger_type": "Charger type",
        "driver": "Driver",
        "license_plate": "License plate",
        "cost_center": "Cost center",
        "total_cost": "Total cost",
        "total_kwh": "Total kWh",
        "avg_charge_power_kw": "Average charging power (kW)",
        "total_charging_hours": "Total charging time (h)",
        "home": "Home",
        "external": "External",
        "all_locations": "All locations",
        "month_name": "Month",
        "month_year": "Month/Year",
        "export_period": "Period",
        "price_per_kwh": "Price/kWh",
        "row_num": "No.",
        "soc_start": "SOC Start (%)",
        "soc_end": "SOC End (%)",
        "odo_start": "Odometer start",
        "odo_end": "Odometer end",
        "meter_old": "Meter reading old",
        "meter_new": "Meter reading new",
        "total_sessions": "Number of sessions",
        "total_km": "Total km",
        # legacy keys kept for compatibility
        "department": "Department",
        "export_date": "Export date",
        "avg_charge_power": "Average charging power (kW)",
        "total_charging_time": "Total charging time",
        "signature": "Signature",
        "period": "Period",
    },
}

def t_export(key: str, lang: str = "de") -> str:
    """Translate an export label key to the given language."""
    lang = lang if lang in EXPORT_LABELS else "de"
    return EXPORT_LABELS[lang].get(key, EXPORT_LABELS["de"].get(key, key))

DATA_DIR      = Path(os.environ.get("DATA_DIR", "/data"))
DB_PATH       = DATA_DIR / "sessions.db"
EXPORT_DIR    = DATA_DIR / "exports"
TEMPLATE_PATH = DATA_DIR / "template.xlsx"
PRICE_PER_KWH = float(os.environ.get("PRICE_PER_KWH", "0.30"))

# ── Column keyword → field name mapping ───────────────────────────────────────
COLUMN_KEYWORDS = {
    "datum":            "date",
    "date":             "date",
    "start":            "start_time",
    "beginn":           "start_time",
    "ende":             "end_time",
    "end":              "end_time",
    "km start":         "odo_start",
    "km-start":         "odo_start",
    "start km":         "odo_start",
    "odometer":         "odo_start",
    "km ende":          "odo_end",
    "km-ende":          "odo_end",
    "end km":           "odo_end",
    "kilometerstand":   "odo_end",
    "kilome":           "odo_end",
    "kwh":              "kwh_charged",
    "geladen":          "kwh_charged",
    "energie":          "kwh_charged",
    "energy":           "kwh_charged",
    "lademenge":        "kwh_charged",
    "ladedauer":        "duration",
    "ladezeit":         "duration",
    "dauer":            "duration",
    "kosten":           "cost_eur",
    "cost":             "cost_eur",
    "betrag":           "cost_eur",
    "ladekosten":       "cost_eur",
    "soc start":        "soc_start",
    "soc anfang":       "soc_start",
    "soc ende":         "soc_end",
    "soc end":          "soc_end",
    "standort":         "location",
    "ort":              "location",
    "location":         "location",
    "nr":               "row_num",
    "#":                "row_num",
    "lfd":              "row_num",
    "zählerstand alt":  "meter_old",
    "zählerstand neu":  "meter_new",
    "zählerstand":      "meter_old",
    "charging_time":    "duration",
    "total_duration":   "duration",
    "duration_hours":   "duration_hours",
    "charge_power":     "charge_power_kw",
    "avg_power":        "charge_power_kw",
    "ladeleistung":     "charge_power_kw",
    "preis":            "price_per_kwh",
    "price":            "price_per_kwh",
    "preis/kwh":        "price_per_kwh",
    "preisquelle":      "price_source",
    "price_source":     "price_source",
    "vertrag":          "charging_contract_name",
    "contract":         "charging_contract_name",
    "ladetyp":          "charger_type",
    "charger_type":     "charger_type",
    "ladeart":          "charger_type",
}

FIELD_LABELS = {
    "date":        "Datum",
    "start_time":  "Start Uhrzeit",
    "end_time":    "Ende Uhrzeit",
    "odo_start":   "KM-Stand Start",
    "odo_end":     "KM-Stand Ende / Kilometerstand",
    "soc_start":   "SOC Start (%)",
    "soc_end":     "SOC Ende (%)",
    "kwh_charged": "Geladene kWh / Lademenge",
    "cost_eur":    "Kosten (€) / Ladekosten",
    "duration":    "Ladedauer",
    "meter_old":   "Zählerstand Alt (kWh)",
    "meter_new":   "Zählerstand Neu (kWh)",
    "location":    "Standort",
    "row_num":     "Nr.",
    "price_source":           "Preisquelle",
    "charging_contract_name": "Vertragsname",
    "charger_type":           "Ladeart",
    None:          "— nicht zugewiesen —",
}

NUM_FMT = {
    "date":        "DD.MM.YYYY",
    "start_time":  "HH:MM",
    "end_time":    "HH:MM",
    "odo_start":   '#,##0 "km"',
    "odo_end":     '#,##0 "km"',
    "soc_start":   '0"%"',
    "soc_end":     '0"%"',
    "kwh_charged": '0.00 "kWh"',
    "cost_eur":    '€#,##0.00',
    "duration":         '[h]:MM',
    "duration_hours":   '0.00 "h"',
    "charge_power_kw":  '0.00 "kW"',
    "meter_old":   '0',
    "meter_new":   '0',
    "row_num":     "0",
    "price_per_kwh":          '€0.0000',
    "charger_type":           "@",
    "price_source":           "@",
    "charging_contract_name": "@",
}

LOCATION_LABELS = {"home": "🏠 Zuhause", "extern": "⚡ Extern", "unknown": "—"}

# ── Header label → data key mapping for auto-fill ────────────────────────────
HEADER_KEYWORDS = {
    "abrechnungsmonat": "month_year",
    "gesamtkosten":     "total_cost",
    "kennzeichen":      "kennzeichen",
    "fahrer":           "fahrer",
    "abteilung":        "abteilung",
    "kostenstelle":     "kostenstelle",
    "kosten pro kw":    "price_per_kwh",
    "kosten pro kwh":   "price_per_kwh",
    "preis pro kwh":    "price_per_kwh",
    "monat":            "month_year",
}

def match_column(txt):
    """Match a column header text to a TABLE_FIELDS key.

    Uses TABLE_FIELDS synonyms (from template_fields.py) with best-confidence
    selection: exact match > substring > partial words. Longer synonyms win on
    ties so "preis/kwh" beats the shorter "kwh" substring of kwh_charged.
    Falls back to the legacy COLUMN_KEYWORDS dict only for exact keyword matches.
    """
    if not txt:
        return None
    norm = str(txt).lower().strip().rstrip(":").strip()
    if not norm:
        return None

    best_field = None
    best_conf = 0.0
    best_syn_len = 0

    for field, info in TABLE_FIELDS.items():
        for syn in info.get("synonyms", []):
            syn_l = syn.lower()
            if norm == syn_l:
                conf = 1.0
            elif syn_l in norm or norm in syn_l:
                conf = 0.8
            else:
                words = syn_l.split()
                if len(words) > 1 and all(w in norm for w in words):
                    conf = 0.7
                else:
                    continue
            # Higher confidence wins; tie broken by synonym length (longer = more specific)
            if conf > best_conf or (conf == best_conf and len(syn) > best_syn_len):
                best_field = field
                best_conf = conf
                best_syn_len = len(syn)

    # Tiebreaker for charge_power_kw vs charger_power_kw (installed rating wins when
    # header contains priority words like "kw/stunde", "wallbox", "11kw", …)
    if best_field == "charge_power_kw" and best_conf < 1.0:
        try:
            from template_fields import CHARGER_POWER_PRIORITY_KEYWORDS
            if any(kw in norm for kw in CHARGER_POWER_PRIORITY_KEYWORDS):
                best_field = "charger_power_kw"
        except ImportError:
            pass

    if best_field and best_conf >= 0.7:
        return best_field

    # Legacy fallback: exact keyword match only (avoids false substring positives)
    for kw, field in COLUMN_KEYWORDS.items():
        if norm == kw:
            return field

    return None

def fetch_sessions(year, month, location="all"):
    if not DB_PATH.exists(): return []
    con = sqlite3.connect(DB_PATH); con.row_factory = sqlite3.Row
    params = [f"{year:04d}-{month:02d}%"]
    where  = ["end_ts IS NOT NULL", "start_ts LIKE ?"]
    if location and location != "all":
        where.append("location = ?")
        params.append(location)
    rows = con.execute(
        f"SELECT * FROM sessions WHERE {' AND '.join(where)} ORDER BY start_ts",
        params
    ).fetchall()
    con.close(); return [dict(r) for r in rows]

def to_row(s, idx, lang="de"):
    dt_s = datetime.fromisoformat(s["start_ts"])
    dt_e = datetime.fromisoformat(s["end_ts"]) if s.get("end_ts") else None

    try:
        from services.pricing_service import get_session_duration_seconds
        total_seconds = get_session_duration_seconds(s)
    except Exception:
        total_seconds = int((dt_e - dt_s).total_seconds()) if dt_e else None
    duration = total_seconds / 86400 if total_seconds is not None else None
    duration_hours = round(total_seconds / 3600, 2) if total_seconds is not None else None

    # Determine charger_type: use stored value or derive from location
    charger_type = s.get("charger_type")
    if not charger_type or charger_type == "unknown":
        loc = s.get("location", "unknown")
        if loc == "home":
            charger_type = "AC"
        elif loc in ("extern", "external"):
            charger_type = "DC"

    row = {
        "row_num":      idx,
        "date":         dt_s.date(),
        "start_time":   dt_s.time(),
        "end_time":     dt_e.time() if dt_e else None,
        "duration":     duration,
        "duration_hours": duration_hours,
        "charge_power_kw":   None,  # calculated below
        "charger_power_kw":  s.get("charger_power_kw"),  # installed wallbox power
        "odo_start":    s.get("odo_start"),
        "odo_end":      s.get("odo_end"),
        "soc_start":    s.get("soc_start"),
        "soc_end":      s.get("soc_end"),
        "kwh_charged":  s.get("kwh_charged"),
        "cost_eur":     s.get("cost_eur"),
        "price_per_kwh": s.get("price_per_kwh"),
        "meter_old":    format_meter_value_kwh(s.get("meter_old")),
        "meter_new":    format_meter_value_kwh(s.get("meter_new")),
        "location":     LOCATION_LABELS.get(s.get("location", "unknown"), s.get("location", "—")),
        "charger_type": charger_type,
        "price_source":          s.get("price_source"),
        "charging_contract_name": s.get("charging_contract_name"),
        "charging_contract_id":  s.get("charging_contract_id"),
    }
    # Calculate charge_power_kw = kwh / duration_h
    kwh = s.get("kwh_charged")
    dur_h = row.get("duration_hours")
    if kwh is not None and dur_h and dur_h > 0:
        row["charge_power_kw"] = round(kwh / dur_h, 2)
    return row

# ── Built-in style helpers ────────────────────────────────────────────────────
C_HDR  = "1F4E79"; C_FG  = "FFFFFF"
C_SUM  = "D6E4F0"; C_ALT = "F2F7FB"
C_HOME = "E8F5E9"; C_EXT = "FFF8E1"
T = Side(style="thin", color="A9C4D8")
BRD = Border(left=T, right=T, top=T, bottom=T)

def cs(ws, r, c, v, bold=False, bg=None, fg="000000", nf=None, al="left"):
    cell = ws.cell(row=r, column=c, value=v)
    cell.font      = Font(name="Arial", bold=bold, color=fg, size=10)
    cell.alignment = Alignment(horizontal=al, vertical="center")
    cell.border    = BRD
    if bg: cell.fill = PatternFill("solid", start_color=bg.lstrip("#"))
    if nf: cell.number_format = nf
    return cell

def row_bg(s):
    loc = s.get("location", "unknown")
    if loc == "home":   return C_HOME
    if loc == "extern": return C_EXT
    return C_ALT

def safe_set(cell, value):
    try: cell.value = value
    except (AttributeError, TypeError): pass

# ── Header section auto-fill ──────────────────────────────────────────────────
def fill_header_section(ws, data_start_row, header_data):
    """Scan rows before data_start_row for label-value pairs and auto-fill."""
    for row_idx in range(1, data_start_row):
        for cell in ws[row_idx]:
            if not cell.value or isinstance(cell, MergedCell):
                continue
            key = str(cell.value).lower().strip().rstrip(':').strip()

            # Match against known header labels
            matched_field = None
            for kw, field in HEADER_KEYWORDS.items():
                if kw in key:
                    matched_field = field
                    break

            if matched_field and matched_field in header_data and header_data[matched_field] is not None:
                # Fill the next writable cell to the right
                target = ws.cell(row=cell.row, column=cell.column + 1)
                if not isinstance(target, MergedCell):
                    safe_set(target, header_data[matched_field])
                    if matched_field == "total_cost":
                        try: target.number_format = '€#,##0.00'
                        except Exception: pass
                    elif matched_field == "price_per_kwh":
                        try: target.number_format = '€0.00'
                        except Exception: pass

            # Special: cell ending with ", den" → fill date in next cell
            if key.endswith(", den") or key.endswith(",den"):
                target = ws.cell(row=cell.row, column=cell.column + 1)
                if not isinstance(target, MergedCell) and not target.value:
                    safe_set(target, datetime.today().date())
                    try: target.number_format = 'DD.MM.YYYY'
                    except Exception: pass

# ── Header values computation ─────────────────────────────────────────────────
def compute_header_values(sessions, year, month, header_info=None, lang="de"):
    """Compute all HEADER_FIELDS values from sessions + config."""
    import calendar as _cal
    if header_info is None:
        header_info = {}
    home_s     = [s for s in sessions if s.get("location") == "home"]
    ext_s      = [s for s in sessions if s.get("location") == "extern"]
    home_kwh   = sum(s.get("kwh_charged") or 0 for s in home_s)
    ext_kwh    = sum(s.get("kwh_charged") or 0 for s in ext_s)
    home_cost  = sum(s.get("cost_eur")    or 0 for s in home_s)
    ext_cost   = sum(s.get("cost_eur")    or 0 for s in ext_s)
    total_kwh  = sum(s.get("kwh_charged") or 0 for s in sessions)
    total_cost = sum(s.get("cost_eur")    or 0 for s in sessions)
    n = len(sessions)
    # Sum per-session km differences where both odo values are present
    total_km = 0
    for s in sessions:
        odo_s = s.get("odo_start")
        odo_e = s.get("odo_end")
        if odo_s is not None and odo_e is not None:
            diff = odo_e - odo_s
            if diff > 0:
                total_km += diff
    if total_km == 0 and n:
        # Fallback: use first/last odo readings
        odo_end_val   = sessions[-1].get("odo_end")   or 0
        odo_start_val = sessions[0].get("odo_start")  or 0
        total_km = max(0, odo_end_val - odo_start_val)
    avg_cons = None
    if total_km and total_km > 0:
        avg_cons = round(total_kwh / total_km * 100, 2)

    # Total charging hours
    total_charging_hours = 0.0
    for s in sessions:
        if s.get("start_ts") and s.get("end_ts"):
            try:
                dt_s = datetime.fromisoformat(s["start_ts"])
                dt_e = datetime.fromisoformat(s["end_ts"])
                total_charging_hours += (dt_e - dt_s).total_seconds() / 3600
            except Exception:
                pass
    total_charging_hours = round(total_charging_hours, 2)
    avg_charge_power_kw = None
    if total_charging_hours > 0 and total_kwh > 0:
        avg_charge_power_kw = round(total_kwh / total_charging_hours, 2)

    # Localized month name and period
    month_name = MONTH_NAMES.get(lang, MONTH_NAMES["de"])[month]
    month_year = f"{month_name} {year}"
    _now_utc = datetime.now(timezone.utc)
    export_date_str = _now_utc.strftime(DATE_FORMATS.get(lang, "%d.%m.%Y"))
    last_day = _cal.monthrange(year, month)[1]
    if lang == "en":
        export_period = f"{year}-{month:02d}-01 – {year}-{month:02d}-{last_day:02d}"
    else:
        export_period = f"01.{month:02d}.{year} – {last_day:02d}.{month:02d}.{year}"

    return {
        "fahrer":                   header_info.get("fahrer") or None,
        "kennzeichen":              header_info.get("kennzeichen") or None,
        "abteilung":                header_info.get("abteilung") or None,
        "kostenstelle":             header_info.get("kostenstelle") or None,
        "month_year":               month_year,
        "month_name":               month_name,
        "export_date":              export_date_str,
        "export_period":            export_period,
        "total_sessions":           n,
        "total_kwh":                round(total_kwh, 2),
        "total_cost":               total_cost,
        "total_home_kwh":           home_kwh,
        "total_external_kwh":       ext_kwh,
        "total_home_cost":          home_cost,
        "total_external_cost":      ext_cost,
        "total_km":                 total_km,
        "avg_consumption_kwh_100km": avg_cons,
        "meter_start_value":        format_meter_value_kwh(sessions[0].get("meter_old")) if n else None,
        "meter_end_value":          format_meter_value_kwh(sessions[-1].get("meter_new")) if n else None,
        "price_per_kwh":            header_info.get("price_per_kwh") or None,
        "total_charging_hours":     total_charging_hours,
        "avg_charge_power_kw":      avg_charge_power_kw,
    }


# ── Placeholder fill helper ───────────────────────────────────────────────────
def _fill_placeholders(wb, header_vals, include_signature=False, sig_img=None):
    """Replace all {{key}} placeholders in all worksheets with values from header_vals.

    Supported placeholders: {{fahrer}}, {{kennzeichen}}, {{abteilung}}, {{kostenstelle}},
    {{month_name}}, {{month_year}}, {{export_date}}, {{export_period}}, {{price_per_kwh}},
    {{total_sessions}}, {{total_kwh}}, {{total_cost}}, {{total_home_kwh}}, {{total_external_kwh}},
    {{total_home_cost}}, {{total_external_cost}}, {{total_km}}, {{total_charging_hours}},
    {{avg_charge_power_kw}}, {{avg_consumption_kwh_100km}}, {{meter_start_value}},
    {{meter_end_value}}, {{signature}}
    """
    ph_pattern = _PH_PATTERN

    # Fields that should be formatted as 2-decimal floats
    FLOAT2_FIELDS = {
        "total_kwh", "total_cost", "total_home_kwh", "total_external_kwh",
        "total_home_cost", "total_external_cost", "total_charging_hours",
        "avg_charge_power_kw", "avg_consumption_kwh_100km",
        "price_per_kwh",
        # meter_start_value / meter_end_value excluded — already rounded to int
    }

    sig_cell_ref = None  # Will be set if {{signature}} placeholder is found

    for ws_sheet in wb.worksheets:
        for row in ws_sheet.iter_rows():
            for cell in row:
                if isinstance(cell, MergedCell):
                    continue
                val = cell.value
                if not isinstance(val, str) or '{{' not in val:
                    continue

                # Handle {{signature}} placeholder
                if '{{signature}}' in val:
                    if include_signature:
                        # Record cell address for later signature insertion, then clear text
                        from openpyxl.utils import get_column_letter as _gcl2
                        sig_cell_ref = f"{_gcl2(cell.column)}{cell.row}"
                        try:
                            cell.value = val.replace('{{signature}}', '')
                        except Exception:
                            pass
                    else:
                        try:
                            cell.value = val.replace('{{signature}}', '')
                        except Exception:
                            pass
                    continue

                def _replace_ph(m, _hv=header_vals, _f2=FLOAT2_FIELDS):
                    key = m.group(1)
                    v = _hv.get(key)
                    if v is None:
                        return ""
                    if key in _f2:
                        try:
                            return f"{float(v):.2f}"
                        except (ValueError, TypeError):
                            pass
                    return str(v)

                new_val = ph_pattern.sub(_replace_ph, val)
                if new_val != val:
                    try:
                        cell.value = new_val
                    except Exception:
                        pass

    return sig_cell_ref  # Return detected signature cell ref if any


# ── Template-based export ─────────────────────────────────────────────────────
def export_with_template(year, month, sessions, location, col_override=None, start_row=None, header_row=None, header_info=None,
                          cell_mapping=None, sheet=None,
                          include_signature=False, signature_path=None, signature_mapping=None,
                          lang="de", footer_start_row=None):
    if header_info is None:
        header_info = {}
    if cell_mapping is None:
        cell_mapping = {}
    ml     = datetime(year, month, 1).strftime("%B %Y")
    suffix = f"_{location}" if location != "all" else ""
    EXPORT_DIR.mkdir(parents=True, exist_ok=True)
    out = EXPORT_DIR / f"EV_Ladeprotokoll_{year:04d}-{month:02d}{suffix}.xlsx"
    shutil.copy(TEMPLATE_PATH, out)
    wb = openpyxl.load_workbook(out, keep_vba=False)
    # Select sheet
    if sheet and sheet in wb.sheetnames:
        ws = wb[sheet]
    else:
        ws = wb.active

    # ── Build column map from explicit mapping ────────────────────────────────
    # Supports both formats: {"1": "date"} and {"1": {"field": "date", ...}}
    col_map = {}
    if col_override and isinstance(col_override, dict):
        for col_str, field_info in col_override.items():
            try:
                col_idx = int(col_str)
            except (ValueError, TypeError):
                continue
            field = field_info.get("field") if isinstance(field_info, dict) else field_info
            if field:
                col_map[col_idx] = field

    # ── Fallback: auto-detect header row if no explicit column mapping ────────
    # When header_row is saved (from a previous template analysis), scan exactly
    # that row. Otherwise call analyze_template() for robust detection, which
    # requires >=2 actual session-column matches and won't mistake summary rows
    # (e.g. "Ladezeit gesamt | Gesamtkosten") for the data table header.
    if not col_map:
        if header_row:
            hr = int(header_row)
            for cnum in range(1, (ws.max_column or 40) + 1):
                try:
                    cell = ws.cell(row=hr, column=cnum)
                    if isinstance(cell, MergedCell) or cell.value is None:
                        continue
                    f = match_column(cell.value)
                    if f:
                        col_map[cnum] = f
                except Exception:
                    pass
        else:
            # No header_row saved — use template_analyzer for robust detection
            try:
                from template_analyzer import analyze_template as _at
                _analysis = _at(TEMPLATE_PATH)
                if _analysis.get("ok") and _analysis.get("column_mapping"):
                    for col_str, info in _analysis["column_mapping"].items():
                        try:
                            col_idx = int(col_str)
                        except (ValueError, TypeError):
                            continue
                        field = info.get("field") if isinstance(info, dict) else info
                        if field:
                            col_map[col_idx] = field
                    if not start_row and _analysis.get("start_row"):
                        start_row = _analysis["start_row"]
                elif _analysis.get("ok") and _analysis.get("header_row"):
                    # analyzer detected header but no column mapping yet — scan that row
                    hr = _analysis["header_row"]
                    for cnum in range(1, (ws.max_column or 40) + 1):
                        try:
                            cell = ws.cell(row=hr, column=cnum)
                            if isinstance(cell, MergedCell) or cell.value is None:
                                continue
                            f = match_column(cell.value)
                            if f:
                                col_map[cnum] = f
                        except Exception:
                            pass
                    if not start_row:
                        start_row = hr + 1
            except Exception:
                pass

    # ── Determine data start row ──────────────────────────────────────────────
    if start_row:
        ds = int(start_row)
    elif header_row:
        ds = int(header_row) + 1
    else:
        ds = (ws.max_row or 1)

    if not col_map:
        col_map = {1:"row_num",2:"date",3:"start_time",4:"end_time",
                   5:"odo_start",6:"odo_end",7:"kwh_charged",8:"cost_eur",9:"location"}

    max_row = ws.max_row or 0

    # capture template row styles from first data row
    tstyles = {}
    if max_row >= ds:
        for ci in col_map:
            c = ws.cell(row=ds, column=ci)
            try:
                tstyles[ci] = {
                    "font":      copy.copy(c.font)      if c.font      else None,
                    "fill":      copy.copy(c.fill)      if c.fill      else None,
                    "border":    copy.copy(c.border)    if c.border    else None,
                    "alignment": copy.copy(c.alignment) if c.alignment else None,
                }
            except Exception:
                pass

    # meter readings: use stored values if available, else calculate from meter_start_value
    has_stored = any(s.get("meter_old") is not None for s in sessions)
    if not has_stored:
        meter_start = float(header_info.get("meter_start_value") or 0)
        cumulative = meter_start
        enriched = []
        for s in sessions:
            s = dict(s)
            s["meter_old"] = round(cumulative, 3)
            cumulative += float(s.get("kwh_charged") or 0)
            s["meter_new"] = round(cumulative, 3)
            enriched.append(s)
        sessions = enriched

    # Determine footer_start_row:
    # - Explicit value from saved config → use directly (no heuristic).
    # - Otherwise call _detect_footer_start_row() which handles both
    #   SUM-formula footers and plain-text footers via a 2-phase algorithm.
    if footer_start_row is not None:
        footer_start_row = int(footer_start_row)
    else:
        footer_start_row = _detect_footer_start_row(ws, ds, max_row or 0, col_map)

    template_data_rows = max(footer_start_row - ds, 0)

    # If more sessions than template rows, insert rows before the footer to make room
    n_extra = len(sessions) - template_data_rows
    if n_extra > 0 and footer_start_row <= (ws.max_row or 0):
        style_row = footer_start_row - 1
        _rstyles = {}
        for ci in col_map:
            try:
                _c = ws.cell(row=style_row, column=ci)
                _rstyles[ci] = {
                    "font":          copy.copy(_c.font)      if _c.font      else None,
                    "fill":          copy.copy(_c.fill)      if _c.fill      else None,
                    "border":        copy.copy(_c.border)    if _c.border    else None,
                    "alignment":     copy.copy(_c.alignment) if _c.alignment else None,
                    "number_format": _c.number_format,
                }
            except Exception:
                pass
        # Capture ALL per-row formulas from template data row (ds) before
        # insert_rows() shifts rows.  This includes UNMAPPED columns such as
        # auto-computed fields (e.g. C3 = =B3*0.30 where C is not in col_map).
        tformulas_all = {}
        for _tfc in range(1, (ws.max_column or 40) + 1):
            try:
                _tc = ws.cell(row=ds, column=_tfc)
                if isinstance(_tc.value, str) and _tc.value.startswith("="):
                    if not _is_range_formula(_tc.value):   # only per-row formulas
                        tformulas_all[_tfc] = _tc.value
            except Exception:
                pass
        old_last_data = footer_start_row - 1  # before insertion
        ws.insert_rows(footer_start_row, n_extra)
        # After insert_rows, footer is now at footer_start_row + n_extra
        new_footer_start = footer_start_row + n_extra
        for _ei in range(n_extra):
            _r = footer_start_row + _ei
            # Apply styles from last template data row (mapped columns)
            for ci, _st in _rstyles.items():
                try:
                    _cell = ws.cell(row=_r, column=ci)
                    if _st["font"]:      _cell.font      = _st["font"]
                    if _st["fill"]:      _cell.fill      = _st["fill"]
                    if _st["border"]:    _cell.border    = _st["border"]
                    if _st["alignment"]: _cell.alignment = _st["alignment"]
                    if _st["number_format"]: _cell.number_format = _st["number_format"]
                except Exception:
                    pass
            # Copy per-row formulas from all columns (mapped AND unmapped)
            for ci, formula in tformulas_all.items():
                try:
                    cell = ws.cell(row=_r, column=ci)
                    if isinstance(cell, MergedCell):
                        continue
                    adjusted = _adjust_formula_rows(formula, _r - ds)
                    cell.value = adjusted
                except Exception:
                    pass
        # Extend SUM/range formulas in footer rows to cover new data rows
        _new_last = old_last_data + n_extra
        for _sr in range(new_footer_start, (ws.max_row or 0) + 1):
            for _sc in range(1, (ws.max_column or 20) + 1):
                try:
                    _cell = ws.cell(row=_sr, column=_sc)
                    if not (isinstance(_cell.value, str) and _cell.value.startswith("=")):
                        continue
                    _cell.value = _re_formula.sub(
                        rf'([A-Za-z]+\$?){old_last_data}(?=[,):+\-*/\s]|$)',
                        lambda m: m.group(1) + str(_new_last),
                        _cell.value
                    )
                except Exception:
                    pass
        max_row = ws.max_row or 0

    # Clear old data: ONLY mapped columns, ONLY rows before footer, skip formula cells
    clear_end = min(ds + max(len(sessions), template_data_rows) - 1, footer_start_row - 1)
    for r in range(ds, clear_end + 1):
        for ci in col_map:
            try:
                cell = ws.cell(row=r, column=ci)
                if isinstance(cell, MergedCell):
                    continue
                if isinstance(cell.value, str) and cell.value.startswith("="):
                    continue
                safe_set(cell, None)
            except Exception:
                pass

    # write data rows
    for i, s in enumerate(sessions):
        rd = to_row(s, i + 1); tr = ds + i
        for ci, field in col_map.items():
            cell = ws.cell(row=tr, column=ci)
            safe_set(cell, rd.get(field))
            if ci in tstyles:
                st = tstyles[ci]
                try:
                    if st["font"]:      cell.font      = st["font"]
                    if st["fill"]:      cell.fill      = st["fill"]
                    if st["border"]:    cell.border    = st["border"]
                    if st["alignment"]: cell.alignment = st["alignment"]
                except Exception:
                    pass
            try:
                if field in NUM_FMT: cell.number_format = NUM_FMT[field]
            except Exception:
                pass

    # Compute full header values
    hv = compute_header_values(sessions, year, month, header_info, lang=lang)

    # auto-fill header section (Abrechnungsmonat, Kennzeichen, Fahrer, etc.)
    # Use hv merged with legacy header_data keys for backward compat
    header_data = dict(hv)
    fill_header_section(ws, ds, header_data)

    # Replace {{placeholder}} in any cell text
    _ph_sig_cell = _fill_placeholders(wb, hv, include_signature=include_signature, sig_img=None)

    # Write cell_mapping: { "B4": "kennzeichen" or {"field": "kennzeichen", ...} }
    if cell_mapping and isinstance(cell_mapping, dict):
        col_letter_re = _COL_ADDR_RE
        for addr, field_info in cell_mapping.items():
            field = field_info if isinstance(field_info, str) else field_info.get("field")
            if not field or field not in hv:
                continue
            value = hv.get(field)
            if value is None:
                continue
            m = col_letter_re.match(str(addr))
            if not m:
                continue
            try:
                col_letters = m.group(1).upper()
                row_num = int(m.group(2))
                # Convert column letters to index
                col_idx = 0
                for ch in col_letters:
                    col_idx = col_idx * 26 + (ord(ch) - 64)
                cell = ws.cell(row=row_num, column=col_idx)
                if isinstance(cell, MergedCell):
                    # Find merge top-left
                    for mr in ws.merged_cells.ranges:
                        if mr.min_row <= row_num <= mr.max_row and mr.min_col <= col_idx <= mr.max_col:
                            cell = ws.cell(row=mr.min_row, column=mr.min_col)
                            break
                # Skip formula cells
                if isinstance(cell.value, str) and cell.value.startswith("="):
                    continue
                safe_set(cell, value)
            except Exception:
                pass

    # ── Signature insertion ──────────────────────────────────────────────────
    if include_signature and signature_path and signature_mapping:
        try:
            from openpyxl.drawing.image import Image as XLImage
            # Support both old "cell" key and new "anchor_cell" key
            sig_cell = signature_mapping.get("anchor_cell") or signature_mapping.get("cell")
            # Use cell detected from {{signature}} placeholder if present
            if not sig_cell and _ph_sig_cell:
                sig_cell = _ph_sig_cell
            # Fallback: scan for any remaining {{signature}} text
            if not sig_cell:
                for row in ws.iter_rows():
                    for cell in row:
                        if cell.value and "{{signature}}" in str(cell.value):
                            from openpyxl.utils import get_column_letter as _gcl
                            sig_cell = f"{_gcl(cell.column)}{cell.row}"
                            cell.value = None  # clear placeholder
                            break
                    if sig_cell:
                        break
            if sig_cell:
                img = XLImage(signature_path)
                sig_w = int(signature_mapping.get("width", 220))
                sig_h = int(signature_mapping.get("height", 80))
                # keep aspect ratio if requested
                if signature_mapping.get("keep_aspect_ratio", True):
                    try:
                        from PIL import Image as _PILImage
                        with _PILImage.open(signature_path) as _pimg:
                            iw, ih = _pimg.size
                        if iw > 0 and ih > 0:
                            ratio = min(sig_w / iw, sig_h / ih)
                            sig_w = int(iw * ratio)
                            sig_h = int(ih * ratio)
                    except Exception:
                        pass
                img.width  = sig_w
                img.height = sig_h
                # Apply offset using OneCellAnchor if offset values are provided
                offset_x = int(signature_mapping.get("offset_x", 0))
                offset_y = int(signature_mapping.get("offset_y", 0))
                placed = False
                if offset_x or offset_y:
                    try:
                        import re as _re_sig
                        from openpyxl.utils import column_index_from_string as _col_idx
                        try:
                            from openpyxl.drawing.spreadsheet_drawing import OneCellAnchor, AnchorMarker
                        except ImportError:
                            from openpyxl.drawing.anchor import OneCellAnchor, AnchorMarker
                        from openpyxl.drawing.xdr import XDRPositiveSize2D
                        try:
                            from openpyxl.utils.units import pixels_to_EMU
                        except (ImportError, Exception):
                            def pixels_to_EMU(px): return int(px * 9525)
                        m2 = _re_sig.match(r'^([A-Za-z]+)(\d+)$', sig_cell)
                        if m2:
                            col_str = m2.group(1).upper()
                            row_num2 = int(m2.group(2)) - 1  # 0-based
                            col_num2 = _col_idx(col_str) - 1  # 0-based
                            marker = AnchorMarker(
                                col=col_num2, colOff=pixels_to_EMU(offset_x),
                                row=row_num2, rowOff=pixels_to_EMU(offset_y)
                            )
                            ext = XDRPositiveSize2D(
                                cx=pixels_to_EMU(sig_w),
                                cy=pixels_to_EMU(sig_h)
                            )
                            anchor = OneCellAnchor(_from=marker, ext=ext)
                            img.anchor = anchor
                            ws.add_image(img)
                            placed = True
                    except Exception as _anchor_err:
                        import logging as _log_sig
                        _log_sig.getLogger(__name__).warning(
                            "Signatur-Offset fehlgeschlagen, Fallback auf einfaches Einfügen: %s", _anchor_err)
                        placed = False
                if not placed:
                    ws.add_image(img, sig_cell)
        except Exception as e:
            import logging as _log_sig2
            _log_sig2.getLogger(__name__).warning("Signatur-Einfügung fehlgeschlagen: %s", e)
    # ── End signature ────────────────────────────────────────────────────────

    wb.save(out)
    print(f"✅ Template-Export: {out} ({len(sessions)} Sessions)")
    return str(out)

# ── Built-in export ───────────────────────────────────────────────────────────
def export_builtin(year, month, sessions, location, config=None, lang="de"):
    month_name = MONTH_NAMES.get(lang, MONTH_NAMES["de"])[month]
    ml         = f"{month_name} {year}"
    loc_label  = {
        "all":    t_export("all_locations", lang),
        "home":   "🏠 " + t_export("home", lang),
        "extern": "⚡ " + t_export("external", lang),
    }.get(location, t_export("all_locations", lang))
    suffix    = f"_{location}" if location != "all" else ""
    EXPORT_DIR.mkdir(parents=True, exist_ok=True)

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = t_export("charging_log", lang)
    ws.freeze_panes = "A3"

    # Title (16 columns: A–P)
    ws.merge_cells("A1:P1")
    t = ws["A1"]
    t.value     = f"{t_export('charging_log', lang)} – {ml}  |  {loc_label}"
    t.font      = Font(name="Arial", bold=True, size=14, color=C_FG)
    t.fill      = PatternFill("solid", start_color=C_HDR)
    t.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 28

    hdrs = [
        t_export("row_num", lang),
        t_export("date", lang),
        t_export("start", lang),
        t_export("end", lang),
        t_export("duration_hours", lang),
        t_export("charger_type", lang),
        t_export("odo_start", lang),
        t_export("odo_end", lang),
        t_export("soc_start", lang),
        t_export("soc_end", lang),
        t_export("kwh", lang),
        t_export("price_per_kwh", lang),
        t_export("cost", lang),
        t_export("location", lang),
        "Preisquelle" if lang != "en" else "Price source",
        "Vertragsname" if lang != "en" else "Contract name",
    ]
    for col, h in enumerate(hdrs, 1):
        cs(ws, 2, col, h, bold=True, bg="#"+C_HDR, fg=C_FG, al="center")

    ds = 3
    for i, s in enumerate(sessions):
        r  = ds + i
        rd = to_row(s, i + 1, lang=lang)
        bg = row_bg(s)
        cs(ws,r,1, rd["row_num"],         bg=bg, al="center")
        cs(ws,r,2, rd["date"],            bg=bg, nf="DD.MM.YYYY", al="center")
        cs(ws,r,3, rd["start_time"],      bg=bg, nf="HH:MM", al="center")
        cs(ws,r,4, rd["end_time"],        bg=bg, nf="HH:MM", al="center")
        cs(ws,r,5, rd["duration_hours"],  bg=bg, nf='0.00 "h"', al="right")
        cs(ws,r,6, rd["charger_type"],    bg=bg, al="center")
        cs(ws,r,7, rd["odo_start"],       bg=bg, nf='#,##0 "km"', al="right")
        cs(ws,r,8, rd["odo_end"],         bg=bg, nf='#,##0 "km"', al="right")
        cs(ws,r,9, rd["soc_start"],       bg=bg, nf='0"%"', al="right")
        cs(ws,r,10,rd["soc_end"],         bg=bg, nf='0"%"', al="right")
        cs(ws,r,11,rd["kwh_charged"],     bg=bg, nf='0.00 "kWh"', al="right")
        cs(ws,r,12,rd["price_per_kwh"],   bg=bg, nf='€0.0000', al="right")
        cs(ws,r,13,rd["cost_eur"],        bg=bg, nf='€#,##0.00', al="right")
        cs(ws,r,14,rd["location"],        bg=bg, al="center")
        cs(ws,r,15,rd["price_source"],    bg=bg, al="left")
        cs(ws,r,16,rd["charging_contract_name"], bg=bg, al="left")

    n = len(sessions)
    if n:
        tr = ds + n
        for col in range(1, 17):
            v  = ("Σ" if col==1
                  else f"=SUM(K{ds}:K{ds+n-1})" if col==11
                  else f"=SUM(M{ds}:M{ds+n-1})" if col==13
                  else "")
            nf = ('0.00 "kWh"' if col==11 else '€#,##0.00' if col==13 else None)
            cs(ws, tr, col, v, bold=True, bg="#"+C_SUM, nf=nf,
               al="center" if col==1 else "right")

    # column widths
    for col, w in enumerate([5,13,8,8,8,8,13,13,10,10,13,10,12,14,14,20],1):
        ws.column_dimensions[get_column_letter(col)].width = w

    # Legend
    leg = ds + n + 2
    ws.merge_cells(f"A{leg}:D{leg}")
    home_lbl = "🏠 " + t_export("home", lang)
    ext_lbl  = "⚡ " + t_export("external", lang)
    c = ws[f"A{leg}"]; c.value=f"{home_lbl} = Grün"; c.font=Font(name="Arial",size=9,color="2E7D32")
    ws.merge_cells(f"E{leg}:H{leg}")
    c = ws[f"E{leg}"]; c.value=f"{ext_lbl} = Gelb"; c.font=Font(name="Arial",size=9,color="F57F17")

    # ── Summary sheet ─────────────────────────────────────────────────────────
    summary_title = t_export("summary", lang)
    ws2 = wb.create_sheet(summary_title)
    ws2.merge_cells("A1:B1")
    h = ws2["A1"]
    h.value     = f"{summary_title} – {ml} – {loc_label}"
    h.font      = Font(name="Arial", bold=True, size=13, color=C_FG)
    h.fill      = PatternFill("solid", start_color=C_HDR)
    h.alignment = Alignment(horizontal="center", vertical="center")
    ws2.row_dimensions[1].height = 24

    home_s    = [s for s in sessions if s.get("location")=="home"]
    ext_s     = [s for s in sessions if s.get("location")=="extern"]
    home_kwh  = sum(s.get("kwh_charged") or 0 for s in home_s)
    ext_kwh   = sum(s.get("kwh_charged") or 0 for s in ext_s)
    home_cost = sum(s.get("cost_eur")    or 0 for s in home_s)
    ext_cost  = sum(s.get("cost_eur")    or 0 for s in ext_s)
    total_km  = ((sessions[-1].get("odo_end") or 0) - (sessions[0].get("odo_start") or 0)) if n else 0
    total_kwh = sum(s.get("kwh_charged") or 0 for s in sessions)
    total_cost_all = sum(s.get("cost_eur") or 0 for s in sessions)
    verbrauch = round(total_kwh / total_km * 100, 1) if total_km > 0 else None
    hv2 = compute_header_values(sessions, year, month, lang=lang)
    total_h_val = round(hv2.get("total_charging_hours") or 0, 2)
    avg_pwr_val = hv2.get("avg_charge_power_kw")

    if lang == "en":
        rows2 = [
            ("Total sessions",          n,              None),
            (f"  🏠 {t_export('home', lang)}",  len(home_s), None),
            (f"  ⚡ {t_export('external', lang)}", len(ext_s), None),
            ("Total kWh",               total_kwh,      '0.00 "kWh"'),
            (f"  🏠 {t_export('home', lang)} kWh", home_kwh, '0.00 "kWh"'),
            (f"  ⚡ {t_export('external', lang)} kWh", ext_kwh, '0.00 "kWh"'),
            ("Total cost",              total_cost_all, "€#,##0.00"),
            (f"  🏠 {t_export('home', lang)} cost", home_cost, "€#,##0.00"),
            ("Odometer start",          sessions[0].get("odo_start") if n else "-", '#,##0 "km"'),
            ("Odometer end",            sessions[-1].get("odo_end")  if n else "-", '#,##0 "km"'),
            ("Total km",                total_km,       '#,##0 "km"'),
            ("Consumption",             verbrauch,      '0.0 "kWh/100km"'),
            ("Total charging time",     total_h_val,    '0.00 "h"'),
        ]
        if avg_pwr_val is not None:
            rows2.append(("Avg. charging power", avg_pwr_val, '0.00 "kW"'))
    else:
        rows2 = [
            ("Ladevorgänge gesamt",     n,              None),
            (f"  🏠 Zuhause",           len(home_s),    None),
            (f"  ⚡ Extern",            len(ext_s),     None),
            ("Gesamt kWh",              total_kwh,      '0.00 "kWh"'),
            (f"  🏠 Zuhause kWh",       home_kwh,       '0.00 "kWh"'),
            (f"  ⚡ Extern kWh",        ext_kwh,        '0.00 "kWh"'),
            ("Gesamtkosten",            total_cost_all, "€#,##0.00"),
            (f"  🏠 Zuhause Kosten",    home_cost,      "€#,##0.00"),
            ("KM Monatsanfang",         sessions[0].get("odo_start") if n else "-", '#,##0 "km"'),
            ("KM Monatsende",           sessions[-1].get("odo_end")  if n else "-", '#,##0 "km"'),
            ("Gefahrene KM",            total_km,       '#,##0 "km"'),
            ("Verbrauch",               verbrauch,      '0.0 "kWh/100km"'),
            ("Ladezeit gesamt",         total_h_val,    '0.00 "h"'),
        ]
        if avg_pwr_val is not None:
            rows2.append(("Ø Ladeleistung", avg_pwr_val, '0.00 "kW"'))
    for ri, (lbl, val, fmt_) in enumerate(rows2, 3):
        bg = "#"+C_ALT if ri % 2 == 0 else None
        cs(ws2, ri, 1, lbl, bold=not lbl.startswith("  "), bg=bg, al="left")
        cs(ws2, ri, 2, val, bg=bg, nf=fmt_, al="right")
    ws2.column_dimensions["A"].width = 26
    ws2.column_dimensions["B"].width = 20

    out = EXPORT_DIR / f"EV_Ladeprotokoll_{year:04d}-{month:02d}{suffix}.xlsx"
    wb.save(out)
    print(f"Export: {out} ({n} Sessions)")
    return str(out)

# ── Entry point ───────────────────────────────────────────────────────────────
def export(year, month, location="all", col_override=None, start_row=None, header_row=None, header_info=None,
           cell_mapping=None, sheet=None,
           include_signature=False, signature_path=None, signature_mapping=None,
           lang="de", return_warnings=False, footer_start_row=None):
    """Export sessions to XLSX. Returns bytes or (bytes, warnings) if return_warnings=True."""
    sessions = fetch_sessions(year, month, location)
    warnings_list = []

    # Collect warnings
    warnings_list.append(f"Sprache: {lang}")
    if include_signature and not signature_path:
        warnings_list.append("Keine Signatur gespeichert")
    if include_signature and signature_mapping and not (
        signature_mapping.get("anchor_cell") or signature_mapping.get("cell")
    ):
        warnings_list.append("Signaturposition nicht definiert")

    if TEMPLATE_PATH.exists():
        path = export_with_template(year, month, sessions, location, col_override, start_row,
                                    header_row=header_row,
                                    header_info=header_info,
                                    cell_mapping=cell_mapping, sheet=sheet,
                                    include_signature=include_signature,
                                    signature_path=signature_path,
                                    signature_mapping=signature_mapping,
                                    lang=lang,
                                    footer_start_row=footer_start_row)
    else:
        path = export_builtin(year, month, sessions, location, lang=lang)

    # Read file as bytes
    with open(path, "rb") as f:
        xlsx_bytes = f.read()

    if return_warnings:
        return xlsx_bytes, warnings_list
    return xlsx_bytes


def preview_export(year, month, location="all", col_override=None, start_row=None, header_row=None,
                   header_info=None, cell_mapping=None, sheet=None,
                   include_signature=False, signature_path=None, signature_mapping=None,
                   lang="de") -> dict:
    """Generate a preview dict by calling the real export() and reading back the XLSX.
    Returns {'ok': True, 'sheets': [...], 'warnings': [...]} with max 200 rows / 30 cols per sheet."""
    import io as _io
    warnings_out = []
    try:
        xlsx_bytes, export_warnings = export(
            year, month, location,
            col_override=col_override, start_row=start_row, header_row=header_row,
            header_info=header_info, cell_mapping=cell_mapping, sheet=sheet,
            include_signature=include_signature, signature_path=signature_path,
            signature_mapping=signature_mapping,
            lang=lang, return_warnings=True,
        )
        warnings_out.extend(export_warnings)
    except Exception as e:
        warnings_out.append(f"Export fehlgeschlagen: {e}")
        return {"ok": False, "sheets": [], "warnings": warnings_out}

    # Read generated XLSX and build grid
    sheets_out = []
    try:
        wb = openpyxl.load_workbook(_io.BytesIO(xlsx_bytes), data_only=True)
        for ws in wb.worksheets:
            rows_out = []
            for ri, row in enumerate(ws.iter_rows(max_row=200, max_col=30, values_only=True)):
                cells = []
                for val in row:
                    if val is None:
                        cells.append("")
                    elif isinstance(val, (int, float)):
                        cells.append(val)
                    else:
                        cells.append(str(val))
                rows_out.append(cells)
            sheets_out.append({"name": ws.title, "rows": rows_out})
    except Exception as e:
        warnings_out.append(f"Vorschau-Lesen fehlgeschlagen: {e}")

    return {
        "ok": True,
        "sheets": sheets_out,
        "warnings": warnings_out,
    }

def export_multi_month_bytes(periods_sessions, loc_filter="all", config=None, lang="de",
                              include_signature=False, signature_path=None, signature_mapping=None):
    """Create a multi-sheet XLSX with one sheet per month.
    periods_sessions: list of (period_info, sessions) tuples.
    Returns (bytes, warnings).
    """
    import io
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment

    warnings_list = []
    wb = openpyxl.Workbook()
    wb.remove(wb.active)  # remove default empty sheet

    mn = MONTH_NAMES.get(lang, MONTH_NAMES["de"])

    for period_info, sessions in periods_sessions:
        start = period_info["start"]
        year, month = start.year, start.month
        sheet_name = f"{mn[month]} {year}"[:31]
        ws = wb.create_sheet(title=sheet_name)
        ws.freeze_panes = "A3"

        # Title row (16 columns: A–P)
        ws.merge_cells("A1:P1")
        t = ws["A1"]
        t.value     = period_info.get("label_de" if lang != "en" else "label_en", sheet_name)
        t.font      = Font(name="Arial", bold=True, size=13, color=C_FG)
        t.fill      = PatternFill("solid", start_color=C_HDR)
        t.alignment = Alignment(horizontal="center", vertical="center")
        ws.row_dimensions[1].height = 26

        hdrs = [
            t_export("row_num", lang),
            t_export("date", lang),
            t_export("start", lang),
            t_export("end", lang),
            t_export("duration_hours", lang),
            t_export("charger_type", lang),
            t_export("odo_start", lang),
            t_export("odo_end", lang),
            t_export("soc_start", lang),
            t_export("soc_end", lang),
            t_export("kwh", lang),
            t_export("price_per_kwh", lang),
            t_export("cost", lang),
            t_export("location", lang),
            "Preisquelle" if lang != "en" else "Price source",
            "Vertragsname" if lang != "en" else "Contract name",
        ]
        for col, h in enumerate(hdrs, 1):
            cs(ws, 2, col, h, bold=True, bg="#"+C_HDR, fg=C_FG, al="center")

        ds = 3
        for i, s in enumerate(sessions):
            r  = ds + i
            rd = to_row(s, i + 1, lang=lang)
            bg = row_bg(s)
            cs(ws, r, 1,  rd["row_num"],         bg=bg, al="center")
            cs(ws, r, 2,  rd["date"],             bg=bg, nf="DD.MM.YYYY", al="center")
            cs(ws, r, 3,  rd["start_time"],       bg=bg, nf="HH:MM", al="center")
            cs(ws, r, 4,  rd["end_time"],         bg=bg, nf="HH:MM", al="center")
            cs(ws, r, 5,  rd["duration_hours"],   bg=bg, nf='0.00 "h"', al="right")
            cs(ws, r, 6,  rd["charger_type"],     bg=bg, al="center")
            cs(ws, r, 7,  rd["odo_start"],        bg=bg, nf='#,##0 "km"', al="right")
            cs(ws, r, 8,  rd["odo_end"],          bg=bg, nf='#,##0 "km"', al="right")
            cs(ws, r, 9,  rd["soc_start"],        bg=bg, nf='0"%"', al="right")
            cs(ws, r, 10, rd["soc_end"],          bg=bg, nf='0"%"', al="right")
            cs(ws, r, 11, rd["kwh_charged"],      bg=bg, nf='0.00 "kWh"', al="right")
            cs(ws, r, 12, rd["price_per_kwh"],    bg=bg, nf='€0.0000', al="right")
            cs(ws, r, 13, rd["cost_eur"],         bg=bg, nf='€#,##0.00', al="right")
            cs(ws, r, 14, rd["location"],         bg=bg, al="center")
            cs(ws, r, 15, rd["price_source"],     bg=bg, al="left")
            cs(ws, r, 16, rd["charging_contract_name"], bg=bg, al="left")

        n = len(sessions)
        if n:
            tr = ds + n
            for col in range(1, 17):
                v  = ("Σ" if col == 1
                      else f"=SUM(K{ds}:K{ds+n-1})" if col == 11
                      else f"=SUM(M{ds}:M{ds+n-1})" if col == 13
                      else "")
                nf = ('0.00 "kWh"' if col == 11 else '€#,##0.00' if col == 13 else None)
                cs(ws, tr, col, v, bold=True, bg="#"+C_SUM, nf=nf,
                   al="center" if col == 1 else "right")

        # Column widths
        col_widths = [5, 13, 8, 8, 8, 8, 13, 13, 10, 10, 13, 10, 12, 14, 14, 20]
        for i, w in enumerate(col_widths, 1):
            ws.column_dimensions[openpyxl.utils.get_column_letter(i)].width = w

    if not wb.worksheets:
        wb.create_sheet("Leer")

    out = io.BytesIO()
    wb.save(out)
    return out.getvalue(), warnings_list


if __name__ == "__main__":
    y, m = (map(int, sys.argv[1].split("-")) if len(sys.argv) > 1
            else (datetime.now().year, datetime.now().month))
    loc = sys.argv[2] if len(sys.argv) > 2 else "all"
    export(y, m, loc)
