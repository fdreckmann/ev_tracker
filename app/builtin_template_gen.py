"""Generate built-in template .xlsx files on demand."""
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from pathlib import Path

_THIN = Side(style="thin", color="BBBBBB")
_BRD  = Border(left=_THIN, right=_THIN, top=_THIN, bottom=_THIN)

def _hdr(ws, r, c, v, bg="1F4E79", fg="FFFFFF", bold=True, size=10):
    cell = ws.cell(row=r, column=c, value=v)
    cell.font      = Font(name="Arial", bold=bold, color=fg, size=size)
    cell.fill      = PatternFill("solid", start_color=bg)
    cell.alignment = Alignment(horizontal="center", vertical="center")
    cell.border    = _BRD
    return cell

def _lbl(ws, r, c, v, bold=False):
    cell = ws.cell(row=r, column=c, value=v)
    cell.font      = Font(name="Arial", bold=bold, size=10)
    cell.alignment = Alignment(horizontal="left", vertical="center")
    return cell

def _val(ws, r, c, v=""):
    cell = ws.cell(row=r, column=c, value=v)
    cell.font      = Font(name="Arial", size=10)
    cell.alignment = Alignment(horizontal="left", vertical="center")
    cell.border    = _BRD
    return cell

def generate_standard(path: Path):
    wb = openpyxl.Workbook()
    ws = wb.active; ws.title = "Ladeprotokoll"
    # Header info rows
    _lbl(ws, 1, 1, "Fahrer:", bold=True); _val(ws, 1, 2, "{{fahrer}}")
    _lbl(ws, 1, 3, "Kennzeichen:", bold=True); _val(ws, 1, 4, "{{kennzeichen}}")
    _lbl(ws, 2, 1, "Monat:", bold=True); _val(ws, 2, 2, "{{month_year}}")
    _lbl(ws, 2, 3, "Abteilung:", bold=True); _val(ws, 2, 4, "{{abteilung}}")
    # Merge title
    ws.merge_cells("A4:H4")
    t = ws["A4"]; t.value = "Ladeprotokoll"
    t.font = Font(name="Arial", bold=True, size=14, color="1F4E79")
    t.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[4].height = 24
    # Column headers (row 5)
    for c, h in enumerate(["Nr.", "Datum", "Start", "Ende", "Dauer", "kWh", "Kosten €", "Standort"], 1):
        _hdr(ws, 5, c, h)
    ws.row_dimensions[5].height = 18
    # Column widths
    for c, w in enumerate([5, 12, 8, 8, 10, 10, 12, 16], 1):
        ws.column_dimensions[get_column_letter(c)].width = w
    # Data rows placeholder (rows 6-25, alternating fill)
    for r in range(6, 26):
        bg = "F2F7FB" if r % 2 == 0 else "FFFFFF"
        for c in range(1, 9):
            cell = ws.cell(row=r, column=c)
            cell.fill   = PatternFill("solid", start_color=bg)
            cell.border = _BRD
    wb.save(path)

def generate_arbeitgeber(path: Path):
    wb = openpyxl.Workbook()
    ws = wb.active; ws.title = "Ladeprotokoll"
    # Header rows
    _lbl(ws, 1, 1, "Fahrer:", bold=True); _val(ws, 1, 2, "{{fahrer}}")
    _lbl(ws, 1, 3, "Kennzeichen:", bold=True); _val(ws, 1, 4, "{{kennzeichen}}")
    _lbl(ws, 2, 1, "Kostenstelle:", bold=True); _val(ws, 2, 2, "{{kostenstelle}}")
    _lbl(ws, 2, 3, "Monat:", bold=True); _val(ws, 2, 4, "{{month_year}}")
    _lbl(ws, 3, 1, "Abteilung:", bold=True); _val(ws, 3, 2, "{{abteilung}}")
    _lbl(ws, 4, 1, "Gesamt kWh:", bold=True); _val(ws, 4, 2, "{{total_kwh}}")
    _lbl(ws, 4, 3, "Gesamtkosten:", bold=True); _val(ws, 4, 4, "{{total_cost}}")
    ws.merge_cells("A6:I6")
    t = ws["A6"]; t.value = "Arbeitgeber-Ladeabrechnung"
    t.font = Font(name="Arial", bold=True, size=13, color="4B0082")
    t.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[6].height = 22
    for c, h in enumerate(["Nr.", "Datum", "Start", "Ende", "kWh", "Zähler Alt", "Zähler Neu", "Kosten €", "Ladeart"], 1):
        _hdr(ws, 7, c, h, bg="4B0082")
    ws.row_dimensions[7].height = 18
    for c, w in enumerate([5, 12, 8, 8, 10, 13, 13, 12, 10], 1):
        ws.column_dimensions[get_column_letter(c)].width = w
    for r in range(8, 30):
        bg = "F5F0FF" if r % 2 == 0 else "FFFFFF"
        for c in range(1, 10):
            cell = ws.cell(row=r, column=c)
            cell.fill   = PatternFill("solid", start_color=bg)
            cell.border = _BRD
    wb.save(path)

def generate_minimal(path: Path):
    wb = openpyxl.Workbook()
    ws = wb.active; ws.title = "Ladevorgänge"
    ws.merge_cells("A1:D1")
    t = ws["A1"]; t.value = "{{month_year}}"
    t.font = Font(name="Arial", bold=True, size=12, color="166534")
    t.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 20
    for c, h in enumerate(["Datum", "kWh", "Kosten €", "Standort"], 1):
        _hdr(ws, 2, c, h, bg="166534")
    for c, w in enumerate([14, 12, 14, 18], 1):
        ws.column_dimensions[get_column_letter(c)].width = w
    for r in range(3, 25):
        bg = "F0FFF4" if r % 2 == 0 else "FFFFFF"
        for c in range(1, 5):
            cell = ws.cell(row=r, column=c)
            cell.fill   = PatternFill("solid", start_color=bg)
            cell.border = _BRD
    wb.save(path)

def generate_steuer(path: Path):
    wb = openpyxl.Workbook()
    ws = wb.active; ws.title = "Nachweis"
    _lbl(ws, 1, 1, "Fahrer:", bold=True); _val(ws, 1, 2, "{{fahrer}}")
    _lbl(ws, 1, 3, "Kennzeichen:", bold=True); _val(ws, 1, 4, "{{kennzeichen}}")
    _lbl(ws, 2, 1, "Monat:", bold=True); _val(ws, 2, 2, "{{month_year}}")
    _lbl(ws, 3, 1, "Zähler Anfang:", bold=True); _val(ws, 3, 2, "{{meter_start_value}}")
    _lbl(ws, 3, 3, "Zähler Ende:", bold=True); _val(ws, 3, 4, "{{meter_end_value}}")
    _lbl(ws, 4, 1, "Gesamt KM:", bold=True); _val(ws, 4, 2, "{{total_km}}")
    _lbl(ws, 4, 3, "Gesamtkosten:", bold=True); _val(ws, 4, 4, "{{total_cost}}")
    ws.merge_cells("A5:H5")
    t = ws["A5"]; t.value = "Steuerlicher Nachweis Ladekosten"
    t.font = Font(name="Arial", bold=True, size=13, color="9A3412")
    t.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[5].height = 22
    for c, h in enumerate(["Nr.", "Datum", "KM Start", "KM Ende", "kWh", "Zähler Alt", "Zähler Neu", "Kosten €"], 1):
        _hdr(ws, 6, c, h, bg="9A3412")
    ws.row_dimensions[6].height = 18
    for c, w in enumerate([5, 12, 12, 12, 10, 13, 13, 12], 1):
        ws.column_dimensions[get_column_letter(c)].width = w
    for r in range(7, 30):
        bg = "FFF7ED" if r % 2 == 0 else "FFFFFF"
        for c in range(1, 9):
            cell = ws.cell(row=r, column=c)
            cell.fill   = PatternFill("solid", start_color=bg)
            cell.border = _BRD
    wb.save(path)

_GENERATORS = {
    "standard":    generate_standard,
    "arbeitgeber": generate_arbeitgeber,
    "minimal":     generate_minimal,
    "steuer":      generate_steuer,
}

def generate_builtin_template(template_id: str, output_path: Path) -> bool:
    """Generate a built-in template .xlsx to output_path. Returns True on success."""
    gen = _GENERATORS.get(template_id)
    if not gen:
        return False
    try:
        output_path.parent.mkdir(parents=True, exist_ok=True)
        gen(output_path)
        return True
    except Exception as e:
        print(f"Template generation error: {e}")
        return False
