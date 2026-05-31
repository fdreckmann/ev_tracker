"""
Tests for production blockers (v2.0.53 go-live fixes).
Covers: DB migration, export param validation, API v1 validation,
        config atomic save, XLSX upload validation, backup restore,
        export template roundtrip, vehicle_id validation, template hash.
"""
import hashlib
import json
import sqlite3
import zipfile
from io import BytesIO
from pathlib import Path


# ─── helpers ──────────────────────────────────────────────────────────────────

def _make_api_token(app, scopes=None):
    import secrets
    from core.db import _get_db, close_db_if_owned
    from datetime import datetime, timezone

    raw = "evtk_" + secrets.token_urlsafe(32)
    token_hash = hashlib.sha256(raw.encode()).hexdigest()
    prefix = raw[:8]
    if scopes is None:
        scopes = ["sessions:read", "sessions:write", "vehicles:read", "system:read"]
    with app.app_context():
        con = _get_db()
        con.execute("""INSERT INTO api_tokens
            (name, token_hash, token_prefix, scopes, is_active, created_at)
            VALUES (?,?,?,?,?,?)""",
            ("blocker-test", token_hash, prefix,
             json.dumps(scopes), 1, datetime.now(timezone.utc).replace(tzinfo=None).isoformat()))
        con.commit()
        close_db_if_owned(con)
    return raw


# ─── 1. DB Migration ──────────────────────────────────────────────────────────

class TestDbMigration:
    def test_old_minimal_db_gets_all_columns(self, tmp_path):
        """A DB with only id/start_ts/end_ts/kwh_charged must gain all base columns."""
        import sys, os
        sys.path.insert(0, str(Path(__file__).parent.parent / "app"))
        db_path = tmp_path / "sessions.db"
        con = sqlite3.connect(db_path)
        con.execute("""CREATE TABLE sessions (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            start_ts TEXT,
            end_ts TEXT,
            kwh_charged REAL
        )""")
        con.execute("INSERT INTO sessions (start_ts, end_ts, kwh_charged) VALUES (?,?,?)",
                    ("2024-01-01T10:00:00", "2024-01-01T11:00:00", 10.0))
        con.commit()
        con.close()

        import core.db as _db
        old_path = _db.DB_PATH
        _db.DB_PATH = db_path
        try:
            import server as _srv
            old_srv_path = _srv.DB_PATH
            _srv.DB_PATH = db_path
            _srv.init_db()
            # Verify columns now exist
            con2 = sqlite3.connect(db_path)
            cols = {r[1] for r in con2.execute("PRAGMA table_info(sessions)").fetchall()}
            con2.close()
            for required in ("cost_eur", "location", "charger_type", "odo_start", "odo_end",
                             "soc_start", "soc_end", "vehicle_id", "provider",
                             "price_per_kwh", "cost_manual"):
                assert required in cols, f"Missing column after migration: {required}"
            _srv.DB_PATH = old_srv_path
        finally:
            _db.DB_PATH = old_path

    def test_migration_is_idempotent(self, app):
        """Running init_db() twice must not raise."""
        import server as _srv
        with app.app_context():
            _srv.init_db()
            _srv.init_db()  # second call — must be harmless


# ─── 5. API v1 Validation ─────────────────────────────────────────────────────

class TestApiV1Validation:
    def test_negative_kwh_returns_400(self, app, client):
        token = _make_api_token(app)
        rv = client.post("/api/v1/sessions",
                         headers={"Authorization": f"Bearer {token}"},
                         json={"start_ts": "2026-05-10T10:00:00",
                               "end_ts": "2026-05-10T11:00:00",
                               "kwh_charged": -5.0})
        assert rv.status_code == 400

    def test_soc_over_100_returns_400(self, app, client):
        token = _make_api_token(app)
        rv = client.post("/api/v1/sessions",
                         headers={"Authorization": f"Bearer {token}"},
                         json={"start_ts": "2026-05-10T10:00:00",
                               "end_ts": "2026-05-10T11:00:00",
                               "kwh_charged": 10.0,
                               "soc_end": 150})
        assert rv.status_code == 400

    def test_end_before_start_returns_400(self, app, client):
        token = _make_api_token(app)
        rv = client.post("/api/v1/sessions",
                         headers={"Authorization": f"Bearer {token}"},
                         json={"start_ts": "2026-05-10T12:00:00",
                               "end_ts": "2026-05-10T10:00:00",
                               "kwh_charged": 10.0})
        assert rv.status_code == 400

    def test_invalid_date_returns_400(self, app, client):
        token = _make_api_token(app)
        rv = client.post("/api/v1/sessions",
                         headers={"Authorization": f"Bearer {token}"},
                         json={"start_ts": "not-a-date",
                               "kwh_charged": 10.0})
        assert rv.status_code == 400

    def test_valid_session_returns_201(self, app, client):
        token = _make_api_token(app)
        rv = client.post("/api/v1/sessions",
                         headers={"Authorization": f"Bearer {token}"},
                         json={"start_ts": "2026-05-10T10:00:00",
                               "end_ts": "2026-05-10T11:00:00",
                               "kwh_charged": 20.0,
                               "location": "home",
                               "charger_type": "ac"})
        assert rv.status_code == 201
        assert rv.get_json().get("ok") is True

    def test_missing_start_ts_returns_400(self, app, client):
        token = _make_api_token(app)
        rv = client.post("/api/v1/sessions",
                         headers={"Authorization": f"Bearer {token}"},
                         json={"kwh_charged": 10.0})
        assert rv.status_code == 400

    def test_negative_cost_returns_400(self, app, client):
        token = _make_api_token(app)
        rv = client.post("/api/v1/sessions",
                         headers={"Authorization": f"Bearer {token}"},
                         json={"start_ts": "2026-05-10T10:00:00",
                               "end_ts": "2026-05-10T11:00:00",
                               "kwh_charged": 10.0,
                               "cost_eur": -1.0})
        assert rv.status_code == 400

    def test_vehicles_name_fallback(self, app, client):
        """api_v1_vehicles must support 'name' or 'car_name' in extra_vehicles."""
        from core.config import load_config, save_config
        token = _make_api_token(app)
        with app.app_context():
            cfg = load_config()
            cfg["extra_vehicles"] = [{"id": "v1", "name": "Tesla", "provider": "manual"}]
            save_config(cfg)
        rv = client.get("/api/v1/vehicles",
                        headers={"Authorization": f"Bearer {token}"})
        assert rv.status_code == 200
        vehicles = rv.get_json()
        v1 = next((v for v in vehicles if v["id"] == "v1"), None)
        assert v1 is not None
        assert v1["name"] == "Tesla"


# ─── 6. Export Parameter Validation ──────────────────────────────────────────

class TestExportParamValidation:
    def test_invalid_year_returns_400(self, authed_client):
        rv = authed_client.get("/api/export?year=abc&month=5")
        assert rv.status_code == 400

    def test_invalid_month_returns_400(self, authed_client):
        rv = authed_client.get("/api/export?year=2026&month=13")
        assert rv.status_code == 400

    def test_month_zero_returns_400(self, authed_client):
        rv = authed_client.get("/api/export?year=2026&month=0")
        assert rv.status_code == 400

    def test_invalid_col_override_returns_400(self, authed_client):
        rv = authed_client.get("/api/export?year=2026&month=5&col_override=kaputt")
        assert rv.status_code == 400

    def test_valid_params_return_xlsx(self, authed_client):
        rv = authed_client.get("/api/export?year=2026&month=5")
        assert rv.status_code == 200
        assert "spreadsheetml" in rv.content_type or "xlsx" in rv.content_type.lower()


# ─── 10. Config Atomic Save ───────────────────────────────────────────────────

class TestConfigAtomicSave:
    def test_repeated_saves_keep_valid_json(self, app, tmp_path):
        from core.config import load_config, save_config
        with app.app_context():
            for i in range(10):
                cfg = load_config()
                cfg["_test_counter"] = i
                save_config(cfg)
            # Read back raw JSON to verify it's valid
            from core.config import CONFIG_FILE
            raw = CONFIG_FILE.read_text()
            parsed = json.loads(raw)
            assert parsed.get("_test_counter") == 9

    def test_legacy_keys_preserved(self, app):
        from core.config import load_config, save_config
        with app.app_context():
            cfg = load_config()
            cfg["_legacy_custom_key"] = "preserved"
            save_config(cfg)
            cfg2 = load_config()
            assert cfg2.get("_legacy_custom_key") == "preserved"


# ─── 12. Backup Restore Robustness ───────────────────────────────────────────

class TestBackupRestoreRobustness:
    def _do_restore(self, zip_path, data_dir):
        """Call the real 2-arg restore implementation from backup_service."""
        from services.backup_service import _restore_backup_impl
        _restore_backup_impl(zip_path, data_dir)

    def test_invalid_zip_rejected(self, tmp_path):
        bad_zip = tmp_path / "bad.zip"
        bad_zip.write_bytes(b"not a zip file at all")
        data_dir = tmp_path / "data"
        data_dir.mkdir()
        try:
            self._do_restore(bad_zip, data_dir)
            assert False, "Should have raised"
        except (ValueError, zipfile.BadZipFile):
            pass

    def test_zip_slip_rejected(self, tmp_path):
        evil_zip = tmp_path / "evil.zip"
        data_dir = tmp_path / "data"
        data_dir.mkdir()
        with zipfile.ZipFile(evil_zip, "w") as zf:
            zf.writestr("../etc/passwd", "root:x:0:0")
        try:
            self._do_restore(evil_zip, data_dir)
            assert False, "Should have raised"
        except ValueError:
            pass

    def test_normal_restore_works(self, tmp_path):
        good_zip = tmp_path / "good.zip"
        data_dir = tmp_path / "data"
        data_dir.mkdir()
        cfg_content = json.dumps({"provider": "manual"})
        with zipfile.ZipFile(good_zip, "w") as zf:
            zf.writestr("config.json", cfg_content)
        self._do_restore(good_zip, data_dir)
        restored = (data_dir / "config.json").read_text()
        assert json.loads(restored)["provider"] == "manual"


# ─── 3. XLSX Upload Validation ───────────────────────────────────────────────

class TestXlsxUploadValidation:
    def test_invalid_xlsx_rejected(self, authed_client):
        """A file with .xlsx extension but invalid content must return 400."""
        data = {"file": (BytesIO(b"not a valid xlsx file"), "bad.xlsx")}
        rv = authed_client.post("/api/template",
                                content_type="multipart/form-data",
                                data=data)
        assert rv.status_code == 400
        body = rv.get_json()
        assert body.get("ok") is False

    def test_valid_xlsx_accepted(self, authed_client, tmp_path):
        """A real openpyxl-created xlsx must be accepted."""
        try:
            import openpyxl
        except ImportError:
            import pytest; pytest.skip("openpyxl not available")
        wb = openpyxl.Workbook()
        ws = wb.active
        ws["A1"] = "Datum"
        ws["B1"] = "kWh"
        buf = BytesIO()
        wb.save(buf)
        buf.seek(0)
        data = {"file": (buf, "valid.xlsx")}
        rv = authed_client.post("/api/template",
                                content_type="multipart/form-data",
                                data=data)
        assert rv.status_code == 200
        assert rv.get_json().get("ok") is True


# ─── 4. Export Template Roundtrip ────────────────────────────────────────────

class TestExportTemplateRoundtrip:
    def test_create_template_saves_all_fields(self, authed_client):
        payload = {
            "name": "Roundtrip Test",
            "column_mapping": {"1": {"field": "date"}, "2": {"field": "kwh_charged"}},
            "cell_mapping": {"B4": "kennzeichen"},
            "signature_mapping": {"anchor_cell": "A10"},
            "start_row": 6,
            "header_row": 5,
            "sheet": "Tabelle1",
        }
        rv = authed_client.post("/api/export/templates",
                                json=payload)
        assert rv.status_code == 200
        tpl = rv.get_json()["template"]
        assert tpl["column_mapping"] == payload["column_mapping"]
        assert tpl["cell_mapping"] == payload["cell_mapping"]
        assert tpl["signature_mapping"] == payload["signature_mapping"]
        assert tpl["start_row"] == 6
        assert tpl["header_row"] == 5
        assert tpl["sheet"] == "Tabelle1"

    def test_legacy_mapping_field_preserved(self, authed_client):
        """Old 'mapping' field must still work and be mirrored to column_mapping."""
        rv = authed_client.post("/api/export/templates",
                                json={"name": "Legacy", "mapping": {"3": "cost_eur"}})
        assert rv.status_code == 200
        tpl = rv.get_json()["template"]
        assert tpl["mapping"] == {"3": "cost_eur"}
        assert tpl["column_mapping"] == {"3": "cost_eur"}

    def test_get_templates_includes_all_fields(self, authed_client):
        """GET /api/export/templates must return column_mapping and cell_mapping."""
        authed_client.post("/api/export/templates",
                           json={"name": "Full", "column_mapping": {"1": "date"},
                                 "cell_mapping": {"C5": "fahrer"}})
        rv = authed_client.get("/api/export/templates")
        assert rv.status_code == 200
        templates = rv.get_json()
        full = next((t for t in templates if t["name"] == "Full"), None)
        assert full is not None
        assert "column_mapping" in full
        assert "cell_mapping" in full
        assert full["cell_mapping"].get("C5") == "fahrer"


# ─── 13. Vehicle ID Validation ────────────────────────────────────────────────

class TestVehicleIdValidation:
    def test_v0_always_valid(self, app):
        from routes.api_v1 import _vehicle_id_valid
        with app.app_context():
            assert _vehicle_id_valid("v0") is True

    def test_unknown_vehicle_invalid(self, app):
        from routes.api_v1 import _vehicle_id_valid
        with app.app_context():
            assert _vehicle_id_valid("v99") is False

    def test_path_traversal_rejected(self, app):
        from routes.api_v1 import _vehicle_id_valid
        with app.app_context():
            assert _vehicle_id_valid("../../etc/passwd") is False
            assert _vehicle_id_valid("v0/../../secret") is False

    def test_known_extra_vehicle_valid(self, app):
        from routes.api_v1 import _vehicle_id_valid
        from core.config import load_config, save_config
        with app.app_context():
            cfg = load_config()
            cfg["extra_vehicles"] = [{"id": "v2", "name": "Extra", "provider": "manual"}]
            save_config(cfg)
            assert _vehicle_id_valid("v2") is True

    def test_api_v1_rejects_unknown_vehicle(self, app, client):
        """POST /api/v1/sessions with unknown vehicle_id must return 400."""
        token = _make_api_token(app)
        rv = client.post("/api/v1/sessions",
                         headers={"Authorization": f"Bearer {token}"},
                         json={"start_ts": "2026-05-10T10:00:00",
                               "end_ts": "2026-05-10T11:00:00",
                               "kwh_charged": 10.0,
                               "vehicle_id": "nonexistent_vehicle"})
        assert rv.status_code == 400

    def test_manual_session_rejects_unknown_vehicle(self, authed_client):
        """POST /api/sessions/manual with unknown vehicle_id must return 400."""
        rv = authed_client.post("/api/sessions/manual",
                                json={"start_ts": "2026-05-10T10:00:00",
                                      "end_ts": "2026-05-10T11:00:00",
                                      "kwh_charged": 10.0,
                                      "vehicle_id": "nonexistent_vehicle"})
        assert rv.status_code == 400


# ─── 14. Template Hash Invalidation ──────────────────────────────────────────

class TestTemplateHash:
    def test_upload_sets_hash(self, authed_client):
        """Uploading a template must set active_template.hash in config."""
        try:
            import openpyxl
        except ImportError:
            import pytest; pytest.skip("openpyxl not available")
        wb = openpyxl.Workbook()
        ws = wb.active
        ws["A1"] = "Test"
        buf = BytesIO()
        wb.save(buf)
        buf.seek(0)
        rv = authed_client.post("/api/template",
                                content_type="multipart/form-data",
                                data={"file": (buf, "test.xlsx")})
        assert rv.status_code == 200
        from core.config import load_config
        with authed_client.application.app_context():
            cfg = load_config()
        tpl = cfg.get("active_template") or {}
        assert tpl.get("hash") is not None
        assert len(tpl["hash"]) == 16

    def test_upload_invalidates_mapping_hash(self, authed_client):
        """Uploading a template must set template_mapping_hash to None."""
        try:
            import openpyxl
        except ImportError:
            import pytest; pytest.skip("openpyxl not available")
        wb = openpyxl.Workbook()
        buf = BytesIO()
        wb.save(buf)
        buf.seek(0)
        authed_client.post("/api/template",
                           content_type="multipart/form-data",
                           data={"file": (buf, "test2.xlsx")})
        from core.config import load_config
        with authed_client.application.app_context():
            cfg = load_config()
        assert cfg.get("template_mapping_hash") is None

    def test_mapping_post_stores_hash(self, authed_client):
        """POST /api/template/mapping must save current template hash."""
        try:
            import openpyxl
        except ImportError:
            import pytest; pytest.skip("openpyxl not available")
        wb = openpyxl.Workbook()
        buf = BytesIO()
        wb.save(buf)
        buf.seek(0)
        authed_client.post("/api/template",
                           content_type="multipart/form-data",
                           data={"file": (buf, "hash_test.xlsx")})
        rv = authed_client.post("/api/template/mapping",
                                json={"column_mapping": {"1": "date"},
                                      "cell_mapping": {},
                                      "start_row": 2,
                                      "header_row": 1,
                                      "sheet": ""})
        assert rv.status_code == 200
        from core.config import load_config
        with authed_client.application.app_context():
            cfg = load_config()
        tpl_hash = (cfg.get("active_template") or {}).get("hash")
        map_hash = cfg.get("template_mapping_hash")
        assert tpl_hash is not None
        assert map_hash == tpl_hash

    def test_mapping_get_returns_hashes(self, authed_client):
        """GET /api/template/mapping must return mapping_hash and template_hash."""
        rv = authed_client.get("/api/template/mapping")
        assert rv.status_code == 200
        body = rv.get_json()
        assert "mapping_hash" in body
        assert "template_hash" in body


# ─── 15. No writes to /data ───────────────────────────────────────────────────

class TestNoDataDirWrites:
    def test_config_write_goes_to_tmp_path(self, app, tmp_path):
        """save_config must write to the patched path, NOT /data/config.json."""
        from core.config import load_config, save_config, CONFIG_FILE
        real_data_config = Path("/data/config.json")
        with app.app_context():
            cfg = load_config()
            cfg["_isolation_test"] = "yes"
            save_config(cfg)
            # Patched path should have our key
            assert CONFIG_FILE.read_text().find("_isolation_test") != -1
            # /data/config.json should NOT be touched
            if real_data_config.exists():
                content = real_data_config.read_text()
                assert "_isolation_test" not in content
