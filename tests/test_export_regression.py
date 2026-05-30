"""
Export regression tests — covers:
- to_row() edge cases (None SOC, None cost, charger_type derivation, location labels)
- compute_header_values() total_kwh / total_cost include ALL locations (bug fix)
- export_builtin: extern AC/DC, unknown location, missing SOC, special chars
- fetch_sessions location filter
- Cell mapping applied correctly in template export
- Multi-vehicle export: sessions from v1 appear in output
- Empty month: no crash
- API endpoint smoke tests for all major scenarios
"""
import sys
import pytest
from io import BytesIO
from pathlib import Path

sys.path.insert(0, str(Path(__file__).parent.parent / "app"))


def _has_openpyxl():
    try:
        import openpyxl
        return True
    except ImportError:
        return False


# ── Shared session factories ──────────────────────────────────────────────────

def _make_session(**kwargs):
    base = {
        "id": 1,
        "start_ts": "2026-05-01T10:00:00",
        "end_ts":   "2026-05-01T11:30:00",
        "kwh_charged": 20.0,
        "cost_eur": 6.00,
        "price_per_kwh": 0.30,
        "location": "home",
        "charger_type": None,
        "soc_start": 20.0,
        "soc_end": 80.0,
        "odo_start": 10000,
        "odo_end": 10150,
        "meter_old": None,
        "meter_new": None,
        "max_power_kw": None,
        "price_source": None,
        "charging_contract_name": None,
        "charging_contract_id": None,
        "charger_power_kw": None,
    }
    base.update(kwargs)
    return base


def _seed_diverse(app, year=2026, month=5):
    """Seed 6 sessions covering all required scenarios."""
    from core.db import _get_db, close_db_if_owned
    with app.app_context():
        con = _get_db()
        sessions = [
            # 1. Home AC — standard
            ("2026-05-01T10:00:00", "2026-05-01T11:00:00",
             15.0, 4.50, "home", "ac", "v0", 0.30, 20.0, 80.0, None, None),
            # 2. Extern AC
            ("2026-05-02T12:00:00", "2026-05-02T13:30:00",
             22.0, 7.70, "extern", "ac", "v0", 0.35, 30.0, 90.0, None, None),
            # 3. Extern DC
            ("2026-05-03T08:00:00", "2026-05-03T08:30:00",
             30.0, 12.00, "extern", "dc", "v0", 0.40, 10.0, 70.0, None, None),
            # 4. Session with missing SOC (None)
            ("2026-05-04T14:00:00", "2026-05-04T15:00:00",
             10.0, 3.00, "home", "ac", "v0", 0.30, None, None, None, None),
            # 5. Unknown location (manually changed)
            ("2026-05-05T09:00:00", "2026-05-05T10:00:00",
             8.0, 2.80, "unknown", None, "v0", 0.35, 50.0, 60.0, None, None),
            # 6. Second vehicle (v1) — extern DC
            ("2026-05-06T16:00:00", "2026-05-06T16:45:00",
             25.0, 9.50, "extern", "dc", "v1", 0.38, 40.0, 75.0, None, None),
        ]
        for (start, end, kwh, cost, loc, ctype, vid, price, soc_s, soc_e, mo, mn) in sessions:
            con.execute("""
                INSERT INTO sessions
                    (start_ts, end_ts, kwh_charged, cost_eur, location, charger_type,
                     vehicle_id, price_per_kwh, soc_start, soc_end, meter_old, meter_new,
                     cost_manual, provider)
                VALUES (?,?,?,?,?,?,?,?,?,?,?,?,1,'manual')
            """, (start, end, kwh, cost, loc, ctype, vid, price, soc_s, soc_e, mo, mn))
        con.commit()
        close_db_if_owned(con)


# ── to_row() edge cases ───────────────────────────────────────────────────────

class TestToRow:
    def test_none_soc_does_not_crash(self):
        from export_excel import to_row
        s = _make_session(soc_start=None, soc_end=None)
        rd = to_row(s, 1)
        assert rd["soc_start"] is None
        assert rd["soc_end"] is None

    def test_none_cost_does_not_crash(self):
        from export_excel import to_row
        s = _make_session(cost_eur=None, price_per_kwh=None)
        rd = to_row(s, 1)
        assert rd["cost_eur"] is None
        assert rd["price_per_kwh"] is None

    def test_charger_type_derived_home(self):
        from export_excel import to_row
        s = _make_session(location="home", charger_type=None)
        rd = to_row(s, 1)
        assert rd["charger_type"] == "AC"

    def test_charger_type_derived_extern(self):
        from export_excel import to_row
        s = _make_session(location="extern", charger_type=None)
        rd = to_row(s, 1)
        assert rd["charger_type"] == "DC"

    def test_explicit_charger_type_preserved(self):
        """Explicitly stored charger_type (not 'unknown') must not be overridden."""
        from export_excel import to_row
        s = _make_session(location="extern", charger_type="ac")
        rd = to_row(s, 1)
        assert rd["charger_type"] == "ac"

    def test_unknown_charger_type_derived(self):
        """charger_type='unknown' should be derived from location."""
        from export_excel import to_row
        s = _make_session(location="home", charger_type="unknown")
        rd = to_row(s, 1)
        assert rd["charger_type"] == "AC"

    def test_location_label_home(self):
        from export_excel import to_row
        rd = to_row(_make_session(location="home"), 1)
        assert "Zuhause" in rd["location"]

    def test_location_label_extern(self):
        from export_excel import to_row
        rd = to_row(_make_session(location="extern"), 1)
        assert "Extern" in rd["location"]

    def test_location_label_unknown(self):
        from export_excel import to_row
        rd = to_row(_make_session(location="unknown"), 1)
        assert rd["location"] == "—"

    def test_row_num_set(self):
        from export_excel import to_row
        rd = to_row(_make_session(), 7)
        assert rd["row_num"] == 7

    def test_duration_hours_calculated(self):
        """1.5h session → duration_hours == 1.5."""
        from export_excel import to_row
        s = _make_session(
            start_ts="2026-05-01T10:00:00",
            end_ts="2026-05-01T11:30:00",
            kwh_charged=15.0,
        )
        rd = to_row(s, 1)
        assert rd["duration_hours"] == pytest.approx(1.5, abs=0.01)

    def test_charge_power_kw_calculated(self):
        """15 kWh / 1.5 h = 10 kW."""
        from export_excel import to_row
        s = _make_session(
            start_ts="2026-05-01T10:00:00",
            end_ts="2026-05-01T11:30:00",
            kwh_charged=15.0,
        )
        rd = to_row(s, 1)
        assert rd["charge_power_kw"] == pytest.approx(10.0, abs=0.1)

    def test_special_chars_in_contract_name(self):
        """Sonderzeichen in Vertragsname dürfen nicht zu Fehler führen."""
        from export_excel import to_row
        s = _make_session(charging_contract_name="Ü&Ö <Test>")
        rd = to_row(s, 1)
        assert rd["charging_contract_name"] == "Ü&Ö <Test>"


# ── compute_header_values() ───────────────────────────────────────────────────

class TestComputeHeaderValues:
    """Tests for the header value computation used in placeholders and cell_mapping."""

    def _hv(self, sessions, **kw):
        from export_excel import compute_header_values
        return compute_header_values(sessions, year=2026, month=5, **kw)

    def test_empty_sessions_no_crash(self):
        hv = self._hv([])
        assert hv["total_sessions"] == 0
        assert hv["total_kwh"] == 0
        assert hv["total_cost"] == 0
        assert hv["meter_start_value"] is None
        assert hv["meter_end_value"] is None

    def test_total_kwh_includes_unknown_location(self):
        """BUG FIX: sessions with location='unknown' must count toward total_kwh."""
        home = _make_session(id=1, kwh_charged=10.0, cost_eur=3.0, location="home")
        unk  = _make_session(id=2, kwh_charged=5.0,  cost_eur=1.5, location="unknown")
        hv = self._hv([home, unk])
        assert hv["total_kwh"] == pytest.approx(15.0)

    def test_total_cost_includes_unknown_location(self):
        """BUG FIX: sessions with location='unknown' must count toward total_cost."""
        home = _make_session(id=1, kwh_charged=10.0, cost_eur=3.00, location="home")
        unk  = _make_session(id=2, kwh_charged=5.0,  cost_eur=1.50, location="unknown")
        hv = self._hv([home, unk])
        assert hv["total_cost"] == pytest.approx(4.50)

    def test_home_extern_breakdown_correct(self):
        home = _make_session(id=1, kwh_charged=10.0, cost_eur=3.0, location="home")
        ext  = _make_session(id=2, kwh_charged=20.0, cost_eur=8.0, location="extern")
        hv = self._hv([home, ext])
        assert hv["total_home_kwh"]     == pytest.approx(10.0)
        assert hv["total_external_kwh"] == pytest.approx(20.0)
        assert hv["total_home_cost"]    == pytest.approx(3.0)
        assert hv["total_external_cost"]== pytest.approx(8.0)
        assert hv["total_kwh"]          == pytest.approx(30.0)
        assert hv["total_cost"]         == pytest.approx(11.0)

    def test_all_three_locations_totals(self):
        """All locations contribute to totals; only home/extern populate breakdown."""
        home = _make_session(id=1, kwh_charged=10.0, cost_eur=3.0, location="home")
        ext  = _make_session(id=2, kwh_charged=15.0, cost_eur=5.0, location="extern")
        unk  = _make_session(id=3, kwh_charged=5.0,  cost_eur=2.0, location="unknown")
        hv = self._hv([home, ext, unk])
        assert hv["total_sessions"]     == 3
        assert hv["total_kwh"]          == pytest.approx(30.0)
        assert hv["total_cost"]         == pytest.approx(10.0)
        assert hv["total_home_kwh"]     == pytest.approx(10.0)
        assert hv["total_external_kwh"] == pytest.approx(15.0)

    def test_none_kwh_treated_as_zero(self):
        s = _make_session(id=1, kwh_charged=None, cost_eur=None)
        hv = self._hv([s])
        assert hv["total_kwh"] == 0
        assert hv["total_cost"] == 0

    def test_total_sessions_count(self):
        sessions = [_make_session(id=i) for i in range(5)]
        hv = self._hv(sessions)
        assert hv["total_sessions"] == 5

    def test_month_year_german(self):
        hv = self._hv([])
        assert "Mai" in hv["month_year"]
        assert "2026" in hv["month_year"]

    def test_month_year_english(self):
        from export_excel import compute_header_values
        hv = compute_header_values([], year=2026, month=5, lang="en")
        assert "May" in hv["month_year"]


# ── export_builtin() ──────────────────────────────────────────────────────────

@pytest.mark.skipif(not _has_openpyxl(), reason="openpyxl not available")
class TestExportBuiltin:
    """Tests for the built-in (no template) Excel export."""

    def _export(self, sessions, year=2026, month=5, location="all", lang="de"):
        import export_excel as _xl
        import tempfile, os
        old_db  = _xl.DB_PATH
        old_dir = _xl.EXPORT_DIR
        old_tpl = _xl.TEMPLATE_PATH
        with tempfile.TemporaryDirectory() as td:
            _xl.EXPORT_DIR  = Path(td) / "exports"
            _xl.EXPORT_DIR.mkdir()
            _xl.TEMPLATE_PATH = Path(td) / "template.xlsx"  # make sure it doesn't exist
            # Patch fetch_sessions to return our test sessions
            orig_fetch = _xl.fetch_sessions
            _xl.fetch_sessions = lambda y, m, loc="all": [
                s for s in sessions
                if loc == "all" or s.get("location") == loc
            ]
            try:
                xlsx_bytes = _xl.export(year, month, location, lang=lang)
            finally:
                _xl.fetch_sessions = orig_fetch
                _xl.DB_PATH = old_db
                _xl.EXPORT_DIR = old_dir
                _xl.TEMPLATE_PATH = old_tpl
        return xlsx_bytes

    def _load(self, xlsx_bytes):
        import openpyxl
        return openpyxl.load_workbook(BytesIO(xlsx_bytes), data_only=True)

    def test_home_session_exported(self):
        s = _make_session(location="home", charger_type="ac", kwh_charged=15.0)
        wb = self._load(self._export([s]))
        ws = wb.active
        # Row 3 = first data row, col 11 = kWh
        assert ws.cell(row=3, column=11).value == pytest.approx(15.0)

    def test_extern_ac_session_exported(self):
        s = _make_session(location="extern", charger_type="ac", kwh_charged=22.0, cost_eur=7.7)
        wb = self._load(self._export([s]))
        ws = wb.active
        assert ws.cell(row=3, column=11).value == pytest.approx(22.0)

    def test_extern_dc_session_exported(self):
        s = _make_session(location="extern", charger_type="dc", kwh_charged=30.0)
        wb = self._load(self._export([s]))
        ws = wb.active
        assert ws.cell(row=3, column=11).value == pytest.approx(30.0)

    def test_missing_soc_does_not_crash(self):
        s = _make_session(soc_start=None, soc_end=None)
        xlsx_bytes = self._export([s])
        wb = self._load(xlsx_bytes)
        assert wb is not None

    def test_none_cost_does_not_crash(self):
        s = _make_session(cost_eur=None)
        xlsx_bytes = self._export([s])
        assert len(xlsx_bytes) > 0

    def test_empty_month_no_crash(self):
        xlsx_bytes = self._export([])
        wb = self._load(xlsx_bytes)
        assert wb is not None

    def test_two_sheets_produced(self):
        """Built-in export always produces data sheet + summary sheet."""
        s = _make_session()
        wb = self._load(self._export([s]))
        assert len(wb.sheetnames) == 2

    def test_summary_total_kwh_includes_unknown_location(self):
        """BUG FIX: summary sheet 'Gesamt kWh' must include unknown-location session."""
        home = _make_session(id=1, kwh_charged=10.0, cost_eur=3.0, location="home")
        unk  = _make_session(id=2, kwh_charged=5.0,  cost_eur=1.5, location="unknown")
        wb = self._load(self._export([home, unk]))
        ws2 = wb.worksheets[1]  # summary sheet
        # Find the 'Gesamt kWh' row value
        total_kwh_val = None
        for row in ws2.iter_rows(values_only=True):
            if row[0] and "Gesamt kWh" in str(row[0]):
                total_kwh_val = row[1]
                break
        assert total_kwh_val is not None, "Gesamt kWh row not found in summary"
        assert total_kwh_val == pytest.approx(15.0), \
            f"Expected 15.0 (home 10 + unknown 5), got {total_kwh_val}"

    def test_summary_total_cost_includes_unknown_location(self):
        """BUG FIX: summary 'Gesamtkosten' must include unknown-location sessions."""
        home = _make_session(id=1, kwh_charged=10.0, cost_eur=3.00, location="home")
        unk  = _make_session(id=2, kwh_charged=5.0,  cost_eur=1.50, location="unknown")
        wb = self._load(self._export([home, unk]))
        ws2 = wb.worksheets[1]
        total_cost_val = None
        for row in ws2.iter_rows(values_only=True):
            if row[0] and "Gesamtkosten" in str(row[0]):
                total_cost_val = row[1]
                break
        assert total_cost_val == pytest.approx(4.50), \
            f"Expected 4.50, got {total_cost_val}"

    def test_location_filter_home_only(self):
        """Export with location=home must exclude extern sessions."""
        home = _make_session(id=1, kwh_charged=10.0, location="home")
        ext  = _make_session(id=2, kwh_charged=20.0, location="extern")
        wb = self._load(self._export([home, ext], location="home"))
        ws = wb.active
        # Only 1 data row (row 3), no row 4
        val_r3 = ws.cell(row=3, column=11).value
        val_r4 = ws.cell(row=4, column=11).value
        assert val_r3 == pytest.approx(10.0)
        assert val_r4 is None or val_r4 == "" or str(val_r4).startswith("=")

    def test_special_chars_in_session_data(self):
        """Sonderzeichen in Vertragsnamen / location dürfen Export nicht brechen."""
        s = _make_session(charging_contract_name="Grüner Strom & Co. <GmbH>")
        xlsx_bytes = self._export([s])
        assert len(xlsx_bytes) > 0

    def test_multi_vehicle_sessions_all_exported(self):
        """Sessions from multiple vehicles are all included in the export."""
        v0 = _make_session(id=1, kwh_charged=10.0, location="home")
        v1 = _make_session(id=2, kwh_charged=25.0, location="extern")
        v1["vehicle_id"] = "v1"
        wb = self._load(self._export([v0, v1]))
        ws = wb.active
        # Both sessions should appear in rows 3 and 4
        kwh_r3 = ws.cell(row=3, column=11).value
        kwh_r4 = ws.cell(row=4, column=11).value
        kwh_vals = sorted(filter(None, [kwh_r3, kwh_r4]))
        assert kwh_vals == pytest.approx(sorted([10.0, 25.0]))

    def test_english_export_labels(self):
        """lang='en' must produce English header labels."""
        s = _make_session()
        wb = self._load(self._export([s], lang="en"))
        ws = wb.active
        # Check title cell (A1 merged)
        title = ws["A1"].value or ""
        assert "Charging log" in title, f"Expected English title, got: {title!r}"

    def test_sum_formula_in_footer(self):
        """Built-in export must have SUM formula in the summary row (col 11 = kWh)."""
        import openpyxl
        sessions = [_make_session(id=i) for i in range(3)]
        xlsx_bytes = self._export(sessions)
        # Load without data_only so we see actual formula strings
        wb = openpyxl.load_workbook(BytesIO(xlsx_bytes), data_only=False)
        ws = wb.active
        # Summary row: ds=3, n=3 → row 6
        kwh_cell = ws.cell(row=6, column=11).value
        assert isinstance(kwh_cell, str) and kwh_cell.startswith("=SUM"), \
            f"Expected SUM formula in kWh footer, got {kwh_cell!r}"


# ── Template-based export: cell mapping ───────────────────────────────────────

@pytest.mark.skipif(not _has_openpyxl(), reason="openpyxl not available")
class TestCellMappingExportIntegration:
    """Verify that cell_mapping values are written to the correct cells in the output."""

    def test_cell_mapping_written_to_output(self, tmp_path):
        import openpyxl
        from export_excel import export_with_template

        # Simple template: header info in B2, data starts row 4
        wb = openpyxl.Workbook()
        ws = wb.active
        ws["A2"] = "Kennzeichen:"
        ws["B2"] = ""          # will be filled by cell_mapping
        ws["A3"] = "Datum"; ws["B3"] = "kWh"
        ws["A4"] = None; ws["B4"] = None
        ws["A5"] = "Gesamt:"
        tpl_path = tmp_path / "template.xlsx"
        wb.save(str(tpl_path))

        import export_excel as _xl
        old_tpl = _xl.TEMPLATE_PATH
        old_exp = _xl.EXPORT_DIR
        try:
            _xl.TEMPLATE_PATH = tpl_path
            _xl.EXPORT_DIR    = tmp_path / "exports"
            _xl.EXPORT_DIR.mkdir(exist_ok=True)

            sessions = [_make_session()]
            result_path = export_with_template(
                2026, 5, sessions,
                location="all",
                col_override={"1": "date", "2": "kwh_charged"},
                start_row=4, header_row=3,
                footer_start_row=5,
                header_info={"kennzeichen": "EV-123", "fahrer": "Testfahrer"},
                cell_mapping={"B2": "kennzeichen"},
            )
            out_wb = openpyxl.load_workbook(result_path, data_only=True)
            out_ws = out_wb.active
            assert out_ws["B2"].value == "EV-123", \
                f"cell_mapping B2→kennzeichen not applied, got {out_ws['B2'].value!r}"
        finally:
            _xl.TEMPLATE_PATH = old_tpl
            _xl.EXPORT_DIR    = old_exp

    def test_cell_mapping_object_format(self, tmp_path):
        """cell_mapping with {'field': 'month_year'} object format must also work."""
        import openpyxl
        from export_excel import export_with_template

        wb = openpyxl.Workbook()
        ws = wb.active
        ws["A1"] = "Monat:"; ws["B1"] = ""
        ws["A2"] = "kWh"
        ws["A3"] = None
        tpl_path = tmp_path / "template.xlsx"
        wb.save(str(tpl_path))

        import export_excel as _xl
        old_tpl = _xl.TEMPLATE_PATH
        old_exp = _xl.EXPORT_DIR
        try:
            _xl.TEMPLATE_PATH = tpl_path
            _xl.EXPORT_DIR    = tmp_path / "exports"
            _xl.EXPORT_DIR.mkdir(exist_ok=True)

            sessions = [_make_session()]
            result_path = export_with_template(
                2026, 5, sessions,
                location="all",
                col_override={"1": "kwh_charged"},
                start_row=3, header_row=2,
                footer_start_row=4,
                cell_mapping={"B1": {"field": "month_year"}},
            )
            out_wb = openpyxl.load_workbook(result_path, data_only=True)
            out_ws = out_wb.active
            assert "Mai" in str(out_ws["B1"].value), \
                f"cell_mapping B1→month_year not applied, got {out_ws['B1'].value!r}"
        finally:
            _xl.TEMPLATE_PATH = old_tpl
            _xl.EXPORT_DIR    = old_exp

    def test_formula_cell_not_overwritten_by_cell_mapping(self, tmp_path):
        """A formula cell must not be overwritten by cell_mapping."""
        import openpyxl
        from export_excel import export_with_template

        wb = openpyxl.Workbook()
        ws = wb.active
        ws["B1"] = "=1+1"   # formula — must be preserved
        ws["A2"] = "kWh"
        ws["A3"] = None
        tpl_path = tmp_path / "template.xlsx"
        wb.save(str(tpl_path))

        import export_excel as _xl
        old_tpl = _xl.TEMPLATE_PATH
        old_exp = _xl.EXPORT_DIR
        try:
            _xl.TEMPLATE_PATH = tpl_path
            _xl.EXPORT_DIR    = tmp_path / "exports"
            _xl.EXPORT_DIR.mkdir(exist_ok=True)

            sessions = [_make_session()]
            result_path = export_with_template(
                2026, 5, sessions,
                location="all",
                col_override={"1": "kwh_charged"},
                start_row=3, header_row=2,
                footer_start_row=4,
                cell_mapping={"B1": "total_kwh"},
            )
            out_wb = openpyxl.load_workbook(result_path, data_only=False)
            out_ws = out_wb.active
            b1_val = out_ws["B1"].value
            assert isinstance(b1_val, str) and b1_val.startswith("="), \
                f"Formula cell B1 was overwritten: got {b1_val!r}"
        finally:
            _xl.TEMPLATE_PATH = old_tpl
            _xl.EXPORT_DIR    = old_exp


# ── Placeholder fill ──────────────────────────────────────────────────────────

@pytest.mark.skipif(not _has_openpyxl(), reason="openpyxl not available")
class TestPlaceholderFill:
    def test_placeholder_replaced(self, tmp_path):
        import openpyxl
        from export_excel import export_with_template

        wb = openpyxl.Workbook()
        ws = wb.active
        ws["A1"] = "{{month_year}}"
        ws["B1"] = "{{total_kwh}}"
        ws["A2"] = "kWh"
        ws["A3"] = None
        tpl_path = tmp_path / "template.xlsx"
        wb.save(str(tpl_path))

        import export_excel as _xl
        old_tpl = _xl.TEMPLATE_PATH
        old_exp = _xl.EXPORT_DIR
        try:
            _xl.TEMPLATE_PATH = tpl_path
            _xl.EXPORT_DIR    = tmp_path / "exports"
            _xl.EXPORT_DIR.mkdir(exist_ok=True)

            sessions = [_make_session(kwh_charged=15.0)]
            result_path = export_with_template(
                2026, 5, sessions,
                location="all",
                col_override={"1": "kwh_charged"},
                start_row=3, header_row=2,
                footer_start_row=4,
            )
            out_wb = openpyxl.load_workbook(result_path, data_only=True)
            out_ws = out_wb.active
            assert "Mai" in str(out_ws["A1"].value), \
                f"{{{{month_year}}}} not replaced, got {out_ws['A1'].value!r}"
            assert "15" in str(out_ws["B1"].value), \
                f"{{{{total_kwh}}}} not replaced, got {out_ws['B1'].value!r}"
        finally:
            _xl.TEMPLATE_PATH = old_tpl
            _xl.EXPORT_DIR    = old_exp


# ── API endpoint smoke tests ──────────────────────────────────────────────────

class TestExportAPIEndpoints:
    """Integration tests via the Flask test client."""

    def test_export_home_only(self, authed_client, app):
        _seed_diverse(app)
        rv = authed_client.get("/api/export?year=2026&month=5&location=home")
        assert rv.status_code == 200
        assert "spreadsheetml" in rv.content_type or "xlsx" in rv.content_type.lower()

    def test_export_extern_only(self, authed_client, app):
        _seed_diverse(app)
        rv = authed_client.get("/api/export?year=2026&month=5&location=extern")
        assert rv.status_code == 200

    def test_export_all_locations(self, authed_client, app):
        _seed_diverse(app)
        rv = authed_client.get("/api/export?year=2026&month=5&location=all")
        assert rv.status_code == 200

    def test_export_english_lang(self, authed_client, app):
        _seed_diverse(app)
        rv = authed_client.get("/api/export?year=2026&month=5&lang=en")
        assert rv.status_code == 200

    def test_export_empty_month(self, authed_client, app):
        rv = authed_client.get("/api/export?year=2024&month=1")
        assert rv.status_code == 200

    def test_export_invalid_month(self, authed_client, app):
        rv = authed_client.get("/api/export?year=2026&month=13")
        assert rv.status_code == 400

    def test_export_invalid_year(self, authed_client, app):
        rv = authed_client.get("/api/export?year=1999&month=5")
        assert rv.status_code == 400

    def test_export_preview_diverse_data(self, authed_client, app):
        _seed_diverse(app)
        rv = authed_client.post("/api/export/preview", json={
            "year": 2026, "month": 5, "location": "all",
        })
        assert rv.status_code == 200
        body = rv.get_json()
        assert body["ok"] is True
        assert len(body["sheets"]) >= 1

    def test_export_valid_xlsx_with_diverse_sessions(self, authed_client, app):
        """Export with diverse data (home, extern, unknown, missing SOC) must produce valid XLSX."""
        _seed_diverse(app)
        rv = authed_client.get("/api/export?year=2026&month=5")
        assert rv.status_code == 200
        try:
            import openpyxl
            wb = openpyxl.load_workbook(BytesIO(rv.data), data_only=True)
            ws = wb.active
            # 6 sessions → data rows 3-8 (in builtin export starting at row 3)
            # At least 1 data row must have kWh value
            kwh_values = [
                ws.cell(row=r, column=11).value
                for r in range(3, 10)
                if ws.cell(row=r, column=11).value is not None
            ]
            assert len(kwh_values) >= 1
        except ImportError:
            pass  # openpyxl not available

    def test_preview_download_token_roundtrip(self, authed_client, app):
        """POST /api/export/preview then GET /api/export/download/<token>."""
        _seed_diverse(app)
        rv_prev = authed_client.post("/api/export/preview", json={
            "year": 2026, "month": 5,
        })
        assert rv_prev.status_code == 200
        token = rv_prev.get_json().get("download_token")
        if token:
            rv_dl = authed_client.get(f"/api/export/download/{token}")
            assert rv_dl.status_code == 200
            assert "spreadsheetml" in rv_dl.content_type or "xlsx" in rv_dl.content_type.lower()


# ── fetch_sessions ────────────────────────────────────────────────────────────

class TestFetchSessions:
    """Unit tests for fetch_sessions() location filtering."""

    def _populate(self, tmp_db):
        import sqlite3
        con = sqlite3.connect(tmp_db)
        con.execute("""CREATE TABLE sessions (
            id INTEGER PRIMARY KEY,
            start_ts TEXT, end_ts TEXT, kwh_charged REAL, cost_eur REAL,
            location TEXT, charger_type TEXT, vehicle_id TEXT,
            soc_start REAL, soc_end REAL, odo_start REAL, odo_end REAL,
            meter_old REAL, meter_new REAL, max_power_kw REAL,
            price_per_kwh REAL, price_source TEXT, charging_contract_name TEXT,
            charging_contract_id TEXT, charger_power_kw REAL,
            cost_manual INTEGER DEFAULT 0, provider TEXT DEFAULT 'manual',
            created_mode TEXT
        )""")
        rows = [
            (1, "2026-05-01T10:00:00", "2026-05-01T11:00:00", 10.0, 3.0, "home", "ac", "v0"),
            (2, "2026-05-02T10:00:00", "2026-05-02T11:00:00", 20.0, 7.0, "extern", "dc", "v0"),
            (3, "2026-05-03T10:00:00", "2026-05-03T11:00:00", 5.0, 1.5, "unknown", None, "v0"),
        ]
        for r in rows:
            con.execute(
                "INSERT INTO sessions (id,start_ts,end_ts,kwh_charged,cost_eur,location,charger_type,vehicle_id) VALUES (?,?,?,?,?,?,?,?)",
                r
            )
        con.commit(); con.close()

    def test_all_returns_all(self, tmp_path):
        import export_excel as _xl
        db = tmp_path / "sessions.db"
        self._populate(str(db))
        old = _xl.DB_PATH
        try:
            _xl.DB_PATH = db
            rows = _xl.fetch_sessions(2026, 5, "all")
            assert len(rows) == 3
        finally:
            _xl.DB_PATH = old

    def test_home_filter(self, tmp_path):
        import export_excel as _xl
        db = tmp_path / "sessions.db"
        self._populate(str(db))
        old = _xl.DB_PATH
        try:
            _xl.DB_PATH = db
            rows = _xl.fetch_sessions(2026, 5, "home")
            assert len(rows) == 1
            assert rows[0]["location"] == "home"
        finally:
            _xl.DB_PATH = old

    def test_extern_filter(self, tmp_path):
        import export_excel as _xl
        db = tmp_path / "sessions.db"
        self._populate(str(db))
        old = _xl.DB_PATH
        try:
            _xl.DB_PATH = db
            rows = _xl.fetch_sessions(2026, 5, "extern")
            assert len(rows) == 1
            assert rows[0]["location"] == "extern"
        finally:
            _xl.DB_PATH = old

    def test_different_month_excluded(self, tmp_path):
        import export_excel as _xl
        db = tmp_path / "sessions.db"
        self._populate(str(db))
        old = _xl.DB_PATH
        try:
            _xl.DB_PATH = db
            rows = _xl.fetch_sessions(2026, 6, "all")
            assert len(rows) == 0
        finally:
            _xl.DB_PATH = old

    def test_missing_db_returns_empty(self, tmp_path):
        import export_excel as _xl
        old = _xl.DB_PATH
        try:
            _xl.DB_PATH = tmp_path / "nonexistent.db"
            rows = _xl.fetch_sessions(2026, 5, "all")
            assert rows == []
        finally:
            _xl.DB_PATH = old
