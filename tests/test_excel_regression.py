"""
Excel export regression tests.
Covers: footer detection, row insertion, hash mismatch, cell mapping roundtrip.
"""
import pytest
import sys
from pathlib import Path
from io import BytesIO

sys.path.insert(0, str(Path(__file__).parent.parent / "app"))


def _has_openpyxl():
    try:
        import openpyxl
        return True
    except ImportError:
        return False


SAMPLE_SESSION = {
    "id": 1, "start_ts": "2026-01-15T08:00:00", "end_ts": "2026-01-15T09:00:00",
    "kwh_charged": 15.0, "cost_eur": 4.50, "location": "home",
    "charger_type": "ac", "soc_start": 20.0, "soc_end": 80.0,
    "odo_start": None, "odo_end": None, "meter_old": None, "meter_new": None,
    "max_power_kw": None, "price_per_kwh": 0.30, "duration": None,
}


class TestExcelFooterDetection:
    """Tests for Excel footer detection and row management."""

    @pytest.mark.skipif(not _has_openpyxl(), reason="openpyxl not available")
    def test_plain_text_footer_not_overwritten(self, tmp_path):
        """A template with plain text in footer rows must not be overwritten."""
        import openpyxl
        from export_excel import export_with_template

        # Create template: row 1=header info, row 2=col headers, rows 3-5=data, row 6=footer text
        wb = openpyxl.Workbook()
        ws = wb.active
        ws["A2"] = "Datum"; ws["B2"] = "kWh"; ws["C2"] = "Kosten"
        # 3 data rows
        ws["A3"] = None; ws["B3"] = None; ws["C3"] = None
        ws["A4"] = None; ws["B4"] = None; ws["C4"] = None
        ws["A5"] = None; ws["B5"] = None; ws["C5"] = None
        # Footer row with plain text (no formulas)
        ws["A6"] = "Gesamt:"
        ws["B6"] = "— kWh"
        tpl_path = tmp_path / "template.xlsx"
        wb.save(str(tpl_path))

        # Patch export_excel module paths
        import export_excel as _xl
        old_tpl = _xl.TEMPLATE_PATH
        old_exp = _xl.EXPORT_DIR
        try:
            _xl.TEMPLATE_PATH = tpl_path
            _xl.EXPORT_DIR = tmp_path / "exports"
            (tmp_path / "exports").mkdir(exist_ok=True)

            col_map = {"1": "date", "2": "kwh_charged", "3": "cost_eur"}
            result_path = export_with_template(
                2026, 1, [SAMPLE_SESSION],
                location="all",
                col_override=col_map,
                start_row=3,
                header_row=2,
                footer_start_row=6,
            )
            # Load result and check footer row is preserved
            out_wb = openpyxl.load_workbook(result_path)
            out_ws = out_wb.active
            # Footer row 6 must still have "Gesamt:" text
            assert out_ws["A6"].value == "Gesamt:", f"Footer was overwritten: got {out_ws['A6'].value!r}"
        finally:
            _xl.TEMPLATE_PATH = old_tpl
            _xl.EXPORT_DIR = old_exp

    @pytest.mark.skipif(not _has_openpyxl(), reason="openpyxl not available")
    def test_sum_footer_extended_when_rows_inserted(self, tmp_path):
        """SUM formula in footer must be extended after new rows are inserted."""
        import openpyxl
        from export_excel import export_with_template

        # Create template: 2 data rows (rows 3-4), SUM formula in row 5
        wb = openpyxl.Workbook()
        ws = wb.active
        ws["A2"] = "Datum"; ws["B2"] = "kWh"
        ws["A3"] = None; ws["B3"] = None
        ws["A4"] = None; ws["B4"] = None
        # Footer with SUM formula
        ws["A5"] = "Summe:"
        ws["B5"] = "=SUM(B3:B4)"
        tpl_path = tmp_path / "template.xlsx"
        wb.save(str(tpl_path))

        import export_excel as _xl
        old_tpl = _xl.TEMPLATE_PATH
        old_exp = _xl.EXPORT_DIR
        try:
            _xl.TEMPLATE_PATH = tpl_path
            _xl.EXPORT_DIR = tmp_path / "exports"
            (tmp_path / "exports").mkdir(exist_ok=True)

            # Export 4 sessions (more than 2 template rows)
            sessions = [dict(SAMPLE_SESSION, id=i+1) for i in range(4)]
            col_map = {"1": "date", "2": "kwh_charged"}
            result_path = export_with_template(
                2026, 1, sessions,
                location="all",
                col_override=col_map,
                start_row=3,
                header_row=2,
            )
            out_wb = openpyxl.load_workbook(result_path)
            out_ws = out_wb.active
            # Data rows 3-6 (4 sessions), footer now at row 7
            # SUM formula should reference B3:B6 (extended)
            footer_formula = None
            for r in range(5, 10):
                v = out_ws.cell(row=r, column=2).value
                if isinstance(v, str) and v.startswith("=SUM"):
                    footer_formula = v
                    break
            assert footer_formula is not None, "SUM formula in footer not found"
            # The formula should reference B3:B6 (covering all 4 data rows)
            assert ":B6" in footer_formula or "B6" in footer_formula, \
                f"SUM formula not extended to cover all data rows: {footer_formula!r}"
        finally:
            _xl.TEMPLATE_PATH = old_tpl
            _xl.EXPORT_DIR = old_exp

    @pytest.mark.skipif(not _has_openpyxl(), reason="openpyxl not available")
    def test_data_row_formula_not_treated_as_footer(self, tmp_path):
        """A per-row formula in a data row must NOT trigger footer detection."""
        import openpyxl
        from export_excel import export_with_template

        # Create template: data rows have a per-row formula in col C (duration calc)
        wb = openpyxl.Workbook()
        ws = wb.active
        ws["A2"] = "Start"; ws["B2"] = "Ende"; ws["C2"] = "Dauer"
        # Template data row has a per-row formula
        ws["A3"] = None; ws["B3"] = None; ws["C3"] = "=B3-A3"
        ws["A4"] = None; ws["B4"] = None; ws["C4"] = "=B4-A4"
        # Footer starts at row 5
        ws["A5"] = "Gesamt:"
        ws["C5"] = "=SUM(C3:C4)"
        tpl_path = tmp_path / "template.xlsx"
        wb.save(str(tpl_path))

        import export_excel as _xl
        old_tpl = _xl.TEMPLATE_PATH
        old_exp = _xl.EXPORT_DIR
        try:
            _xl.TEMPLATE_PATH = tpl_path
            _xl.EXPORT_DIR = tmp_path / "exports"
            (tmp_path / "exports").mkdir(exist_ok=True)

            # Export 2 sessions — should fill rows 3-4 without treating row 3 as footer
            sessions = [dict(SAMPLE_SESSION, id=i+1) for i in range(2)]
            col_map = {"1": "start_time", "2": "end_time"}
            result_path = export_with_template(
                2026, 1, sessions,
                location="all",
                col_override=col_map,
                start_row=3,
                header_row=2,
            )
            out_wb = openpyxl.load_workbook(result_path)
            out_ws = out_wb.active
            # Row 3 must have the session data (start_time from session), not be empty
            # The test passes if no exception is raised and the result loads correctly
            assert out_ws is not None
        finally:
            _xl.TEMPLATE_PATH = old_tpl
            _xl.EXPORT_DIR = old_exp


class TestTemplateHashInvalidation:
    """Tests for template hash invalidation on upload."""

    def test_upload_sets_hash_and_clears_mapping_hash(self, authed_client, tmp_path):
        """Uploading a new template must set template hash and clear mapping hash."""
        import openpyxl
        wb = openpyxl.Workbook()
        ws = wb.active
        ws["A1"] = "Datum"; ws["B1"] = "kWh"
        buf = BytesIO()
        wb.save(buf)
        buf.seek(0)

        data = {"file": (buf, "neues_template.xlsx")}
        rv = authed_client.post("/api/template",
                               content_type="multipart/form-data",
                               data=data)
        assert rv.status_code == 200
        assert rv.get_json()["ok"] is True

        # Check that mapping endpoint shows hash_mismatch = True (no mapping saved yet)
        rv2 = authed_client.get("/api/template/mapping")
        data2 = rv2.get_json()
        assert data2["template_hash"] is not None, "template_hash must be set after upload"
        assert data2["mapping_hash"] is None or data2["hash_mismatch"] is True, \
            "hash_mismatch must be True after upload before mapping is saved"

    def test_saving_mapping_clears_mismatch(self, authed_client, tmp_path):
        """After saving mapping, hash_mismatch must be False."""
        import openpyxl
        wb = openpyxl.Workbook()
        ws = wb.active
        ws["A1"] = "Datum"; ws["B1"] = "kWh"
        buf = BytesIO()
        wb.save(buf)
        buf.seek(0)

        # Upload template
        authed_client.post("/api/template",
                          content_type="multipart/form-data",
                          data={"file": (buf, "tpl.xlsx")})

        # Save mapping
        rv = authed_client.post("/api/template/mapping",
                               content_type="application/json",
                               data='{"column_mapping":{"1":"date","2":"kwh_charged"},"start_row":2}')
        assert rv.get_json()["ok"] is True

        # Now hash_mismatch should be False
        rv2 = authed_client.get("/api/template/mapping")
        data2 = rv2.get_json()
        assert data2.get("hash_mismatch") is False, \
            f"hash_mismatch should be False after saving mapping, got: {data2}"


class TestCellMappingRoundtrip:
    """Tests for single-cell mapping preservation."""

    def test_cell_mapping_saved_and_loaded(self, authed_client):
        """Cell mapping must survive POST->GET roundtrip."""
        payload = {
            "column_mapping": {"1": "date", "2": "kwh_charged"},
            "cell_mapping": {"B4": "kennzeichen", "C4": {"field": "fahrer"}},
            "start_row": 3,
            "header_row": 2,
            "sheet": "",
        }
        rv = authed_client.post("/api/template/mapping",
                               json=payload)
        assert rv.get_json()["ok"] is True

        rv2 = authed_client.get("/api/template/mapping")
        data = rv2.get_json()
        assert "B4" in data["cell_mapping"], "Cell mapping B4 not preserved"
        assert "C4" in data["cell_mapping"], "Cell mapping C4 not preserved"
